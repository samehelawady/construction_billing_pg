from django.contrib import admin
from django.contrib.admin import AdminSite, sites as admin_sites
from django.forms import TextInput
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from .models import (
    Client, Project, BOQItem, Invoice, InvoiceItem, CompanyProfile,
    ExpenseCategory, SubExpense, Expense,
    Employee, EmployeeTransfer, PayrollRecord, PayrollCostCenter, PayrollAllocation,
    PricingProject, PricingBOQItem, VariationOrder,
    Supplier, SupplierInvoice, SupplierPayment,
)
from django.urls import reverse
from decimal import Decimal
from django.db.models import Sum, Count, Q
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.urls import path
from datetime import date, timedelta
import calendar
from django.contrib import messages
from django.shortcuts import redirect
from django import forms
from .utils import money
from django.contrib.humanize.templatetags.humanize import intcomma
from django.utils.timezone import make_aware
from datetime import datetime
import json
from django_admin_inline_paginator.admin import TabularInlinePaginated
from django.forms.models import BaseInlineFormSet


# =============================================================================
# REPORTS DASHBOARD (Proxy Model for Reports Hub)
# =============================================================================

class LimitedInlineFormSet(BaseInlineFormSet):
    """Limits inline forms to max_num most recent items without slicing the queryset."""

    def get_queryset(self):
        qs = super().get_queryset()
        if self.max_num and qs.count() > self.max_num:
            recent_ids = list(qs.order_by('-id').values_list('pk', flat=True)[:self.max_num])
            return qs.filter(pk__in=recent_ids)
        return qs


class ReportsDashboard(CompanyProfile):
    """Proxy model to create a Reports hub in the admin sidebar."""

    class Meta:
        proxy = True
        verbose_name = "📊 Reports Dashboard"
        verbose_name_plural = "📊 Reports & Analytics"


# =============================================================================
# PROFESSIONAL BASE MIXIN WITH SHARED REPORT STYLING
# =============================================================================

class ProfessionalReportMixin:
    """
    Provides consistent, professional HTML report styling across all admin views.
    All reports share the same CSS framework for visual consistency.
    """

    # Shared CSS framework - modern, clean, professional
    SHARED_CSS = """
    <style>
        @page { size: A4 portrait; margin: 10mm; }
        * { box-sizing: border-box; margin: 0; padding: 0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }
        body { font-family: "Segoe UI", system-ui, -apple-system, sans-serif; font-size: 9.5px; color: #212529; line-height: 1.5; background: #fff; }
        .report-container { max-width: 100%; padding: 8mm; }
        .page { page-break-after: always; break-after: page; }
        .page:last-child { page-break-after: auto; }
        .report-header { text-align: center; margin-bottom: 20px; padding-bottom: 15px; border-bottom: 3px solid #1a237e; }
        .report-header .logo { max-height: 120px; max-width: 240px; object-fit: contain; margin-bottom: 10px; }
        .report-title { font-size: 18px; font-weight: 700; color: #1a237e; letter-spacing: 0.5px; text-transform: uppercase; }
        .report-subtitle { font-size: 11px; color: #6c757d; margin-top: 4px; }
        .card { background: #ffffff; border: 1px solid #dee2e6; border-radius: 8px; padding: 16px; margin-bottom: 16px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); }
        .card-header { font-size: 12px; font-weight: 700; color: #1a237e; margin-bottom: 12px; padding-bottom: 8px; border-bottom: 2px solid #1a237e; display: flex; align-items: center; gap: 8px; }
        .card-header .icon { font-size: 14px; }
        .meta-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px; margin-bottom: 20px; }
        .meta-item { background: #f8f9fa; padding: 10px 14px; border-radius: 4px; border-left: 3px solid #1a237e; }
        .meta-label { font-size: 8px; text-transform: uppercase; color: #6c757d; letter-spacing: 0.5px; margin-bottom: 2px; }
        .meta-value { font-size: 12px; font-weight: 600; color: #212529; }
        .data-table { width: 100%; border-collapse: separate; border-spacing: 0; font-size: 9px; margin-top: 8px; }
        .data-table thead th { background: #1a237e; color: white; padding: 8px 6px; font-weight: 600; text-align: left; font-size: 8px; text-transform: uppercase; letter-spacing: 0.3px; position: sticky; top: 0; }
        .data-table thead th:first-child { border-radius: 4px 0 0 0; }
        .data-table thead th:last-child { border-radius: 0 4px 0 0; }
        .data-table td { border-bottom: 1px solid #dee2e6; padding: 6px; vertical-align: top; }
        .data-table tr:nth-child(even) { background: #f8f9fa; }
        .data-table tr:hover { background: #e3f2fd; }
        .data-table .num { text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; }
        .data-table .text { text-align: left; }
        .data-table .center { text-align: center; }
        .total-row td { background: #e8eaf6 !important; font-weight: 700; border-top: 2px solid #1a237e; border-bottom: 2px solid #1a237e; }
        .grand-total td { background: #1a237e !important; color: white !important; font-weight: 700; font-size: 11px; }
        .grand-total .num { color: white !important; }
        .badge { display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 8px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.3px; }
        .badge-success { background: #e8f5e9; color: #2e7d32; }
        .badge-warning { background: #fff3e0; color: #f57c00; }
        .badge-danger { background: #ffebee; color: #c62828; }
        .badge-info { background: #e3f2fd; color: #0277bd; }
        .badge-primary { background: #e8eaf6; color: #1a237e; }
        .progress-bar { width: 100%; height: 20px; background: #e0e0e0; border-radius: 10px; overflow: hidden; margin-top: 8px; }
        .progress-fill { height: 100%; background: linear-gradient(90deg, #3949ab, #00897b); border-radius: 10px; transition: width 0.5s ease; display: flex; align-items: center; justify-content: flex-end; padding-right: 8px; color: white; font-size: 8px; font-weight: 700; }
        .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
        .grid-3 { display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; }
        .grid-4 { display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; }
        .metric-card { text-align: center; padding: 16px; background: #ffffff; border-radius: 8px; border: 1px solid #dee2e6; box-shadow: 0 2px 12px rgba(0,0,0,0.08); }
        .metric-value { font-size: 20px; font-weight: 700; color: #1a237e; margin-bottom: 4px; }
        .metric-label { font-size: 8px; color: #6c757d; text-transform: uppercase; letter-spacing: 0.5px; }
        .metric-delta { font-size: 9px; margin-top: 4px; }
        .signature-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 40px; margin-top: 40px; padding-top: 20px; }
        .signature-block { text-align: center; }
        .signature-line { border-top: 1px solid #333; margin-top: 50px; padding-top: 8px; font-size: 9px; font-weight: 600; }
        .cover-page { display: flex; flex-direction: column; min-height: calc(100vh - 20mm); justify-content: space-between; }
        .cover-header { margin-bottom: 30px; }
        .cover-title { font-size: 14px; font-weight: 700; color: #1a237e; margin-bottom: 20px; }
        .cover-ref { margin-bottom: 20px; }
        .cover-ref-row { font-size: 10px; margin-bottom: 6px; display: flex; }
        .cover-ref-label { font-weight: 700; min-width: 80px; color: #1a237e; }
        .cover-body { font-size: 10px; line-height: 1.8; margin-bottom: 30px; }
        .cover-body p { margin-bottom: 12px; }
        .cover-closing { margin-top: auto; font-size: 10px; }
        .cover-signature-space { height: 60px; border-bottom: 1px solid #333; margin-bottom: 8px; }
        .cover-sig-name { font-weight: 700; font-size: 11px; }
        .cover-sig-title { color: #6c757d; font-size: 9px; margin-top: 2px; }
        @media print { .no-print { display: none !important; } .card { box-shadow: none; border: 1px solid #ddd; } }
        .text-right { text-align: right; }
        .text-center { text-align: center; }
        .text-left { text-align: left; }
        .font-bold { font-weight: 700; }
        .text-primary { color: #1a237e; }
        .text-success { color: #2e7d32; }
        .text-danger { color: #c62828; }
        .text-warning { color: #f57c00; }
        .text-muted { color: #6c757d; }
        .mb-1 { margin-bottom: 4px; }
        .mb-2 { margin-bottom: 8px; }
        .mb-3 { margin-bottom: 12px; }
        .mt-2 { margin-top: 8px; }
        .mt-3 { margin-top: 12px; }
        .p-2 { padding: 8px; }
        .border-top { border-top: 1px solid #dee2e6; padding-top: 8px; margin-top: 8px; }
    </style>
    """

    def _get_logo_html(self, logo_url, align='right'):
        """Generate consistent logo HTML."""
        if not logo_url:
            return ''
        return f'<div style="text-align:{align}; margin-bottom:10px;"><img src="{logo_url}" alt="Logo" style="max-height:120px; max-width:240px; object-fit:contain;"></div>'

    def _report_base_wrapper(self, title, subtitle, body_html, logo_url='', extra_css=''):
        """
        Wraps any report body in the professional shared CSS framework.
        All reports use this for consistency.
        """
        return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
{self.SHARED_CSS}
{extra_css}
</head>
<body>
    <div class="report-container">
        {self._get_logo_html(logo_url)}
        <div class="report-header">
            <div class="report-title">{title}</div>
            <div class="report-subtitle">{subtitle}</div>
        </div>
        {body_html}
    </div>
    <script>
        window.onload = function() {{ 
            setTimeout(function() {{ window.print(); }}, 400); 
        }};
    </script>
</body>
</html>"""

    def _build_meta_grid(self, items):
        """Build a meta information grid from dict items."""
        cells = ""
        for label, value in items.items():
            cells += f"""
            <div class="meta-item">
                <div class="meta-label">{label}</div>
                <div class="meta-value">{value}</div>
            </div>
            """
        return f'<div class="meta-grid">{cells}</div>'

    def _build_metric_cards(self, metrics):
        """Build metric cards row. metrics = [(label, value, color_class, delta)]"""
        cards = ""
        for label, value, color, delta in metrics:
            delta_html = f'<div class="metric-delta {delta[1]}">{delta[0]}</div>' if delta else ''
            cards += f"""
            <div class="metric-card">
                <div class="metric-value {color}">{value}</div>
                <div class="metric-label">{label}</div>
                {delta_html}
            </div>
            """
        return f'<div class="grid-4">{cards}</div>'
# =============================================================================
# COMPANY SCOPING MIXIN
# =============================================================================

class CompanyScopedAdminMixin:
    company_field_path = None

    def get_active_company(self, request):
        return CompanyProfile.get_active(request)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        company = self.get_active_company(request)
        if not company:
            return qs  # Or return qs.none() if you want strict mode

        if self.company_field_path:
            # Filter applies to ALL users, including superusers
            null_filter = {self.company_field_path + '__isnull': True}
            qs = qs.filter(
                Q(**{self.company_field_path: company}) | Q(**null_filter)
            )
        return qs

    def get_object_or_404_scoped(self, request, model, **kwargs):
        """Get object through the scoped queryset, with fallback."""
        qs = self.get_queryset(request).filter(**kwargs)
        try:
            return qs.get()
        except model.DoesNotExist:
            # Fallback for transition period
            try:
                return model.objects.get(**kwargs)
            except model.DoesNotExist:
                from django.http import Http404
                raise Http404(f"No {model._meta.verbose_name} matches the given query.")

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        company = self.get_active_company(request)
        if company:
            if db_field.name == 'company':
                kwargs['queryset'] = CompanyProfile.objects.filter(is_active=True)
            elif db_field.name == 'client':
                kwargs['queryset'] = Client.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
            elif db_field.name == 'project':
                kwargs['queryset'] = Project.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
            elif db_field.name == 'employee':
                kwargs['queryset'] = Employee.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
            elif db_field.name == 'pricing_project':
                kwargs['queryset'] = PricingProject.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
            elif db_field.name == 'reference_boq_item':
                kwargs['queryset'] = BOQItem.objects.filter(
                    Q(project__company=company) | Q(project__company__isnull=True)
                )
            elif db_field.name == 'reference_projects':
                kwargs['queryset'] = Project.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
            elif db_field.name == 'boq_item':
                kwargs['queryset'] = BOQItem.objects.filter(
                    Q(project__company=company) | Q(project__company__isnull=True)
                )
            elif db_field.name == 'category':
                kwargs['queryset'] = ExpenseCategory.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
            elif db_field.name == 'sub_category':
                kwargs['queryset'] = SubExpense.objects.filter(
                    Q(parent__company=company) | Q(parent__company__isnull=True)
                )
            elif db_field.name == 'payroll_record':
                kwargs['queryset'] = PayrollRecord.objects.filter(
                    Q(employee__company=company) | Q(employee__company__isnull=True)
                )
            elif db_field.name == 'to_project':
                kwargs['queryset'] = Project.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
            elif db_field.name == 'invoice':
                kwargs['queryset'] = Invoice.objects.filter(
                    Q(project__company=company) | Q(project__company__isnull=True)
                )
            elif db_field.name == 'supplier':
                kwargs['queryset'] = Supplier.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def formfield_for_manytomany(self, db_field, request, **kwargs):
        company = self.get_active_company(request)
        if company:
            if db_field.name == 'reference_projects':
                kwargs['queryset'] = Project.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
        return super().formfield_for_manytomany(db_field, request, **kwargs)

    def save_model(self, request, obj, form, change):
        company = self.get_active_company(request)
        if company and hasattr(obj, 'company') and not obj.company_id:
            if hasattr(obj, 'project') and obj.project and hasattr(obj.project, 'company'):
                obj.company = obj.project.company
            else:
                obj.company = company
        super().save_model(request, obj, form, change)


# =============================================================================
# DYNAMIC ADMIN SITE WITH COMPANY CONTEXT
# =============================================================================

class EnhancedAdminSite(AdminSite):
    """
    Professional admin site with:
    - Organized sidebar groups
    - Reports hub
    - Company context in header
    - Improved dashboard
    """
    site_header = "Billing & Project Management"
    site_title = "BPM Portal"
    index_title = "Dashboard"
    index_template = "admin/enhanced_index.html"
    app_index_template = "admin/enhanced_app_index.html"

    def each_context(self, request):
        context = super().each_context(request)
        company = CompanyProfile.get_active(request)

        if company:
            context['site_header'] = company.company_name
            context['site_title'] = f"{company.company_name} - BPM"
            context['index_title'] = f"{company.company_name} Dashboard"

        if request.user.is_authenticated:
            context['available_companies'] = CompanyProfile.objects.filter(is_active=True)
            context['active_company'] = company

            # Add organized navigation groups
            context['nav_groups'] = self._build_navigation(request)

        return context

    def _count_completed_projects(self, company):
        """Count projects with final paid tax invoice and all retention recovered."""
        if not company:
            return 0
        completed = 0
        for proj in Project.objects.filter(company=company):
            final_inv = Invoice.objects.filter(
                project=proj, inv_type='T', is_final_invoice=True, status='Paid'
            ).first()
            if not final_inv:
                continue
            latest_inv = Invoice.objects.filter(
                project=proj, inv_type='T'
            ).exclude(is_advance_invoice=True).order_by('-inv_number').first()
            if latest_inv:
                ret_total = latest_inv.cumulative_retention_total
                ret_recovered = latest_inv.cumulative_retention_a_recovered + latest_inv.cumulative_retention_b_recovered
                if ret_total <= ret_recovered:
                    completed += 1
        return completed

    def _build_navigation(self, request):
        """Build organized navigation groups for the sidebar."""
        company = CompanyProfile.get_active(request)
        # Counts for badges
        client_count = Client.objects.filter(company=company).count() if company else 0
        # FIX: Active projects based on retention status
        active_project_count = 0
        if company:
            for proj in Project.objects.filter(company=company):
                latest_inv = Invoice.objects.filter(
                    project=proj, is_advance_invoice=False
                ).order_by('-inv_number').first()
                if latest_inv:
                    ret_total = latest_inv.cumulative_retention_total
                    ret_recovered = latest_inv.cumulative_retention_a_recovered + latest_inv.cumulative_retention_b_recovered
                    if ret_total > ret_recovered:
                        active_project_count += 1
                else:
                    active_project_count += 1

        # FIX: Count unpaid/draft/approved invoices for badge
        invoice_alert_count = 0
        if company:
            invoice_alert_count = Invoice.objects.filter(
                project__company=company
            ).exclude(
                Q(status='Paid') & Q(inv_type='T')
            ).exclude(
                Q(inv_type='P', status='Paid')
            ).count()
        invoice_count = Invoice.objects.filter(project__company=company).count() if company else 0
        employee_count = Employee.objects.filter(company=company, is_active=True).count() if company else 0
        completed_projects = self._count_completed_projects(company)
        return [
            {
                'name': 'Core Business',
                'icon': '💼',
                'items': [
                    {'name': 'Clients', 'url': reverse('admin:billing_client_changelist'), 'count': client_count,
                     'badge_color': 'primary'},
                    {'name': 'Projects', 'url': reverse('admin:billing_project_changelist'), 'count': active_project_count,
                     'badge_color': 'success'},
                    {'name': 'BOQ Items', 'url': reverse('admin:billing_boqitem_changelist'), 'count': None},
                ]
            },
            {
                'name': 'Financial',
                'icon': '💰',
                'items': [
                    {'name': 'Invoices', 'url': reverse('admin:billing_invoice_changelist'), 'count': invoice_alert_count,
                     'badge_color': 'warning'},
                    {'name': 'Expenses', 'url': reverse('admin:billing_expense_changelist'), 'count': None},
                    {'name': 'Expense Categories', 'url': reverse('admin:billing_expensecategory_changelist'),
                     'count': None},
                ]
            },
            {
                'name': 'Workforce',
                'icon': '👷',
                'items': [
                    {'name': 'Employees', 'url': reverse('admin:billing_employee_changelist'), 'count': employee_count,
                     'badge_color': 'info'},
                    {'name': 'Payroll Records', 'url': reverse('admin:billing_payrollrecord_changelist'),
                     'count': None},
                    {'name': 'Transfers', 'url': reverse('admin:billing_employeetransfer_changelist'), 'count': None},
                ]
            },
            {
                'name': 'Pricing & Setup',
                'icon': '📋',
                'items': [
                    {'name': 'Pricing Projects', 'url': reverse('admin:billing_pricingproject_changelist'),
                     'count': None},
                    {'name': 'Company Profiles', 'url': reverse('admin:billing_companyprofile_changelist'),
                     'count': None},
                ]
            },
            {
                'name': 'Reports & Analytics',
                'icon': '📊',
                'items': [
                    {'name': '📊 Reports Dashboard', 'url': reverse('admin:billing_reportsdashboard_changelist'),
                     'count': None, 'highlight': True},
                ]
            },
            {
                'name': 'Completed Projects', 'url': reverse('admin:billing_project_changelist') + '?completed=1',
                'count': completed_projects, 'badge_color': 'success'
            },
            {'name': 'Suppliers', 'url': reverse('admin:billing_supplier_changelist'), 'count': None},
            {'name': 'Supplier Invoices', 'url': reverse('admin:billing_supplierinvoice_changelist'), 'count': None},
        ]

    def index(self, request, extra_context=None):
        if extra_context is None:
            extra_context = {}

        company = CompanyProfile.get_active(request)
        extra_context['active_company'] = company

        if company:
            extra_context['dashboard_metrics'] = self._build_dashboard_metrics(company)
            extra_context['recent_invoices'] = self._get_recent_invoices(company)
            extra_context['project_progress'] = self._get_project_progress(company)
            # >>> ADD THESE TWO LINES <<<
            extra_context['project_charts'] = self._build_project_charts(company)
            extra_context['overall_chart'] = self._build_overall_chart(company)

        return super().index(request, extra_context)

    def _build_dashboard_metrics(self, company):
        """Build key metrics for the dashboard using only DB fields."""
        from django.db.models import Sum, Q
        from decimal import Decimal


        total_projects = Project.objects.filter(company=company).count()
        # FIX: Active projects based on retention status, not BOQ completion
        active_projects = 0
        for proj in Project.objects.filter(company=company):
            latest_inv = Invoice.objects.filter(
                project=proj, is_advance_invoice=False
            ).order_by('-inv_number').first()
            if latest_inv:
                ret_total = latest_inv.cumulative_retention_total
                ret_recovered = latest_inv.cumulative_retention_a_recovered + latest_inv.cumulative_retention_b_recovered
                if ret_total > ret_recovered:
                    active_projects += 1
            else:
                # No invoices yet = active project
                active_projects += 1

        # current_certified_net_before_vat is a PROPERTY, not a DB field.
        # We must iterate invoices and call the property in Python.
        total_invoiced = Decimal("0")
        for inv in Invoice.objects.filter(
                project__company=company,
                inv_type='T'
        ).exclude(is_advance_invoice=True):
            total_invoiced += inv.current_certified_net_before_vat

        # Include variation amounts in total invoiced
        for proj in Project.objects.filter(company=company):
            total_invoiced += getattr(proj, 'variation_amount', Decimal("0"))

        # total_after_vat is also a PROPERTY — iterate in Python
        outstanding = Decimal("0")
        for inv in Invoice.objects.filter(
                project__company=company
        ).exclude(status='Paid'):
            outstanding += inv.total_after_vat

        total_employees = Employee.objects.filter(company=company, is_active=True).count()

        # Completed projects: final tax invoice paid + all retention recovered
        completed_projects = 0
        for proj in Project.objects.filter(company=company):
            # Check if there's a final tax invoice that is paid
            final_inv = Invoice.objects.filter(
                project=proj,
                inv_type='T',
                is_final_invoice=True,
                status='Paid'
            ).first()

            if final_inv:
                # Check all retention is recovered
                latest_inv = Invoice.objects.filter(
                    project=proj,
                    inv_type='T'
                ).exclude(is_advance_invoice=True).order_by('-inv_number').first()

                if latest_inv:
                    ret_total = latest_inv.cumulative_retention_total
                    ret_recovered = latest_inv.cumulative_retention_a_recovered + latest_inv.cumulative_retention_b_recovered
                    if ret_total <= ret_recovered:  # All retention recovered
                        completed_projects += 1

        return [
            {'label': 'Total Projects', 'value': intcomma(total_projects), 'icon': '🏗️', 'color': 'primary'},
            {'label': 'Active Projects', 'value': intcomma(active_projects), 'icon': '📁', 'color': 'success'},
            {'label': 'Total Invoiced', 'value': f"AED {intcomma(int(total_invoiced))}", 'icon': '💵',
             'color': 'warning'},
            {'label': 'Outstanding', 'value': f"AED {intcomma(int(outstanding))}", 'icon': '⚠️', 'color': 'danger'},
            {'label': 'Active Staff', 'value': intcomma(total_employees), 'icon': '👥', 'color': 'info'},
            {'label': 'Completed Projects', 'value': intcomma(completed_projects), 'icon': '✅', 'color': 'success'},
        ]

    def _get_recent_invoices(self, company, limit=10):
        """FIX: Get unpaid/draft/approved invoices for dashboard - not just recent."""
        return Invoice.objects.filter(
            project__company=company
        ).exclude(
            Q(status='Paid') & Q(inv_type='T')  # Exclude paid tax invoices
        ).exclude(
            Q(inv_type='P', status='Paid')  # Exclude paid proforma
        ).select_related('project', 'project__client').order_by('-date')[:limit]

    def _get_project_progress(self, company):
        """FIX: Show ALL active projects' progress, not just a few."""
        progress_data = []

        for proj in Project.objects.filter(company=company).order_by('project_id_code'):
            latest_inv = Invoice.objects.filter(
                project=proj, is_advance_invoice=False
            ).order_by('-inv_number').first()

            work_done = latest_inv.certified_work_done if latest_inv else Decimal("0")
            # Include variations in the contract value for progress % calculation
            variation_amount = getattr(proj, 'variation_amount', Decimal("0"))
            amended_po = proj.po_amount + variation_amount
            pct = (work_done / amended_po * 100) if amended_po > 0 else 0

            # Check if project is active (retention not fully claimed)
            is_active = True
            if latest_inv:
                ret_total = latest_inv.cumulative_retention_total
                ret_recovered = latest_inv.cumulative_retention_a_recovered + latest_inv.cumulative_retention_b_recovered
                is_active = ret_total > ret_recovered

            progress_data.append({
                'project': proj,
                'progress': pct,
                'work_done': work_done,
                'po_amount': amended_po,
                'is_active': is_active,
            })

        return progress_data

    def _build_project_charts(self, company):
        """Build chart data for each project (revenue, direct expenses, manpower)."""
        from django.db.models import Sum
        from decimal import Decimal
        import calendar

        charts = []
        projects = list(Project.objects.filter(company=company).order_by('project_id_code'))[:12]

        for proj in projects:
            latest_inv = Invoice.objects.filter(
                project=proj, is_advance_invoice=False
            ).order_by('-inv_number').first()

            # Revenue includes certified work + variations
            base_revenue = latest_inv.certified_work_done if latest_inv else Decimal("0")
            variation_amount = getattr(proj, 'variation_amount', Decimal("0"))
            revenue = base_revenue + variation_amount

            direct_expenses = Expense.objects.filter(
                project=proj
            ).aggregate(total=Sum('amount'))['total'] or Decimal("0")

            # Primary: allocated payroll totals (DB field)
            manpower = PayrollAllocation.objects.filter(
                project=proj
            ).aggregate(total=Sum('total_allocated'))['total'] or Decimal("0")

            # Fallback: calculate from PayrollCostCenter + PayrollRecord in Python
            if manpower == 0:
                cost_centers = PayrollCostCenter.objects.filter(
                    project=proj
                ).select_related('payroll_record__employee')

                manpower = Decimal("0")
                for cc in cost_centers:
                    pr = cc.payroll_record
                    emp = pr.employee

                    if cc.days_count and cc.days_count > 0:
                        # Daily rate fallback: total_salary / 30
                        daily_rate = getattr(emp, 'daily_rate', None) or (emp.total_salary / Decimal("30"))
                        salary_portion = daily_rate * cc.days_count

                        annual_admin = (
                                emp.annual_benefits + emp.annual_eid_cost +
                                emp.annual_visa_cost + emp.annual_ticket_cost
                        )
                        admin_portion = (annual_admin / Decimal("312")) * cc.days_count

                        manpower += salary_portion + admin_portion

            charts.append({
                'id': proj.id,
                'code': proj.project_id_code,
                'name': proj.project_name,
                'revenue': float(revenue),
                'direct_expenses': float(direct_expenses),
                'manpower': float(manpower),
            })
        return charts

    def _build_overall_chart(self, company):
        """Build aggregated chart data for all projects (revenue, expenses, profit/loss)."""
        from django.db.models import Sum
        from decimal import Decimal
        import calendar

        labels = []
        revenues = []
        expenses = []
        profits = []

        for proj in Project.objects.filter(company=company).order_by('project_id_code'):
            latest_inv = Invoice.objects.filter(
                project=proj, is_advance_invoice=False
            ).order_by('-inv_number').first()

            # Revenue includes certified work + variations
            base_revenue = latest_inv.certified_work_done if latest_inv else Decimal("0")
            variation_amount = getattr(proj, 'variation_amount', Decimal("0"))
            revenue = base_revenue + variation_amount

            direct_expenses = Expense.objects.filter(
                project=proj
            ).aggregate(total=Sum('amount'))['total'] or Decimal("0")

            # PayrollAllocation first (DB field)
            manpower = PayrollAllocation.objects.filter(
                project=proj
            ).aggregate(total=Sum('total_allocated'))['total'] or Decimal("0")

            # Fallback calculation from cost centers
            if manpower == 0:
                cost_centers = PayrollCostCenter.objects.filter(
                    project=proj
                ).select_related('payroll_record__employee')

                manpower = Decimal("0")
                for cc in cost_centers:
                    pr = cc.payroll_record
                    emp = pr.employee

                    if cc.days_count and cc.days_count > 0:
                        daily_rate = getattr(emp, 'daily_rate', None) or (emp.total_salary / Decimal("30"))
                        salary_portion = daily_rate * cc.days_count

                        annual_admin = (
                                emp.annual_benefits + emp.annual_eid_cost +
                                emp.annual_visa_cost + emp.annual_ticket_cost
                        )
                        admin_portion = (annual_admin / Decimal("312")) * cc.days_count

                        manpower += salary_portion + admin_portion

            total_expenses = direct_expenses + manpower
            profit = revenue - total_expenses

            labels.append(proj.project_id_code)
            revenues.append(float(revenue))
            expenses.append(float(total_expenses))
            profits.append(float(profit))

        return {
            'labels_json': json.dumps(labels),
            'revenues_json': json.dumps(revenues),
            'expenses_json': json.dumps(expenses),
            'profits_json': json.dumps(profits),
        }
# Replace the default admin site
admin.site = EnhancedAdminSite(name="admin")
admin_sites.site = admin.site


# =============================================================================
# REPORTS DASHBOARD ADMIN (The Reports Hub)
# =============================================================================

@admin.register(ReportsDashboard)
class ReportsDashboardAdmin(ProfessionalReportMixin,CompanyScopedAdminMixin, admin.ModelAdmin):

    #In case I need it in the futrue
    #def get_active_company(self, request):
        #return CompanyProfile.get_active(request)
    """
    Central Reports Hub - provides access to all reports from one place.
    This is a proxy model admin that serves as a navigation hub.
    """
    change_list_template = "admin/reports_dashboard.html"

    # Remove add/change/delete permissions since this is a hub
    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        """
        Build the reports dashboard with categorized report cards.
        """
        company = self.get_active_company(request)

        # In ReportsDashboardAdmin.changelist_view(), replace the report_categories list with this:

        report_categories = [
            {
                'name': 'Client Reports',
                'icon': '👤',
                'description': 'Statements, outstanding balances, deductions, and project progress by client',
                'color': '#1a237e',
                'reports': [
                    {
                        'name': 'Statement of Account',
                        'description': 'Full client ledger with invoice history, payments, and aging',
                        'url': reverse('admin:client_statement_list'),
                        'icon': '📄',
                        'badge': 'Financial',
                    },
                    {
                        'name': 'Outstanding Invoices',
                        'description': 'Unpaid invoices grouped by draft and approved status',
                        'url': reverse('admin:client_outstanding_list'),
                        'icon': '⚠️',
                        'badge': 'Aging',
                    },
                    {
                        'name': 'Deductions Report',
                        'description': 'Complete tracking of materials, retention, advance, and other deductions',
                        'url': reverse('admin:client_deductions_list'),
                        'icon': '📦',
                        'badge': 'Deductions',
                    },
                    {
                        'name': 'Project Progress',
                        'description': 'Visual progress cards with retention and advance status',
                        'url': reverse('admin:client_progress_list'),
                        'icon': '📈',
                        'badge': 'Analytics',
                    },
                    {
                        'name': 'Materials Supplied Report',
                        'description': 'Consolidated view of all materials supplied by all clients across all projects',
                        'url': reverse('admin:materials_supplied_report'),
                        'icon': '📦',
                        'badge': 'Materials',
                    },
                    {
                        'name': 'Retention Held Report',
                        'description': 'Consolidated view of all retention held for all clients across all projects',
                        'url': reverse('admin:retention_held_report'),
                        'icon': '🔒',
                        'badge': 'Retention',
                    },
                ]
            },
            {
                'name': 'Cash Flow',
                'icon': '💰',
                'description': 'Cash flow predictions, supplier aging, and liquidity forecasting',
                'color': '#1565c0',
                'reports': [
                    {
                        'name': 'Cash Flow Prediction',
                        'description': '13-week rolling cash flow with inflow from client invoices and outflow from supplier payments',
                        'url': reverse('admin:cash_flow_report'),
                        'icon': '📈',
                        'badge': 'Forecast',
                    },
                ]
            },
            {
                'name': 'Project Reports',
                'icon': '🏗️',
                'description': 'Analytics, cost breakdowns, deductions tracking, and profitability analysis',
                'color': '#00695c',
                'reports': [
                    {
                        'name': 'Project Analytics',
                        'description': 'Comprehensive project overview with invoices, BOQ, and cash flow',
                        'url': reverse('admin:project_analytics_list'),
                        'icon': '📊',
                        'badge': 'Overview',
                    },
                    {
                        'name': 'Cost & Profitability',
                        'description': 'BOQ-level cost breakdown with revenue vs expenses and P&L',
                        'url': reverse('admin:project_cost_profit_list'),
                        'icon': '💹',
                        'badge': 'P&L',
                    },
                    {
                        'name': 'Project Deductions',
                        'description': 'Detailed deduction tracking by invoice for a single project',
                        'url': reverse('admin:project_deductions_list'),
                        'icon': '📦',
                        'badge': 'Deductions',
                    },
                ]
            },
            {
                'name': 'Invoice Reports',
                'icon': '📑',
                'description': 'Printable invoices with cover letters and detail pages',
                'color': '#e65100',
                'reports': [
                    {
                        'name': 'Print Invoice',
                        'description': 'Professional 3-page invoice (Cover + Summary + Detail)',
                        'url': reverse('admin:invoice_print_list'),
                        'icon': '🖨️',
                        'badge': 'Print',
                    },
                ]
            },
            {
                'name': 'Payroll Reports',
                'icon': '💰',
                'description': 'Timesheets, labor costs, and payment method reports',
                'color': '#4a148c',
                'reports': [
                    {
                        'name': 'Worker Time Sheet',
                        'description': 'Daily attendance, OT, and payroll summary per employee',
                        'url': reverse('admin:payroll_timesheet_list'),
                        'icon': '⏱️',
                        'badge': 'Individual',
                    },
                    {
                        'name': 'Labor Cost Report',
                        'description': 'Monthly project labor cost breakdown by worker',
                        'url': reverse('admin:payroll_labor_cost_list'),
                        'icon': '👷',
                        'badge': 'Monthly',
                    },
                    {
                        'name': 'Staff Payroll Report',
                        'description': 'Office staff bank transfer payment summary',
                        'url': reverse('admin:payroll_staff_report_list'),
                        'icon': '🏢',
                        'badge': 'Bank',
                    },
                    {
                        'name': 'WPS Report',
                        'description': 'Site workers WPS agency payment report',
                        'url': reverse('admin:payroll_wps_report_list'),
                        'icon': '💳',
                        'badge': 'WPS',
                    },
                    {
                        'name': 'Cash Payroll Report',
                        'description': 'Cash payment workers summary with signatures',
                        'url': reverse('admin:payroll_cash_report_list'),
                        'icon': '💵',
                        'badge': 'Cash',
                    },
                ]
            },
        ]

        extra_context = extra_context or {}
        extra_context['report_categories'] = report_categories
        extra_context['company'] = company
        extra_context['title'] = "Reports & Analytics Dashboard"

        return super().changelist_view(request, extra_context)

    def materials_supplied_report(self, request):
        """
        Professional Materials Supplied Report - aggregates all materials received
        from all clients across all projects. Shows client-level summaries with
        project breakdowns.
        """
        company = self.get_active_company(request)
        logo_url = company.logo.url if company and company.logo else ''

        # Build query with company scoping
        invoice_qs = Invoice.objects.filter(
            inv_type='T',
            material_supplied_by_client__gt=0
        ).exclude(is_advance_invoice=True).select_related('project', 'project__client')

        if company:
            invoice_qs = invoice_qs.filter(project__company=company)

        # Get all clients with materials (distinct and ordered)
        client_ids = invoice_qs.values_list('project__client', flat=True).distinct()
        clients_with_materials = Client.objects.filter(id__in=client_ids).order_by('name')

        if company:
            clients_with_materials = clients_with_materials.filter(
                Q(company=company) | Q(company__isnull=True)
            )

        # Build client sections
        client_sections = ""
        grand_total_materials = Decimal("0")
        total_clients = 0
        total_invoices = 0
        total_projects = 0

        for client in clients_with_materials:
            # Get projects for this client
            project_qs = Project.objects.filter(client=client)
            if company:
                project_qs = project_qs.filter(Q(company=company) | Q(company__isnull=True))

            project_rows = ""
            client_total = Decimal("0")
            client_projects = 0
            client_invoices = 0

            for proj in project_qs:
                # Get invoices with materials for this project
                proj_invoices = invoice_qs.filter(project=proj).order_by('inv_number')

                if not proj_invoices.exists():
                    continue

                client_projects += 1
                proj_materials = Decimal("0")
                inv_rows = ""

                for inv in proj_invoices:
                    mat = inv.material_supplied_by_client or Decimal("0")
                    proj_materials += mat
                    client_invoices += 1
                    status_badge = '<span class="badge badge-success">Paid</span>' if inv.status == 'Paid' else '<span class="badge badge-warning">Draft</span>' if inv.status == 'Draft' else '<span class="badge badge-info">Approved</span>'

                    inv_rows += f"""
                    <tr>
                        <td class="center">{inv.inv_number}</td>
                        <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                        <td class="text"><b>{inv}</b></td>
                        <td class="center">{status_badge}</td>
                        <td class="num font-bold text-danger">({mat:,.2f})</td>
                        <td class="text">{inv.project.project_name}</td>
                    </tr>
                    """

                if proj_materials > 0:
                    client_total += proj_materials
                    project_rows += f"""
                    <div style="margin-bottom: 12px; border: 1px solid #ffcdd2; border-radius: 6px; overflow: hidden;">
                        <div style="background: #ffebee; padding: 8px 12px; font-size: 10px; font-weight: 700; color: #c62828;">
                            <span style="margin-right: 8px;">🏗️</span> {proj.project_id_code} — {proj.project_name}
                            <span style="float: right; color: #c62828;">Project Total: ({proj_materials:,.2f})</span>
                        </div>
                        <table class="data-table" style="margin: 0; font-size: 8.5px;">
                            <thead>
                                <tr>
                                    <th style="width: 8%;">Inv #</th>
                                    <th style="width: 12%;">Date</th>
                                    <th style="width: 25%;">Invoice Reference</th>
                                    <th style="width: 10%;">Status</th>
                                    <th style="width: 15%;" class="num">Amount</th>
                                    <th style="width: 30%;">Notes</th>
                                </tr>
                            </thead>
                            <tbody>{inv_rows}</tbody>
                            <tfoot>
                                <tr style="background: #ffebee; font-weight: bold;">
                                    <td colspan="4"><b>PROJECT SUBTOTAL</b></td>
                                    <td class="num text-danger"><b>({proj_materials:,.2f})</b></td>
                                    <td></td>
                                </tr>
                            </tfoot>
                        </table>
                    </div>
                    """

            if client_total > 0:
                total_clients += 1
                total_invoices += client_invoices
                total_projects += client_projects
                grand_total_materials += client_total

                client_sections += f"""
                <div class="card" style="page-break-inside: avoid; margin-bottom: 20px; border-left: 4px solid #c62828;">
                    <div class="card-header" style="display: flex; justify-content: space-between; align-items: center; color: #c62828; border-color: #ffcdd2;">
                        <span><span class="icon">👤</span> {client.name}</span>
                        <span class="badge badge-danger">Client Total: ({client_total:,.2f})</span>
                    </div>
                    <div style="padding: 12px;">
                        <div class="meta-grid" style="margin-bottom: 12px; grid-template-columns: repeat(4, 1fr);">
                            <div class="meta-item" style="border-left-color: #c62828;">
                                <div class="meta-label">TRN</div>
                                <div class="meta-value">{client.vat_number or 'N/A'}</div>
                            </div>
                            <div class="meta-item" style="border-left-color: #c62828;">
                                <div class="meta-label">Active Projects</div>
                                <div class="meta-value">{client_projects}</div>
                            </div>
                            <div class="meta-item" style="border-left-color: #c62828;">
                                <div class="meta-label">Invoices</div>
                                <div class="meta-value">{client_invoices}</div>
                            </div>
                            <div class="meta-item" style="border-left-color: #c62828;">
                                <div class="meta-label">Client Total Materials</div>
                                <div class="meta-value text-danger">({client_total:,.2f})</div>
                            </div>
                        </div>
                        {project_rows}
                    </div>
                </div>
                """

        # Build summary by client table
        summary_rows = ""
        for client in clients_with_materials:
            client_total = Decimal("0")
            proj_count = 0
            inv_count = 0

            project_qs = Project.objects.filter(client=client)
            if company:
                project_qs = project_qs.filter(Q(company=company) | Q(company__isnull=True))

            for proj in project_qs:
                proj_invoices = invoice_qs.filter(project=proj)
                proj_mat = sum(inv.material_supplied_by_client or Decimal("0") for inv in proj_invoices)
                if proj_mat > 0:
                    client_total += proj_mat
                    proj_count += 1
                    inv_count += proj_invoices.count()

            if client_total > 0:
                pct_of_grand = (client_total / grand_total_materials * 100) if grand_total_materials > 0 else Decimal(
                    "0")
                summary_rows += f"""
                <tr>
                    <td class="text"><b>{client.name}</b></td>
                    <td class="center">{client.vat_number or 'N/A'}</td>
                    <td class="center">{proj_count}</td>
                    <td class="center">{inv_count}</td>
                    <td class="num font-bold text-danger">({client_total:,.2f})</td>
                    <td class="center">
                        <div class="progress-bar" style="height: 16px; margin-top: 0;">
                            <div class="progress-fill" style="width: {min(float(pct_of_grand), 100)}%; background: linear-gradient(90deg, #c62828, #f57c00); font-size: 7px;">
                                {pct_of_grand:.1f}%
                            </div>
                        </div>
                    </td>
                </tr>
                """

        body = f"""
        {self._build_meta_grid({
            'Report Date': date.today().strftime('%d-%b-%Y'),
            'Company': company.company_name if company else 'N/A',
            'Total Clients with Materials': total_clients,
            'Total Projects Affected': total_projects,
            'Total Invoices': total_invoices,
            'Grand Total Materials': f"AED {grand_total_materials:,.2f}"
        })}

        <!-- Executive Summary Dashboard -->
        <div class="card" style="background: linear-gradient(135deg, #c62828 0%, #d32f2f 100%); color: white; margin-bottom: 20px;">
            <div class="card-header" style="color: white; border-color: rgba(255,255,255,0.3);">
                <span class="icon">📦</span> MATERIALS SUPPLIED BY CLIENT — EXECUTIVE SUMMARY
            </div>
            <div class="grid-4">
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 24px;">{total_clients}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Clients with Materials</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 24px;">{total_projects}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Projects Affected</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 24px;">{total_invoices}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Total Invoices</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 24px;">({grand_total_materials:,.2f})</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Grand Total Materials</div>
                </div>
            </div>
        </div>

        <!-- Client Summary Table -->
        <div class="card" style="margin-bottom: 20px;">
            <div class="card-header" style="color: #c62828; border-color: #ffcdd2;">
                <span class="icon">📊</span> CLIENT SUMMARY — MATERIALS SUPPLIED
            </div>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Client</th>
                        <th class="center">TRN</th>
                        <th class="center">Projects</th>
                        <th class="center">Invoices</th>
                        <th class="num">Total Materials</th>
                        <th style="width: 20%;">% of Grand Total</th>
                    </tr>
                </thead>
                <tbody>{summary_rows if summary_rows else '<tr><td colspan="6" style="text-align:center; color:#999; padding:20px;">No materials data found</td></tr>'}</tbody>
                <tfoot>
                    <tr class="grand-total">
                        <td colspan="4"><b>GRAND TOTAL</b></td>
                        <td class="num"><b>({grand_total_materials:,.2f})</b></td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <!-- Detailed Client Breakdown -->
        <div style="font-size: 14px; font-weight: 700; color: #1a237e; margin: 25px 0 15px 0; padding-bottom: 8px; border-bottom: 3px solid #1a237e;">
            DETAILED CLIENT BREAKDOWN
        </div>

        {client_sections if client_sections else '<div class="card" style="text-align:center; padding:40px; color:#666;"><p>No materials supplied by clients found.</p></div>'}

        <!-- Signatures -->
        <div class="signature-grid" style="margin-top: 40px;">
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">Finance Director</div>
            </div>
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">Technical Manager</div>
            </div>
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">General Manager</div>
            </div>
        </div>
        """

        return HttpResponse(self._report_base_wrapper(
            "MATERIALS SUPPLIED BY CLIENT — CONSOLIDATED REPORT",
            "All Materials Received Across All Clients & Projects",
            body,
            logo_url
        ))

    def retention_held_report(self, request):
        """
        Professional Retention Held Report - aggregates all held retention
        for all clients across all projects. Shows client-level summaries with
        project breakdowns for Retention A and B.
        """
        company = self.get_active_company(request)
        logo_url = company.logo.url if company and company.logo else ''

        # Get all clients with retention (have tax invoices with retention)
        client_ids = Invoice.objects.filter(
            inv_type='T'
        ).exclude(
            is_advance_invoice=True
        ).values_list('project__client', flat=True).distinct()

        clients_with_retention = Client.objects.filter(id__in=client_ids).order_by('name')

        if company:
            clients_with_retention = clients_with_retention.filter(
                Q(company=company) | Q(company__isnull=True)
            )

        # Build client sections
        client_sections = ""
        grand_total_ret_a_held = Decimal("0")
        grand_total_ret_b_held = Decimal("0")
        grand_total_ret_held = Decimal("0")
        grand_total_ret_a_deducted = Decimal("0")
        grand_total_ret_b_deducted = Decimal("0")
        grand_total_ret_a_recovered = Decimal("0")
        grand_total_ret_b_recovered = Decimal("0")
        total_clients = 0
        total_projects = 0

        for client in clients_with_retention:
            project_qs = Project.objects.filter(client=client)
            if company:
                project_qs = project_qs.filter(Q(company=company) | Q(company__isnull=True))

            project_cards = ""
            client_ret_a_held = Decimal("0")
            client_ret_b_held = Decimal("0")
            client_ret_held = Decimal("0")
            client_ret_a_ded = Decimal("0")
            client_ret_b_ded = Decimal("0")
            client_ret_a_rec = Decimal("0")
            client_ret_b_rec = Decimal("0")
            client_projects = 0

            for proj in project_qs:
                latest_inv = Invoice.objects.filter(
                    project=proj,
                    inv_type='T'
                ).exclude(is_advance_invoice=True).order_by('-inv_number').first()

                if not latest_inv:
                    continue

                ret_a_cum = latest_inv.cumulative_retention_a or Decimal("0")
                ret_b_cum = latest_inv.cumulative_retention_b or Decimal("0")
                ret_a_rec = latest_inv.cumulative_retention_a_recovered or Decimal("0")
                ret_b_rec = latest_inv.cumulative_retention_b_recovered or Decimal("0")

                ret_a_held = money(ret_a_cum - ret_a_rec)
                ret_b_held = money(ret_b_cum - ret_b_rec)
                total_held = ret_a_held + ret_b_held

                if total_held <= 0:
                    continue

                client_projects += 1
                client_ret_a_held += ret_a_held
                client_ret_b_held += ret_b_held
                client_ret_held += total_held
                client_ret_a_ded += ret_a_cum
                client_ret_b_ded += ret_b_cum
                client_ret_a_rec += ret_a_rec
                client_ret_b_rec += ret_b_rec

                # Get invoice history for this project's retention
                inv_rows = ""
                invoices = Invoice.objects.filter(
                    project=proj,
                    inv_type='T'
                ).exclude(is_advance_invoice=True).order_by('inv_number')

                for inv in invoices:
                    curr_ret_a = inv.current_retention_a or Decimal("0")
                    curr_ret_b = inv.current_retention_b or Decimal("0")
                    if curr_ret_a > 0 or curr_ret_b > 0:
                        status_badge = '<span class="badge badge-success">Paid</span>' if inv.status == 'Paid' else '<span class="badge badge-warning">Draft</span>' if inv.status == 'Draft' else '<span class="badge badge-info">Approved</span>'

                        inv_rows += f"""
                        <tr>
                            <td class="center">{inv.inv_number}</td>
                            <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                            <td class="num text-danger">({curr_ret_a:,.2f})</td>
                            <td class="num text-danger">({curr_ret_b:,.2f})</td>
                            <td class="num font-bold text-danger">({curr_ret_a + curr_ret_b:,.2f})</td>
                            <td class="center">{status_badge}</td>
                        </tr>
                        """

                project_cards += f"""
                <div style="margin-bottom: 12px; border: 1px solid #ffe0b2; border-radius: 6px; overflow: hidden;">
                    <div style="background: #fff3e0; padding: 8px 12px; font-size: 10px; font-weight: 700; color: #e65100; display: flex; justify-content: space-between;">
                        <span><span style="margin-right: 8px;">🏗️</span> {proj.project_id_code} — {proj.project_name}</span>
                        <span>Net Held: ({total_held:,.2f})</span>
                    </div>
                    <div style="padding: 10px; background: #fafafa;">
                        <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-bottom: 10px;">
                            <div style="text-align: center; padding: 8px; background: white; border-radius: 6px; border: 1px solid #ffe0b2; border-left: 3px solid #ed6c02;">
                                <div style="font-size: 14px; font-weight: 700; color: #ed6c02;">({ret_a_held:,.2f})</div>
                                <div style="font-size: 8px; color: #666; text-transform: uppercase; margin-top: 4px;">Ret A Held ({proj.retention_a_percent}%)</div>
                            </div>
                            <div style="text-align: center; padding: 8px; background: white; border-radius: 6px; border: 1px solid #ffe0b2; border-left: 3px solid #ed6c02;">
                                <div style="font-size: 14px; font-weight: 700; color: #ed6c02;">({ret_b_held:,.2f})</div>
                                <div style="font-size: 8px; color: #666; text-transform: uppercase; margin-top: 4px;">Ret B Held ({proj.retention_b_percent}%)</div>
                            </div>
                            <div style="text-align: center; padding: 8px; background: white; border-radius: 6px; border: 1px solid #ffcdd2; border-left: 3px solid #c62828;">
                                <div style="font-size: 16px; font-weight: 700; color: #c62828;">({total_held:,.2f})</div>
                                <div style="font-size: 8px; color: #666; text-transform: uppercase; margin-top: 4px;">Total Net Held</div>
                            </div>
                        </div>
                        <table class="data-table" style="margin: 0; font-size: 8.5px;">
                            <thead>
                                <tr>
                                    <th style="width: 8%;">Inv #</th>
                                    <th style="width: 12%;">Date</th>
                                    <th style="width: 15%;" class="num">Ret A</th>
                                    <th style="width: 15%;" class="num">Ret B</th>
                                    <th style="width: 15%;" class="num">Total</th>
                                    <th style="width: 10%;">Status</th>
                                </tr>
                            </thead>
                            <tbody>{inv_rows if inv_rows else '<tr><td colspan="6" style="text-align:center; color:#999; padding:8px;">No retention deductions</td></tr>'}</tbody>
                            <tfoot>
                                <tr style="background: #fff3e0; font-weight: bold;">
                                    <td colspan="2"><b>PROJECT CUMULATIVE</b></td>
                                    <td class="num text-danger"><b>({ret_a_cum:,.2f})</b></td>
                                    <td class="num text-danger"><b>({ret_b_cum:,.2f})</b></td>
                                    <td class="num text-danger"><b>({ret_a_cum + ret_b_cum:,.2f})</b></td>
                                    <td></td>
                                </tr>
                                <tr style="background: #e8f5e9; font-weight: bold;">
                                    <td colspan="2"><b>RECOVERED</b></td>
                                    <td class="num text-success"><b>{ret_a_rec:,.2f}</b></td>
                                    <td class="num text-success"><b>{ret_b_rec:,.2f}</b></td>
                                    <td class="num text-success"><b>{ret_a_rec + ret_b_rec:,.2f}</b></td>
                                    <td></td>
                                </tr>
                            </tfoot>
                        </table>
                    </div>
                </div>
                """

            if client_ret_held > 0:
                total_clients += 1
                total_projects += client_projects
                grand_total_ret_a_held += client_ret_a_held
                grand_total_ret_b_held += client_ret_b_held
                grand_total_ret_held += client_ret_held
                grand_total_ret_a_deducted += client_ret_a_ded
                grand_total_ret_b_deducted += client_ret_b_ded
                grand_total_ret_a_recovered += client_ret_a_rec
                grand_total_ret_b_recovered += client_ret_b_rec

                client_sections += f"""
                <div class="card" style="page-break-inside: avoid; margin-bottom: 20px; border-left: 4px solid #ed6c02;">
                    <div class="card-header" style="display: flex; justify-content: space-between; align-items: center; color: #e65100; border-color: #ffe0b2;">
                        <span><span class="icon">👤</span> {client.name}</span>
                        <span class="badge badge-warning">Client Total Held: ({client_ret_held:,.2f})</span>
                    </div>
                    <div style="padding: 12px;">
                        <div class="meta-grid" style="margin-bottom: 12px; grid-template-columns: repeat(4, 1fr);">
                            <div class="meta-item" style="border-left-color: #ed6c02;">
                                <div class="meta-label">TRN</div>
                                <div class="meta-value">{client.vat_number or 'N/A'}</div>
                            </div>
                            <div class="meta-item" style="border-left-color: #ed6c02;">
                                <div class="meta-label">Active Projects</div>
                                <div class="meta-value">{client_projects}</div>
                            </div>
                            <div class="meta-item" style="border-left-color: #ed6c02;">
                                <div class="meta-label">Ret A Held</div>
                                <div class="meta-value text-warning">({client_ret_a_held:,.2f})</div>
                            </div>
                            <div class="meta-item" style="border-left-color: #ed6c02;">
                                <div class="meta-label">Ret B Held</div>
                                <div class="meta-value text-warning">({client_ret_b_held:,.2f})</div>
                            </div>
                        </div>
                        {project_cards}
                    </div>
                </div>
                """

        # Build summary table
        summary_rows = ""
        for client in clients_with_retention:
            client_ret_a = Decimal("0")
            client_ret_b = Decimal("0")
            client_total = Decimal("0")
            proj_count = 0

            project_qs = Project.objects.filter(client=client)
            if company:
                project_qs = project_qs.filter(Q(company=company) | Q(company__isnull=True))

            for proj in project_qs:
                latest_inv = Invoice.objects.filter(
                    project=proj,
                    inv_type='T'
                ).exclude(is_advance_invoice=True).order_by('-inv_number').first()

                if latest_inv:
                    ret_a_held = money(latest_inv.cumulative_retention_a - latest_inv.cumulative_retention_a_recovered)
                    ret_b_held = money(latest_inv.cumulative_retention_b - latest_inv.cumulative_retention_b_recovered)
                    total = ret_a_held + ret_b_held

                    if total > 0:
                        client_ret_a += ret_a_held
                        client_ret_b += ret_b_held
                        client_total += total
                        proj_count += 1

            if client_total > 0:
                pct_of_grand = (client_total / grand_total_ret_held * 100) if grand_total_ret_held > 0 else Decimal("0")
                summary_rows += f"""
                <tr>
                    <td class="text"><b>{client.name}</b></td>
                    <td class="center">{client.vat_number or 'N/A'}</td>
                    <td class="center">{proj_count}</td>
                    <td class="num text-warning">({client_ret_a:,.2f})</td>
                    <td class="num text-warning">({client_ret_b:,.2f})</td>
                    <td class="num font-bold text-danger">({client_total:,.2f})</td>
                    <td class="center">
                        <div class="progress-bar" style="height: 16px; margin-top: 0;">
                            <div class="progress-fill" style="width: {min(float(pct_of_grand), 100)}%; background: linear-gradient(90deg, #ed6c02, #f57c00); font-size: 7px;">
                                {pct_of_grand:.1f}%
                            </div>
                        </div>
                    </td>
                </tr>
                """

        body = f"""
        {self._build_meta_grid({
            'Report Date': date.today().strftime('%d-%b-%Y'),
            'Company': company.company_name if company else 'N/A',
            'Total Clients with Retention': total_clients,
            'Total Projects Affected': total_projects,
            'Grand Total Retention Held': f"AED {grand_total_ret_held:,.2f}"
        })}

        <!-- Executive Summary Dashboard -->
        <div class="card" style="background: linear-gradient(135deg, #e65100 0%, #ed6c02 100%); color: white; margin-bottom: 20px;">
            <div class="card-header" style="color: white; border-color: rgba(255,255,255,0.3);">
                <span class="icon">🔒</span> RETENTION HELD — EXECUTIVE SUMMARY
            </div>
            <div class="grid-4">
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 24px;">{total_clients}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Clients with Retention</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 24px;">({grand_total_ret_a_held:,.2f})</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Total Retention A Held</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 24px;">({grand_total_ret_b_held:,.2f})</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Total Retention B Held</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 24px;">({grand_total_ret_held:,.2f})</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Grand Total Retention Held</div>
                </div>
            </div>
        </div>

        <!-- Client Summary Table -->
        <div class="card" style="margin-bottom: 20px;">
            <div class="card-header" style="color: #e65100; border-color: #ffe0b2;">
                <span class="icon">📊</span> CLIENT SUMMARY — RETENTION HELD
            </div>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Client</th>
                        <th class="center">TRN</th>
                        <th class="center">Projects</th>
                        <th class="num">Ret A Held</th>
                        <th class="num">Ret B Held</th>
                        <th class="num">Total Held</th>
                        <th style="width: 15%;">% of Grand Total</th>
                    </tr>
                </thead>
                <tbody>{summary_rows if summary_rows else '<tr><td colspan="7" style="text-align:center; color:#999; padding:20px;">No retention data found</td></tr>'}</tbody>
                <tfoot>
                    <tr class="grand-total">
                        <td colspan="3"><b>GRAND TOTAL</b></td>
                        <td class="num"><b>({grand_total_ret_a_held:,.2f})</b></td>
                        <td class="num"><b>({grand_total_ret_b_held:,.2f})</b></td>
                        <td class="num"><b>({grand_total_ret_held:,.2f})</b></td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <!-- Detailed Client Breakdown -->
        <div style="font-size: 14px; font-weight: 700; color: #1a237e; margin: 25px 0 15px 0; padding-bottom: 8px; border-bottom: 3px solid #1a237e;">
            DETAILED CLIENT BREAKDOWN
        </div>

        {client_sections if client_sections else '<div class="card" style="text-align:center; padding:40px; color:#666;"><p>No retention held found.</p></div>'}

        <!-- Retention Recovery Status -->
        <div class="card" style="margin-top: 20px; background: #fafafa; border: 2px solid #ed6c02;">
            <div class="card-header" style="color: #e65100; border-color: #ffe0b2;">
                <span class="icon">📋</span> RETENTION RECOVERY STATUS OVERVIEW
            </div>
            <table class="data-table" style="font-size: 10px;">
                <thead>
                    <tr>
                        <th>Metric</th>
                        <th class="num">Retention A</th>
                        <th class="num">Retention B</th>
                        <th class="num">Total</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td><b>Total Deducted (Cumulative)</b></td>
                        <td class="num text-danger">({grand_total_ret_a_deducted:,.2f})</td>
                        <td class="num text-danger">({grand_total_ret_b_deducted:,.2f})</td>
                        <td class="num text-danger">({grand_total_ret_a_deducted + grand_total_ret_b_deducted:,.2f})</td>
                    </tr>
                    <tr>
                        <td><b>Total Recovered</b></td>
                        <td class="num text-success">{grand_total_ret_a_recovered:,.2f}</td>
                        <td class="num text-success">{grand_total_ret_b_recovered:,.2f}</td>
                        <td class="num text-success">{grand_total_ret_a_recovered + grand_total_ret_b_recovered:,.2f}</td>
                    </tr>
                    <tr style="background: #fff3e0; font-weight: bold;">
                        <td><b>NET HELD (Current)</b></td>
                        <td class="num text-warning">({grand_total_ret_a_held:,.2f})</td>
                        <td class="num text-warning">({grand_total_ret_b_held:,.2f})</td>
                        <td class="num text-danger">({grand_total_ret_held:,.2f})</td>
                    </tr>
                </tbody>
            </table>
        </div>

        <!-- Signatures -->
        <div class="signature-grid" style="margin-top: 40px;">
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">Finance Director</div>
            </div>
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">Technical Manager</div>
            </div>
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">General Manager</div>
            </div>
        </div>
        """

        return HttpResponse(self._report_base_wrapper(
            "RETENTION HELD — CONSOLIDATED REPORT",
            "All Retention Held Across All Clients & Projects",
            body,
            logo_url
        ))

    def cash_flow_report(self, request):
        """
        Professional Cash Flow Prediction Report.

        INFLOW: From approved/pending client tax invoices with estimated collection dates
                based on invoice date + project payment terms.

        OUTFLOW: From supplier invoices with expected_payment_date.
        Includes: Supplier payments, payroll estimates, and other known outflows.

        Shows 13-week rolling forecast with weekly buckets.
        """
        from datetime import datetime, timedelta
        from collections import OrderedDict
        from dateutil.relativedelta import relativedelta

        company = self.get_active_company(request)
        logo_url = company.logo.url if company and company.logo else ''
        today = date.today()

        # ─────────────────────────────────────────────────────────────
        # CONFIGURATION
        # ─────────────────────────────────────────────────────────────
        WEEKS_AHEAD = 13  # 13-week rolling forecast
        WEEK_START = today - timedelta(days=today.weekday())  # Monday of current week

        # ─────────────────────────────────────────────────────────────
        # BUILD WEEK BUCKETS
        # ─────────────────────────────────────────────────────────────
        weeks = []
        for i in range(WEEKS_AHEAD):
            week_start = WEEK_START + timedelta(weeks=i)
            week_end = week_start + timedelta(days=6)
            weeks.append({
                'start': week_start,
                'end': week_end,
                'label': f"{week_start.strftime('%d %b')} — {week_end.strftime('%d %b')}",
                'month_label': week_start.strftime('%b %Y'),
            })

        # ─────────────────────────────────────────────────────────────
        # FETCH INFLOW DATA (Client Invoice Collections)
        # ─────────────────────────────────────────────────────────────

        # Approved tax invoices that are NOT paid yet - these are expected collections
        client_invoices = Invoice.objects.filter(
            inv_type='T',
            status__in=['Approved', 'Draft']  # Approved = sent to client, Draft = being prepared
        ).exclude(is_advance_invoice=True).select_related('project', 'project__client')

        if company:
            client_invoices = client_invoices.filter(project__company=company)

        # For each invoice, estimate collection date = invoice date + project payment terms
        inflow_items = []
        for inv in client_invoices:
            payment_terms = inv.project.payment_terms or 30
            estimated_collection = inv.date + timedelta(days=payment_terms)

            # Adjust for weekends (move to next Monday if falls on weekend)
            while estimated_collection.weekday() >= 5:  # Saturday=5, Sunday=6
                estimated_collection += timedelta(days=1)

            # Probability factor based on status
            probability = Decimal("0.9") if inv.status == 'Approved' else Decimal("0.6")

            inflow_items.append({
                'type': 'Client Invoice',
                'description': str(inv),
                'project': inv.project.project_name,
                'client': inv.project.client.name,
                'amount': inv.total_after_vat,  # Total including VAT - what client will pay
                'net_amount': inv.current_certified_net_before_vat,
                'estimated_date': estimated_collection,
                'probability': probability,
                'weighted_amount': inv.total_after_vat * probability,
                'status': inv.status,
                'confidence': 'High' if inv.status == 'Approved' else 'Medium',
            })

        # Also include PAID invoices with collection_date in the future (rare but possible)
        paid_future = Invoice.objects.filter(
            inv_type='T',
            status='Paid',
            collection_date__gt=today
        ).exclude(is_advance_invoice=True).select_related('project', 'project__client')

        if company:
            paid_future = paid_future.filter(project__company=company)

        for inv in paid_future:
            inflow_items.append({
                'type': 'Client Invoice (Confirmed)',
                'description': str(inv),
                'project': inv.project.project_name,
                'client': inv.project.client.name,
                'amount': inv.total_after_vat,
                'net_amount': inv.current_certified_net_before_vat,
                'estimated_date': inv.collection_date,
                'probability': Decimal("1.0"),
                'weighted_amount': inv.total_after_vat,
                'status': 'Paid (Pending Collection)',
                'confidence': 'Confirmed',
            })

        # ─────────────────────────────────────────────────────────────
        # FETCH OUTFLOW DATA (Supplier Payments)
        # ─────────────────────────────────────────────────────────────

        supplier_invoices = SupplierInvoice.objects.filter(
            status__in=['Draft', 'Approved', 'Scheduled']
        ).exclude(status='Paid').exclude(status='Cancelled').select_related('supplier', 'project')

        if company:
            supplier_invoices = supplier_invoices.filter(company=company)

        outflow_items = []
        for inv in supplier_invoices:
            pay_date = inv.expected_payment_date or inv.due_date

            outflow_items.append({
                'type': 'Supplier Payment',
                'description': f"{inv.supplier.name} — {inv.supplier_inv_number}",
                'supplier': inv.supplier.name,
                'project': inv.project.project_name if inv.project else 'Overhead',
                'amount': inv.balance_due,
                'total_amount': inv.total_amount,
                'expected_date': pay_date,
                'status': inv.status,
                'category': inv.supplier.category,
                'is_recurring': inv.is_recurring,
            })

        # ─────────────────────────────────────────────────────────────
        # FETCH PAYROLL OUTFLOW (Estimated)
        # ─────────────────────────────────────────────────────────────

        # Get active employees and estimate monthly payroll
        emp_filter = {'is_active': True}
        if company:
            emp_filter['company'] = company

        active_employees = Employee.objects.filter(**emp_filter)
        monthly_payroll = Decimal("0")
        for emp in active_employees:
            monthly_payroll += emp.total_salary + emp.monthly_admin_cost

        # Generate payroll outflow items for next 3 months
        payroll_items = []
        for i in range(3):
            payroll_date = today.replace(day=1) + relativedelta(months=i)
            # Payroll typically paid by end of month or early next month
            pay_date = payroll_date + relativedelta(months=1, day=5)  # 5th of next month

            payroll_items.append({
                'type': 'Payroll',
                'description': f"Monthly Payroll — {payroll_date.strftime('%b %Y')}",
                'amount': monthly_payroll,
                'expected_date': pay_date,
                'status': 'Projected',
                'category': 'Payroll',
            })

        # ─────────────────────────────────────────────────────────────
        # AGGREGATE BY WEEK
        # ─────────────────────────────────────────────────────────────

        week_data = []
        cumulative_inflow = Decimal("0")
        cumulative_outflow = Decimal("0")

        for week in weeks:
            week_inflow = Decimal("0")
            week_weighted_inflow = Decimal("0")
            week_outflow = Decimal("0")
            week_inflow_items = []
            week_outflow_items = []

            # Client inflows
            for item in inflow_items:
                if week['start'] <= item['estimated_date'] <= week['end']:
                    week_inflow += item['amount']
                    week_weighted_inflow += item['weighted_amount']
                    week_inflow_items.append(item)

            # Supplier outflows
            for item in outflow_items:
                if week['start'] <= item['expected_date'] <= week['end']:
                    week_outflow += item['amount']
                    week_outflow_items.append(item)

            # Payroll outflows
            for item in payroll_items:
                if week['start'] <= item['expected_date'] <= week['end']:
                    week_outflow += item['amount']
                    week_outflow_items.append(item)

            net_flow = week_inflow - week_outflow
            cumulative_inflow += week_inflow
            cumulative_outflow += week_outflow
            running_balance = cumulative_inflow - cumulative_outflow

            week_data.append({
                'week': week,
                'inflow': week_inflow,
                'weighted_inflow': week_weighted_inflow,
                'outflow': week_outflow,
                'net_flow': net_flow,
                'cumulative_inflow': cumulative_inflow,
                'cumulative_outflow': cumulative_outflow,
                'running_balance': running_balance,
                'inflow_items': week_inflow_items,
                'outflow_items': week_outflow_items,
            })

        # ─────────────────────────────────────────────────────────────
        # SUMMARY METRICS
        # ─────────────────────────────────────────────────────────────

        total_inflow = sum(w['inflow'] for w in week_data)
        total_weighted_inflow = sum(w['weighted_inflow'] for w in week_data)
        total_outflow = sum(w['outflow'] for w in week_data)
        net_position = total_inflow - total_outflow

        # Current liquidity position (simplified)
        opening_balance = Decimal("0")  # User should input this
        closing_balance = opening_balance + net_position

        # Risk assessment
        risk_level = "Low"
        risk_color = "#2e7d32"
        if closing_balance < 0:
            risk_level = "CRITICAL"
            risk_color = "#c62828"
        elif closing_balance < total_outflow * Decimal("0.2"):
            risk_level = "High"
            risk_color = "#d32f2f"
        elif closing_balance < total_outflow * Decimal("0.5"):
            risk_level = "Medium"
            risk_color = "#f57c00"

        # ─────────────────────────────────────────────────────────────
        # BUILD WEEKLY TABLE ROWS
        # ─────────────────────────────────────────────────────────────

        week_rows = ""
        for w in week_data:
            net_color = "#2e7d32" if w['net_flow'] >= 0 else "#c62828"
            balance_color = "#2e7d32" if w['running_balance'] >= 0 else "#c62828"

            # Inflow detail tooltip
            inflow_detail = "<br>".join([
                f"• {item['client']}: AED {item['amount']:,.2f} ({item['confidence']})"
                for item in w['inflow_items'][:3]
            ])
            if len(w['inflow_items']) > 3:
                inflow_detail += f"<br>• ... and {len(w['inflow_items']) - 3} more"

            # Outflow detail tooltip
            outflow_detail = "<br>".join([
                f"• {item['supplier'] if 'supplier' in item else item['description']}: AED {item['amount']:,.2f}"
                for item in w['outflow_items'][:3]
            ])
            if len(w['outflow_items']) > 3:
                outflow_detail += f"<br>• ... and {len(w['outflow_items']) - 3} more"

            week_rows += f"""
            <tr style="border-left: 4px solid {net_color};">
                <td style="font-weight: bold; padding: 10px 8px;">{w['week']['label']}</td>
                <td class="num" style="color: #2e7d32; font-weight: 600;">{w['inflow']:,.2f}</td>
                <td class="num" style="color: #6a1b9a; font-size: 9px;">{w['weighted_inflow']:,.2f}</td>
                <td class="num" style="color: #c62828; font-weight: 600;">{w['outflow']:,.2f}</td>
                <td class="num" style="color: {net_color}; font-weight: bold; font-size: 11px;">{w['net_flow']:,.2f}</td>
                <td class="num" style="color: {balance_color}; font-weight: bold;">{w['running_balance']:,.2f}</td>
                <td style="font-size: 8px; color: #666; max-width: 150px;">{inflow_detail or '—'}</td>
                <td style="font-size: 8px; color: #666; max-width: 150px;">{outflow_detail or '—'}</td>
            </tr>
            """

        # ─────────────────────────────────────────────────────────────
        # INFLOW DETAIL TABLE
        # ─────────────────────────────────────────────────────────────

        inflow_rows = ""
        for item in sorted(inflow_items, key=lambda x: x['estimated_date']):
            conf_color = "#2e7d32" if item['confidence'] == 'Confirmed' else "#f57c00" if item['confidence'] == 'High' else "#ed6c02"
            inflow_rows += f"""
            <tr>
                <td class="center">{item['estimated_date'].strftime('%d-%b-%Y')}</td>
                <td class="text"><b>{item['description']}</b></td>
                <td class="text">{item['client']}</td>
                <td class="text">{item['project']}</td>
                <td class="num font-bold">{item['amount']:,.2f}</td>
                <td class="num" style="color: #6a1b9a;">{item['weighted_amount']:,.2f}</td>
                <td class="center"><span class="badge badge-success">{item['status']}</span></td>
                <td class="center" style="color: {conf_color}; font-weight: bold;">{item['confidence']}</td>
            </tr>
            """

        # ─────────────────────────────────────────────────────────────
        # OUTFLOW DETAIL TABLE
        # ─────────────────────────────────────────────────────────────

        outflow_rows = ""
        for item in sorted(outflow_items + payroll_items,
                           key=lambda x: x.get('expected_date', x.get('estimated_date', today))):
            pay_date = item.get('expected_date', item.get('estimated_date', today))
            is_payroll = item['type'] == 'Payroll'
            status_badge = '<span class="badge badge-primary">Scheduled</span>' if is_payroll else f'<span class="badge badge-warning">{item["status"]}</span>'

            outflow_rows += f"""
            <tr>
                <td class="center">{pay_date.strftime('%d-%b-%Y')}</td>
                <td class="text"><b>{item['description']}</b></td>
                <td class="text">{item.get('supplier', item.get('category', 'N/A'))}</td>
                <td class="text">{item.get('project', 'Overhead')}</td>
                <td class="num font-bold text-danger">({item['amount']:,.2f})</td>
                <td class="center">{status_badge}</td>
                <td class="center">{'🔄' if item.get('is_recurring') else ''}</td>
            </tr>
            """

        # ─────────────────────────────────────────────────────────────
        # SUPPLIER SUMMARY CARDS
        # ─────────────────────────────────────────────────────────────

        supplier_summary = SupplierInvoice.objects.filter(
            status__in=['Draft', 'Approved', 'Scheduled']
        ).exclude(status='Paid').exclude(status='Cancelled')

        if company:
            supplier_summary = supplier_summary.filter(company=company)

        supplier_totals = {}
        for inv in supplier_summary.select_related('supplier'):
            cat = inv.supplier.category
            if cat not in supplier_totals:
                supplier_totals[cat] = Decimal("0")
            supplier_totals[cat] += inv.balance_due

        supplier_cards = ""
        cat_colors = {
            'Material': '#1565c0',
            'Subcontractor': '#6a1b9a',
            'Equipment': '#ed6c02',
            'Service': '#2e7d32',
            'Utility': '#00897b',
            'Other': '#757575',
        }

        for cat, total in sorted(supplier_totals.items(), key=lambda x: x[1], reverse=True):
            color = cat_colors.get(cat, '#757575')
            pct = (total / total_outflow * 100) if total_outflow > 0 else 0
            supplier_cards += f"""
            <div class="metric-card" style="border-left: 4px solid {color};">
                <div class="metric-value" style="color: {color}; font-size: 18px;">AED {total:,.2f}</div>
                <div class="metric-label">{cat}</div>
                <div style="font-size: 8px; color: #666; margin-top: 4px;">{pct:.1f}% of total outflow</div>
            </div>
            """

        # ─────────────────────────────────────────────────────────────
        # BUILD FINAL HTML
        # ─────────────────────────────────────────────────────────────

        body = f"""
        {self._build_meta_grid({
            'Report Date': today.strftime('%d-%b-%Y'),
            'Forecast Period': f"{weeks[0]['label']} to {weeks[-1]['label']}",
            'Company': company.company_name if company else 'N/A',
            'Risk Level': risk_level,
        })}

        <!-- Risk Alert Banner -->
        <div style="background: {risk_color}; color: white; padding: 12px 16px; border-radius: 8px; margin-bottom: 20px; text-align: center;">
            <div style="font-size: 11px; text-transform: uppercase; letter-spacing: 1px; opacity: 0.9; margin-bottom: 4px;">Cash Position Risk Assessment</div>
            <div style="font-size: 22px; font-weight: 700;">{risk_level}</div>
            <div style="font-size: 10px; opacity: 0.8; margin-top: 4px;">
                Net Position: AED {net_position:,.2f} | 
                Weighted Inflow: AED {total_weighted_inflow:,.2f}
            </div>
        </div>

        <!-- Executive Summary Dashboard -->
        <div class="card" style="background: linear-gradient(135deg, #1a237e 0%, #283593 100%); color: white; margin-bottom: 20px;">
            <div class="card-header" style="color: white; border-color: rgba(255,255,255,0.3);">
                <span class="icon">💰</span> CASH FLOW EXECUTIVE SUMMARY (13-Week Forecast)
            </div>
            <div class="grid-5">
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 22px;">AED {total_inflow:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Total Expected Inflow</div>
                    <div style="font-size: 8px; color: rgba(255,255,255,0.7); margin-top: 4px;">From {len(inflow_items)} client invoices</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 22px;">AED {total_weighted_inflow:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Weighted Inflow (Risk-Adjusted)</div>
                    <div style="font-size: 8px; color: rgba(255,255,255,0.7); margin-top: 4px;">Probability-adjusted</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 22px;">AED {total_outflow:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Total Expected Outflow</div>
                    <div style="font-size: 8px; color: rgba(255,255,255,0.7); margin-top: 4px;">Suppliers + Payroll</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: {'#2e7d32' if net_position >= 0 else '#ffcdd2'}; font-size: 22px;">AED {net_position:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Net Cash Flow</div>
                    <div style="font-size: 8px; color: rgba(255,255,255,0.7); margin-top: 4px;">Inflow minus Outflow</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 22px;">{len(inflow_items)} / {len(outflow_items) + len(payroll_items)}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Inflow / Outflow Items</div>
                    <div style="font-size: 8px; color: rgba(255,255,255,0.7); margin-top: 4px;">Client inv / Supplier inv</div>
                </div>
            </div>
        </div>

        <!-- Weekly Cash Flow Table -->
        <div class="card" style="margin-bottom: 20px;">
            <div class="card-header">
                <span class="icon">📅</span> WEEKLY CASH FLOW FORECAST
            </div>
            <table class="data-table" style="font-size: 9px;">
                <thead>
                    <tr>
                        <th style="width: 14%;">Week</th>
                        <th class="num" style="width: 12%;">Inflow</th>
                        <th class="num" style="width: 12%;">Weighted Inflow</th>
                        <th class="num" style="width: 12%;">Outflow</th>
                        <th class="num" style="width: 12%;">Net Flow</th>
                        <th class="num" style="width: 12%;">Running Balance</th>
                        <th style="width: 13%;">Inflow Sources</th>
                        <th style="width: 13%;">Outflow Sources</th>
                    </tr>
                </thead>
                <tbody>
                    {week_rows}
                </tbody>
                <tfoot>
                    <tr class="grand-total">
                        <td><b>TOTAL 13-WEEK</b></td>
                        <td class="num"><b>{total_inflow:,.2f}</b></td>
                        <td class="num"><b>{total_weighted_inflow:,.2f}</b></td>
                        <td class="num"><b>{total_outflow:,.2f}</b></td>
                        <td class="num"><b>{net_position:,.2f}</b></td>
                        <td class="num">—</td>
                        <td colspan="2"></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <!-- Supplier Category Breakdown -->
        <div class="card" style="margin-bottom: 20px;">
            <div class="card-header" style="color: #c62828; border-color: #ffcdd2;">
                <span class="icon">📦</span> OUTFLOW BY SUPPLIER CATEGORY
            </div>
            <div class="grid-3">
                {supplier_cards if supplier_cards else '<div style="text-align:center; color:#999; padding:20px;">No supplier invoices pending</div>'}
            </div>
        </div>

        <!-- Inflow Detail -->
        <div style="font-size: 14px; font-weight: 700; color: #1a237e; margin: 25px 0 15px 0; padding-bottom: 8px; border-bottom: 3px solid #1a237e;">
            EXPECTED CLIENT COLLECTIONS (INFLOW)
        </div>
        <div class="card" style="margin-bottom: 20px;">
            <div class="card-header" style="color: #2e7d32; border-color: #c8e6c9;">
                <span class="icon">📥</span> CLIENT INVOICE COLLECTION SCHEDULE
            </div>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Est. Collection</th>
                        <th>Invoice</th>
                        <th>Client</th>
                        <th>Project</th>
                        <th class="num">Amount</th>
                        <th class="num">Weighted</th>
                        <th class="center">Status</th>
                        <th class="center">Confidence</th>
                    </tr>
                </thead>
                <tbody>
                    {inflow_rows if inflow_rows else '<tr><td colspan="8" style="text-align:center; color:#999; padding:20px;">No pending client invoices</td></tr>'}
                </tbody>
            </table>
        </div>

        <!-- Outflow Detail -->
        <div style="font-size: 14px; font-weight: 700; color: #1a237e; margin: 25px 0 15px 0; padding-bottom: 8px; border-bottom: 3px solid #1a237e;">
            PLANNED PAYMENTS (OUTFLOW)
        </div>
        <div class="card" style="margin-bottom: 20px;">
            <div class="card-header" style="color: #c62828; border-color: #ffcdd2;">
                <span class="icon">📤</span> SUPPLIER & PAYROLL PAYMENT SCHEDULE
            </div>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Payment Date</th>
                        <th>Description</th>
                        <th>Supplier/Category</th>
                        <th>Project</th>
                        <th class="num">Amount</th>
                        <th class="center">Status</th>
                        <th class="center">Recurring</th>
                    </tr>
                </thead>
                <tbody>
                    {outflow_rows if outflow_rows else '<tr><td colspan="7" style="text-align:center; color:#999; padding:20px;">No pending payments</td></tr>'}
                </tbody>
            </table>
        </div>

        <!-- Key Alerts -->
        <div class="card" style="margin-bottom: 20px; background: #fff3e0; border: 2px solid #ed6c02;">
            <div class="card-header" style="color: #e65100; border-color: #ffe0b2;">
                <span class="icon">⚠️</span> CASH FLOW ALERTS & RECOMMENDATIONS
            </div>
            <div style="padding: 12px; font-size: 10px; line-height: 1.8;">
                <ul style="margin: 0; padding-left: 16px;">
                    <li><b>Inflow Concentration:</b> {len(inflow_items)} client invoices pending collection. 
                        Weighted inflow is AED {total_weighted_inflow:,.2f} (risk-adjusted).</li>
                    <li><b>Outflow Obligations:</b> {len(outflow_items)} supplier invoices + {len(payroll_items)} payroll periods = 
                        AED {total_outflow:,.2f} total commitments.</li>
                    <li><b>Collection Timing:</b> Inflows are estimated based on invoice date + payment terms. 
                        Actual collection may vary based on client payment behavior.</li>
                    <li><b>Recommendation:</b> {"URGENT: Negative cash position predicted. Consider accelerating collections or deferring non-critical payments." if net_position < 0 else "Cash position is positive. Monitor weekly to ensure sufficient liquidity for large outflows."}</li>
                </ul>
            </div>
        </div>

        <!-- Signatures -->
        <div class="signature-grid" style="margin-top: 40px;">
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">Finance Director</div>
            </div>
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">General Manager</div>
            </div>
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">Date</div>
            </div>
        </div>
        """

        # Add extra CSS for grid-5
        extra_css = """
        <style>
            .grid-5 { display: grid; grid-template-columns: repeat(5, 1fr); gap: 12px; }
            @media (max-width: 1200px) { .grid-5 { grid-template-columns: repeat(3, 1fr); } }
            @media (max-width: 768px) { .grid-5 { grid-template-columns: repeat(2, 1fr); } }
        </style>
        """

        return HttpResponse(self._report_base_wrapper(
            "CASH FLOW PREDICTION REPORT",
            f"13-Week Rolling Forecast — {today.strftime('%d %B %Y')}",
            body,
            logo_url,
            extra_css
        ))

    def get_urls(self):
        urls = super().get_urls()
        # These URLs serve as entry points that redirect to actual report lists
        custom = [
            path('client-statements/', self.admin_site.admin_view(self.client_statement_list),
                 name='client_statement_list'),
            path('client-outstanding/', self.admin_site.admin_view(self.client_outstanding_list),
                 name='client_outstanding_list'),
            path('client-progress/', self.admin_site.admin_view(self.client_progress_list),
                 name='client_progress_list'),
            path('client-deductions/', self.admin_site.admin_view(self.client_deductions_list),
                 name='client_deductions_list'),  # <-- ADD THIS
            path('project-analytics/', self.admin_site.admin_view(self.project_analytics_list),
                 name='project_analytics_list'),
            path('project-cost-profit/', self.admin_site.admin_view(self.project_cost_profit_list),
                 name='project_cost_profit_list'),
            path('project-deductions/', self.admin_site.admin_view(self.project_deductions_list),
                 name='project_deductions_list'),  # <-- ADD THIS
            path('invoice-print/', self.admin_site.admin_view(self.invoice_print_list), name='invoice_print_list'),
            path('payroll-timesheet/', self.admin_site.admin_view(self.payroll_timesheet_list),
                 name='payroll_timesheet_list'),
            path('payroll-labor-cost/', self.admin_site.admin_view(self.payroll_labor_cost_list),
                 name='payroll_labor_cost_list'),
            path('payroll-staff/', self.admin_site.admin_view(self.payroll_staff_report_list),
                 name='payroll_staff_report_list'),
            path('payroll-wps/', self.admin_site.admin_view(self.payroll_wps_report_list),
                 name='payroll_wps_report_list'),
            path('payroll-cash/', self.admin_site.admin_view(self.payroll_cash_report_list),
                 name='payroll_cash_report_list'),
            path('materials-supplied/', self.admin_site.admin_view(self.materials_supplied_report),
                 name='materials_supplied_report'),
            path('retention-held/', self.admin_site.admin_view(self.retention_held_report),
                 name='retention_held_report'),
            path('cash-flow/', self.admin_site.admin_view(self.cash_flow_report), name='cash_flow_report'),
        ]
        return custom + urls

    # These views redirect to the first available object or show a selector
    def client_statement_list(self, request):
        return self._report_selector(request, Client, 'admin:client_statement', 'Client Statement of Account')

    def client_outstanding_list(self, request):
        return self._report_selector(request, Client, 'admin:client_outstanding', 'Outstanding Invoices Report')

    def client_progress_list(self, request):
        return self._report_selector(request, Client, 'admin:client_progress', 'Client Project Progress')

    def client_deductions_list(self, request):
        return self._report_selector(request, Client, 'admin:client_deductions', 'Client Deductions Report')

    def project_analytics_list(self, request):
        return self._report_selector(request, Project, 'admin:project_analytics', 'Project Analytics')

    def project_cost_profit_list(self, request):
        return self._report_selector(request, Project, 'admin:project_cost_profit', 'Cost & Profitability')

    def project_deductions_list(self, request):
        return self._report_selector(request, Project, 'admin:project_deductions', 'Project Deductions Report')

    def invoice_print_list(self, request):
        return self._report_selector(request, Invoice, 'admin:invoice_print', 'Print Invoice')

    def payroll_timesheet_list(self, request):
        return self._report_selector(request, PayrollRecord, 'admin:payroll_timesheet', 'Worker Time Sheet')

    def payroll_labor_cost_list(self, request):
        return redirect(reverse('admin:payroll_labor_cost'))

    def payroll_staff_report_list(self, request):
        return redirect(reverse('admin:payroll_staff_report') + '?month=' + date.today().strftime('%Y-%m'))

    def payroll_wps_report_list(self, request):
        return redirect(reverse('admin:payroll_wps_report') + '?month=' + date.today().strftime('%Y-%m'))

    def payroll_cash_report_list(self, request):
        return redirect(reverse('admin:payroll_cash_report') + '?month=' + date.today().strftime('%Y-%m'))

    def _report_selector(self, request, model, url_name, report_title):
        """FIX: Show a professional selector page with all available objects instead of redirecting to first."""
        company = self.get_active_company(request)
        qs = model.objects.all()
        if company and hasattr(model, 'company'):
            qs = qs.filter(company=company)
        elif company and hasattr(model, 'project'):
            qs = qs.filter(project__company=company)

        objects_list = list(qs)

        if not objects_list:
            # Show empty state if no objects
            html = f"""
            <div style="padding:40px; text-align:center;">
                <h2 style="color:#1a237e; margin-bottom:20px;">No {model._meta.verbose_name_plural} Found</h2>
                <p style="color:#666; margin-bottom:20px;">Please create a {model._meta.verbose_name} first to view reports.</p>
                <a href="{reverse(f'admin:billing_{model._meta.model_name}_changelist')}" 
                   style="display:inline-block; background:#1a237e; color:white; padding:12px 24px; 
                          border-radius:6px; text-decoration:none; font-weight:600;">
                    Go to {model._meta.verbose_name_plural}
                </a>
            </div>
            """
            return HttpResponse(html)

        # Build selector cards
        cards = ""
        for obj in objects_list:
            # Get display info based on model type
            if model == Client:
                title = obj.name
                subtitle = f"TRN: {obj.vat_number or 'N/A'}"
                badge = f"{obj.projects.count()} projects"
                icon = "👤"
            elif model == Project:
                title = f"{obj.project_id_code} — {obj.project_name}"
                subtitle = f"Client: {obj.client.name}"
                badge = f"AED {obj.po_amount:,.2f}"
                icon = "🏗️"
            elif model == Invoice:
                title = str(obj)
                subtitle = f"Project: {obj.project.project_name}"
                badge = obj.status
                icon = "📄"
                if badge == 'Paid':
                    badge_color = 'badge-success'
                elif badge == 'Approved':
                    badge_color = 'badge-info'
                else:
                    badge_color = 'badge-warning'
            elif model == PayrollRecord:
                title = obj.employee.name
                subtitle = f"Month: {obj.month.strftime('%b %Y')}"
                badge = obj.employee.get_payment_type_display()
                icon = "⏱️"
            else:
                title = str(obj)
                subtitle = ""
                badge = ""
                icon = "📊"

            url = reverse(url_name, args=[obj.pk])

            # Determine badge color
            if model == Invoice:
                badge_html = f'<span class="badge {badge_color}">{badge}</span>'
            else:
                badge_html = f'<span class="badge badge-primary">{badge}</span>'

            cards += f"""
            <a href="{url}" class="selector-card" target="_blank">
                <div class="selector-icon">{icon}</div>
                <div class="selector-info">
                    <div class="selector-title">{title}</div>
                    <div class="selector-subtitle">{subtitle}</div>
                </div>
                <div class="selector-badge">{badge_html}</div>
                <div class="selector-arrow">›</div>
            </a>
            """

        html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
{self.SHARED_CSS}
<style>
    .selector-container {{ max-width: 900px; margin: 0 auto; padding: 20px; }}
    .selector-header {{ text-align: center; margin-bottom: 30px; padding-bottom: 20px; border-bottom: 3px solid #1a237e; }}
    .selector-header h1 {{ font-size: 22px; color: #1a237e; margin-bottom: 8px; }}
    .selector-header p {{ color: #6c757d; font-size: 12px; }}
    .selector-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(350px, 1fr)); gap: 12px; }}
    .selector-card {{ 
        display: flex; align-items: center; gap: 12px; 
        padding: 14px 16px; background: #fff; border: 1px solid #dee2e6; 
        border-radius: 8px; text-decoration: none; color: inherit;
        transition: all 0.2s ease; box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    }}
    .selector-card:hover {{ 
        border-color: #1a237e; box-shadow: 0 4px 12px rgba(26,35,126,0.15); 
        transform: translateY(-1px);
    }}
    .selector-icon {{ font-size: 24px; flex-shrink: 0; }}
    .selector-info {{ flex: 1; min-width: 0; }}
    .selector-title {{ font-size: 13px; font-weight: 600; color: #212529; margin-bottom: 3px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
    .selector-subtitle {{ font-size: 10px; color: #6c757d; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
    .selector-badge {{ flex-shrink: 0; }}
    .selector-arrow {{ font-size: 20px; color: #1a237e; font-weight: 300; margin-left: 4px; }}
    .selector-count {{ 
        text-align: center; margin-bottom: 20px; 
        background: #e8eaf6; padding: 10px; border-radius: 6px; 
        font-size: 12px; color: #1a237e; font-weight: 600;
    }}
    .back-link {{ 
        display: inline-block; margin-bottom: 20px; 
        color: #1a237e; text-decoration: none; font-size: 12px; font-weight: 600;
    }}
    .back-link:hover {{ text-decoration: underline; }}
</style>
</head>
<body>
    <div class="report-container">
        <div class="selector-container">
            <a href="{reverse('admin:billing_reportsdashboard_changelist')}" class="back-link">← Back to Reports Dashboard</a>
            <div class="selector-header">
                <h1>{report_title}</h1>
                <p>Select a {model._meta.verbose_name} to generate the report</p>
            </div>
            <div class="selector-count">
                {len(objects_list)} {model._meta.verbose_name_plural} available
            </div>
            <div class="selector-grid">
                {cards}
            </div>
        </div>
    </div>
</body>
</html>"""
        return HttpResponse(html)

# =============================================================================
# COMPANY PROFILE ADMIN
# =============================================================================

@admin.register(CompanyProfile)
class CompanyProfileAdmin(ProfessionalReportMixin, admin.ModelAdmin):
    list_display = ["company_name", "trn_number", "is_active", "logo_preview"]
    list_filter = ["is_active"]
    search_fields = ["company_name", "trn_number"]

    def logo_preview(self, obj):
        if obj.logo:
            return format_html('<img src="{}" style="max-height:80px; max-width:!80px; border-radius:4px;" />', obj.logo.url)
        return "—"
    logo_preview.short_description = "Logo"
# =============================================================================
# CLIENT ADMIN
# =============================================================================

@admin.register(Client)
class ClientAdmin(ProfessionalReportMixin, CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'company'
    list_display = ["name", "contact_person", "vat_number", "company", "statement_button", "outstanding_button",
                    "progress_button", "deductions_button"]
    list_filter = ["company"]

    def statement_button(self, obj):
        url = reverse('admin:client_statement', args=[obj.pk])
        return format_html(
            '<a class="button" href="{}" target="_blank" style="background:#1a237e; color:white; padding:3px 10px; border-radius:4px; font-size:10px; text-decoration:none; font-weight:600;">📄 Statement</a>',
            url)

    statement_button.short_description = "Report"

    def outstanding_button(self, obj):
        url = reverse('admin:client_outstanding', args=[obj.pk])
        return format_html(
            '<a class="button" href="{}" target="_blank" style="background:#c62828; color:white; padding:3px 10px; border-radius:4px; font-size:10px; text-decoration:none; font-weight:600;">⚠️ Outstanding</a>',
            url)

    outstanding_button.short_description = "Report"

    def progress_button(self, obj):
        url = reverse('admin:client_progress', args=[obj.pk])
        return format_html(
            '<a class="button" href="{}" target="_blank" style="background:#00695c; color:white; padding:3px 10px; border-radius:4px; font-size:10px; text-decoration:none; font-weight:600;">📈 Progress</a>',
            url)

    progress_button.short_description = "Report"

    def deductions_button(self, obj):
        url = reverse('admin:client_deductions', args=[obj.pk])
        return format_html(
            '<a class="button" href="{}" target="_blank" style="background:#6a1b9a; color:white; padding:3px 10px; border-radius:4px; font-size:10px; text-decoration:none; font-weight:600;">📦 Deductions</a>',
            url)

    deductions_button.short_description = "Report"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('<int:pk>/statement/', self.admin_site.admin_view(self.statement_view), name='client_statement'),
            path('<int:pk>/outstanding/', self.admin_site.admin_view(self.outstanding_view), name='client_outstanding'),
            path('<int:pk>/progress/', self.admin_site.admin_view(self.progress_view), name='client_progress'),
            path('<int:pk>/deductions/', self.admin_site.admin_view(self.client_deductions_view),
                 name='client_deductions'),  # NEW
        ]
        return custom + urls

    # ========== PROFESSIONAL STATEMENT REPORT ==========
    def statement_view(self, request, pk):
        company = self.get_active_company(request)
        client = self.get_object_or_404_scoped(request, Client, pk=pk)
        logo_url = company.logo.url if company and company.logo else ''

        base_invoice_filter = {'project__client': client}
        if company:
            base_invoice_filter['project__company'] = company

        tax_invoices = Invoice.objects.filter(
            **base_invoice_filter, inv_type='T'
        ).exclude(is_advance_invoice=True).select_related('project').order_by('project__project_id_code', 'date',
                                                                              'inv_number')

        proforma_invoices = Invoice.objects.filter(
            **base_invoice_filter, inv_type='P'
        ).exclude(is_advance_invoice=True).select_related('project').order_by('project__project_id_code', 'date',
                                                                              'inv_number')

        from collections import OrderedDict
        projects_data = OrderedDict()
        for inv in tax_invoices:
            proj = inv.project
            if proj not in projects_data:
                projects_data[proj] = []
            projects_data[proj].append(inv)

        project_sections = ""
        grand_total_debit = Decimal("0")
        grand_total_credit = Decimal("0")
        grand_total_balance = Decimal("0")
        grand_total_vat = Decimal("0")
        grand_total_gross = Decimal("0")

        for proj, invoices_list in projects_data.items():
            rows = ""
            proj_debit = Decimal("0")
            proj_credit = Decimal("0")
            proj_balance = Decimal("0")
            proj_vat = Decimal("0")
            proj_gross = Decimal("0")
            payment_terms = proj.payment_terms or 30

            for inv in invoices_list:
                debit = inv.current_certified_net_before_vat
                credit = debit if inv.status == 'Paid' else Decimal("0.00")
                inv_balance = debit - credit

                if inv_balance > 0:
                    vat = inv.vat_amount
                    gross = inv.total_with_vat
                    vat_display = f"{vat:,.2f}"
                    gross_display = f"{gross:,.2f}"
                else:
                    vat = Decimal("0")
                    gross = Decimal("0")
                    vat_display = "—"
                    gross_display = "—"

                payment_date = inv.payment_date.strftime('%d-%b-%Y') if inv.payment_date else '—'
                due_date = inv.date + timedelta(days=payment_terms)
                due_display = due_date.strftime('%d-%b-%Y')

                # Age calculation
                if inv.status == 'Paid':
                    if inv.payment_date and due_date:
                        age_days = (inv.payment_date - due_date).days
                        if age_days > 0:
                            age_display = f'<span class="badge badge-danger">{age_days} days late</span>'
                        elif age_days < 0:
                            age_display = f'<span class="badge badge-success">{abs(age_days)} days early</span>'
                        else:
                            age_display = '<span class="badge badge-success">On time</span>'
                    else:
                        age_display = '—'
                else:
                    days_to_due = (due_date - date.today()).days
                    if days_to_due > 0:
                        age_display = f'<span class="badge badge-info">{days_to_due} days remaining</span>'
                    elif days_to_due < 0:
                        age_display = f'<span class="badge badge-danger">{abs(days_to_due)} days overdue</span>'
                    else:
                        age_display = '<span class="badge badge-warning">Due today</span>'

                proj_debit += debit
                proj_credit += credit
                proj_balance += inv_balance
                proj_vat += vat
                proj_gross += gross

                status_badge = '<span class="badge badge-success">Paid</span>' if inv.status == 'Paid' else '<span class="badge badge-warning">Pending</span>'

                rows += f"""<<tr>
                    <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                    <td class="text"><b>{inv}</b></td>
                    <td class="center">{status_badge}</td>
                    <td class="center">{due_display}</td>
                    <td class="num font-bold text-primary">{debit:,.2f}</td>
                    <td class="num text-success">{credit:,.2f}</td>
                    <td class="center">{payment_date}</td>
                    <td class="num font-bold">{inv_balance:,.2f}</td>
                    <td class="num">{vat_display}</td>
                    <td class="num">{gross_display}</td>
                    <td class="center">{age_display}</td>
                </tr>"""

            grand_total_debit += proj_debit
            grand_total_credit += proj_credit
            grand_total_balance += proj_balance
            grand_total_vat += proj_vat
            grand_total_gross += proj_gross

            project_sections += f"""
            <div class="card" style="page-break-inside: avoid;">
                <div class="card-header">
                    <span class="icon">🏗️</span> {proj.project_id_code} — {proj.project_name}
                </div>
                <div class="meta-grid" style="margin-bottom:12px;">
                    <div class="meta-item"><div class="meta-label">PO Number</div><div class="meta-value">{proj.po_number or 'N/A'}</div></div>
                    <div class="meta-item"><div class="meta-label">PO Amount</div><div class="meta-value">AED {proj.po_amount:,.2f}</div></div>
                    <div class="meta-item"><div class="meta-label">Payment Terms</div><div class="meta-value">{payment_terms} days</div></div>
                </div>
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Date</th>
                            <th>Invoice #</th>
                            <th>Status</th>
                            <th>Due Date</th>
                            <th class="num">Debit</th>
                            <th class="num">Credit</th>
                            <th>Payment Date</th>
                            <th class="num">Balance</th>
                            <th class="num">VAT</th>
                            <th class="num">Total w/ VAT</th>
                            <th>Aging</th>
                        </tr>
                    </thead>
                    <tbody>{rows}</tbody>
                    <tfoot>
                        <tr class="total-row">
                            <td colspan="4"><b>PROJECT SUBTOTAL</b></td>
                            <td class="num text-primary"><b>{proj_debit:,.2f}</b></td>
                            <td class="num text-success"><b>{proj_credit:,.2f}</b></td>
                            <td></td>
                            <td class="num"><b>{proj_balance:,.2f}</b></td>
                            <td class="num"><b>{proj_vat:,.2f}</b></td>
                            <td class="num"><b>{proj_gross:,.2f}</b></td>
                            <td></td>
                        </tr>
                    </tfoot>
                </table>
            </div>
            """

        grand_vat_display = f"{grand_total_vat:,.2f}" if grand_total_vat > 0 else "—"
        grand_gross_display = f"{grand_total_gross:,.2f}" if grand_total_gross > 0 else "—"

        proforma_section = ""
        if proforma_invoices.exists():
            prof_rows = ""
            for inv in proforma_invoices:
                net = inv.current_net_before_vat
                vat = inv.vat_amount
                total = inv.total_with_vat
                prof_age_days = (date.today() - inv.date).days
                prof_rows += f"""<<tr>
                    <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                    <td class="text"><b>{inv}</b></td>
                    <td class="text">{inv.project.project_id_code} — {inv.project.project_name}</td>
                    <td class="center"><span class="badge badge-warning">{inv.status}</span></td>
                    <td class="num font-bold text-primary">{net:,.2f}</td>
                    <td class="num">{vat:,.2f}</td>
                    <td class="num font-bold">{total:,.2f}</td>
                    <td class="center"><span class="badge badge-info">{prof_age_days} days</span></td>
                </tr>"""

            proforma_section = f"""
            <div class="card" style="margin-top:20px;">
                <div class="card-header">
                    <span class="icon">📋</span> Proforma Invoices 
                    <span class="badge badge-warning" style="margin-left:10px;">NOT INCLUDED IN BALANCES</span>
                </div>
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Date</th>
                            <th>Invoice #</th>
                            <th>Project</th>
                            <th>Status</th>
                            <th class="num">Net</th>
                            <th class="num">VAT</th>
                            <th class="num">Total</th>
                            <th>Age</th>
                        </tr>
                    </thead>
                    <tbody>{prof_rows}</tbody>
                </table>
            </div>
            """

        body = f"""
        {self._build_meta_grid({
            'Client': client.name,
            'TRN': client.vat_number or 'N/A',
            'Report Date': date.today().strftime('%d-%b-%Y'),
            'Company': company.company_name if company else 'N/A'
        })}

        {project_sections if project_sections else '<div class="card" style="text-align:center; padding:40px; color:#666;"><p>No tax invoices found for this client.</p></div>'}

        {f"""
        <div class="card" style="background: #1a237e; color: white; margin-top: 20px;">
            <div class="card-header" style="color: white; border-color: rgba(255,255,255,0.3);">
                <span class="icon">💰</span> GRAND TOTAL ACROSS ALL PROJECTS
            </div>
            <div class="grid-4">
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {grand_total_debit:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Total Debit</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {grand_total_credit:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Total Credit</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {grand_total_balance:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Total Balance</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {grand_gross_display}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Total w/ VAT</div>
                </div>
            </div>
        </div>
        """ if project_sections else ""}

        {proforma_section}
        """

        return HttpResponse(self._report_base_wrapper(
            "STATEMENT OF ACCOUNT",
            f"Client Ledger — {client.name}",
            body,
            logo_url
        ))

    # ========== PROFESSIONAL OUTSTANDING REPORT ==========
    def outstanding_view(self, request, pk):
        company = self.get_active_company(request)
        client = self.get_object_or_404_scoped(request, Client, pk=pk)
        logo_url = company.logo.url if company and company.logo else ''

        base_filter = {'project__client': client}
        if company:
            base_filter['project__company'] = company

        invoices = Invoice.objects.filter(
            **base_filter
        ).exclude(status='Paid').select_related('project').order_by('date')

        draft_rows = ""
        approved_rows = ""
        total_draft = Decimal("0")
        total_approved = Decimal("0")

        for inv in invoices:
            gross = inv.total_with_vat
            days = (date.today() - inv.date).days
            status_badge = '<span class="badge badge-warning">Draft</span>' if inv.status == 'Draft' else '<span class="badge badge-danger">Approved</span>'

            row = f"""<<tr>
                <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                <td class="text"><b>{inv}</b></td>
                <td class="text">{inv.project.project_name}</td>
                <td class="center">{inv.get_inv_type_display()}</td>
                <td class="center">{status_badge}</td>
                <td class="num font-bold text-primary">{gross:,.2f}</td>
                <td class="center"><span class="badge badge-danger">{days} days</span></td>
            </tr>"""

            if inv.status == 'Draft':
                draft_rows += row
                total_draft += gross
            else:
                approved_rows += row
                total_approved += gross

        body = f"""
        {self._build_meta_grid({
            'Client': client.name,
            'TRN': client.vat_number or 'N/A',
            'Report Date': date.today().strftime('%d-%b-%Y'),
            'Total Outstanding': f"AED {(total_draft + total_approved):,.2f}"
        })}

        <div class="card">
            <div class="card-header"><span class="icon">📝</span> Draft Invoices</div>
            <table class="data-table">
                <thead>
                    <tr><th>Date</th><th>Invoice #</th><th>Project</th><th>Type</th><th>Status</th><th class="num">Amount</th><th>Age</th></tr>
                </thead>
                <tbody>{draft_rows or '<tr><td colspan="7" style="text-align:center;color:#999;padding:20px;">No draft invoices</td></tr>'}</tbody>
                <tfoot>
                    <tr class="total-row">
                        <td colspan="5"><b>DRAFT TOTAL</b></td>
                        <td class="num"><b>{total_draft:,.2f}</b></td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <div class="card" style="margin-top:16px;">
            <div class="card-header"><span class="icon">✅</span> Approved / Sent Invoices</div>
            <table class="data-table">
                <thead>
                    <tr><th>Date</th><th>Invoice #</th><th>Project</th><th>Type</th><th>Status</th><th class="num">Amount</th><th>Age</th></tr>
                </thead>
                <tbody>{approved_rows or '<tr><td colspan="7" style="text-align:center;color:#999;padding:20px;">No approved invoices</td></tr>'}</tbody>
                <tfoot>
                    <tr class="total-row">
                        <td colspan="5"><b>APPROVED TOTAL</b></td>
                        <td class="num"><b>{total_approved:,.2f}</b></td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <div class="card" style="background: #1a237e; color: white; margin-top: 20px;">
            <div style="display: flex; justify-content: space-between; align-items: center; padding: 16px;">
                <div style="font-size: 14px; font-weight: 700;">GRAND TOTAL OUTSTANDING</div>
                <div style="font-size: 24px; font-weight: 700;">AED {(total_draft + total_approved):,.2f}</div>
            </div>
        </div>
        """

        return HttpResponse(self._report_base_wrapper(
            "OUTSTANDING INVOICES REPORT",
            f"Unpaid Invoices — {client.name}",
            body,
            logo_url
        ))

    # ========== PROFESSIONAL PROGRESS REPORT ==========
    def progress_view(self, request, pk):
        company = self.get_active_company(request)
        client = self.get_object_or_404_scoped(request, Client, pk=pk)
        logo_url = company.logo.url if company and company.logo else ''

        projects_filter = {'client': client}
        if company:
            projects_filter['company'] = company
        projects = Project.objects.filter(**projects_filter).prefetch_related('invoices', 'boq_items')

        cards = ""
        for proj in projects:
            inv_filter = {'project': proj, 'is_advance_invoice': False}
            if company:
                inv_filter['project__company'] = company

            latest_inv = Invoice.objects.filter(
                **inv_filter
            ).filter(
                Q(inv_type='T') | Q(inv_type='P', status='Approved')
            ).order_by('-inv_number').first()

            work_done = latest_inv.certified_work_done if latest_inv else Decimal("0")
            # Include variations in PO amount for progress calculation
            variation_amount = getattr(proj, 'variation_amount', Decimal("0"))
            amended_po = proj.po_amount + variation_amount
            balance = money(amended_po - work_done)
            progress_pct = (work_done / amended_po * 100) if amended_po > 0 else Decimal("0")

            ret_a_cum = latest_inv.cumulative_retention_a if latest_inv else Decimal("0")
            ret_b_cum = latest_inv.cumulative_retention_b if latest_inv else Decimal("0")
            ret_a_rec = latest_inv.cumulative_retention_a_recovered if latest_inv else Decimal("0")
            ret_b_rec = latest_inv.cumulative_retention_b_recovered if latest_inv else Decimal("0")
            net_ret = money((ret_a_cum + ret_b_cum) - (ret_a_rec + ret_b_rec))

            adv_rec = latest_inv.cumulative_advance_recovered if latest_inv else Decimal("0")
            adv_total = proj.total_advance_value

            # Color coding based on progress
            # Color coding based on progress - MATCH ANALYTICS REPORT
            progress_color = "linear-gradient(90deg, #447e9b, #2e7d32)"

            # Variation display
            var_display = f'<span style="color:#6a1b9a; font-size:8px;"> (+{variation_amount:,.2f} var)</span>' if variation_amount > 0 else ''

            cards += f"""
            <div class="card" style="page-break-inside: avoid;">
                <div class="card-header" style="display: flex; justify-content: space-between; align-items: center;">
                    <span><span class="icon">🏗️</span> {proj.project_id_code} — {proj.project_name}</span>
                    <span class="badge badge-primary">{progress_pct:.1f}% Complete</span>
                </div>

                <div class="grid-4" style="margin-bottom: 16px;">
                    <div class="metric-card">
                        <div class="metric-value text-primary">AED {amended_po:,.2f}{var_display}</div>
                        <div class="metric-label">PO Amount</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value text-success">AED {work_done:,.2f}</div>
                        <div class="metric-label">Work Done</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value text-danger">AED {balance:,.2f}</div>
                        <div class="metric-label">Balance</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value text-warning">AED {net_ret:,.2f}</div>
                        <div class="metric-label">Net Retention</div>
                    </div>
                </div>

                <div style="margin-bottom: 16px;">
                    <div style="display: flex; justify-content: space-between; font-size: 9px; margin-bottom: 4px; color: var(--text-muted);">
                        <span>Progress</span>
                        <span class="font-bold">{progress_pct:.1f}%</span>
                    </div>
                    <div class="progress-bar">
                        <div class="progress-fill" style="width: {min(float(progress_pct), 100)}%; background: {progress_color};">
                            {progress_pct:.0f}%
                        </div>
                    </div>
                </div>

                <div class="grid-3">
                    <div class="card" style="padding: 12px; margin-bottom: 0;">
                        <div style="font-size: 10px; font-weight: 700; color: var(--primary); margin-bottom: 8px; text-align: center; border-bottom: 1px solid var(--border); padding-bottom: 6px;">
                            Retention A ({proj.retention_a_percent}%)
                        </div>
                        <div style="display: flex; justify-content: space-between; font-size: 9px; margin-bottom: 4px;">
                            <span class="text-muted">Deducted:</span>
                            <span class="font-bold">AED {ret_a_cum:,.2f}</span>
                        </div>
                        <div style="display: flex; justify-content: space-between; font-size: 9px; margin-bottom: 4px;">
                            <span class="text-muted">Recovered:</span>
                            <span class="font-bold text-success">AED {ret_a_rec:,.2f}</span>
                        </div>
                        <div style="display: flex; justify-content: space-between; font-size: 9px; border-top: 1px solid var(--border); padding-top: 4px; margin-top: 4px;">
                            <span class="text-muted">Held:</span>
                            <span class="font-bold text-warning">AED {money(ret_a_cum - ret_a_rec):,.2f}</span>
                        </div>
                    </div>

                    <div class="card" style="padding: 12px; margin-bottom: 0;">
                        <div style="font-size: 10px; font-weight: 700; color: var(--primary); margin-bottom: 8px; text-align: center; border-bottom: 1px solid var(--border); padding-bottom: 6px;">
                            Retention B ({proj.retention_b_percent}%)
                        </div>
                        <div style="display: flex; justify-content: space-between; font-size: 9px; margin-bottom: 4px;">
                            <span class="text-muted">Deducted:</span>
                            <span class="font-bold">AED {ret_b_cum:,.2f}</span>
                        </div>
                        <div style="display: flex; justify-content: space-between; font-size: 9px; margin-bottom: 4px;">
                            <span class="text-muted">Recovered:</span>
                            <span class="font-bold text-success">AED {ret_b_rec:,.2f}</span>
                        </div>
                        <div style="display: flex; justify-content: space-between; font-size: 9px; border-top: 1px solid var(--border); padding-top: 4px; margin-top: 4px;">
                            <span class="text-muted">Held:</span>
                            <span class="font-bold text-warning">AED {money(ret_b_cum - ret_b_rec):,.2f}</span>
                        </div>
                    </div>

                    <div class="card" style="padding: 12px; margin-bottom: 0;">
                        <div style="font-size: 10px; font-weight: 700; color: var(--primary); margin-bottom: 8px; text-align: center; border-bottom: 1px solid var(--border); padding-bottom: 6px;">
                            Advance ({proj.advance_percent}%)
                        </div>
                        <div style="display: flex; justify-content: space-between; font-size: 9px; margin-bottom: 4px;">
                            <span class="text-muted">Taken:</span>
                            <span class="font-bold">AED {adv_total:,.2f}</span>
                        </div>
                        <div style="display: flex; justify-content: space-between; font-size: 9px; margin-bottom: 4px;">
                            <span class="text-muted">Recovered:</span>
                            <span class="font-bold text-success">AED {adv_rec:,.2f}</span>
                        </div>
                        <div style="display: flex; justify-content: space-between; font-size: 9px; border-top: 1px solid var(--border); padding-top: 4px; margin-top: 4px;">
                            <span class="text-muted">Balance:</span>
                            <span class="font-bold text-warning">AED {money(adv_total - adv_rec):,.2f}</span>
                        </div>
                    </div>
                </div>
            </div>
            """

        body = f"""
        {self._build_meta_grid({
            'Client': client.name,
            'TRN': client.vat_number or 'N/A',
            'Report Date': date.today().strftime('%d-%b-%Y'),
            'Active Projects': projects.count()
        })}

        <div style="display: flex; flex-direction: column; gap: 16px;">
            {cards if cards else '<div class="card" style="text-align:center; padding:40px; color:#666;"><p>No projects found for this client.</p></div>'}
        </div>
        """

        return HttpResponse(self._report_base_wrapper(
            "CLIENT PROJECT PROGRESS",
            f"Progress Overview — {client.name}",
            body,
            logo_url
        ))

    # =============================================================================
    # CLIENT DEDUCTIONS REPORT (New)
    # =============================================================================

    def client_deductions_view(self, request, pk):
        """
        Professional Client Deductions Report - tracks all deductions across all projects
        per client including: materials supplied by client, retention A/B, advance recovery,
        back charges, liquidated damages, and estimated back charges.
        """
        company = self.get_active_company(request)
        client = self.get_object_or_404_scoped(request, Client, pk=pk)
        logo_url = company.logo.url if company and company.logo else ''

        base_filter = {'project__client': client}
        if company:
            base_filter['project__company'] = company

        # Get all projects for this client
        projects = Project.objects.filter(
            **{k.replace('project__', ''): v for k, v in base_filter.items() if k.startswith('project__')})
        if company:
            projects = projects.filter(Q(company=company) | Q(company__isnull=True))

        # Build project-by-project deduction breakdown
        project_sections = ""
        grand_materials = Decimal("0")
        grand_ret_a = Decimal("0")
        grand_ret_b = Decimal("0")
        grand_ret_a_rec = Decimal("0")
        grand_ret_b_rec = Decimal("0")
        grand_advance = Decimal("0")
        grand_advance_rec = Decimal("0")
        grand_back_charges = Decimal("0")
        grand_est_back_charges = Decimal("0")
        grand_liquidated = Decimal("0")
        grand_net_deductions = Decimal("0")

        for proj in projects:
            invoices = Invoice.objects.filter(
                project=proj, inv_type='T'
            ).exclude(is_advance_invoice=True).order_by('inv_number')

            latest_inv = invoices.order_by('-inv_number').first()

            # Materials supplied by client per invoice
            materials_rows = ""
            proj_materials = Decimal("0")
            for inv in invoices:
                mat = inv.material_supplied_by_client or Decimal("0")
                if mat > 0:
                    proj_materials += mat
                    materials_rows += f"""
                    <tr>
                        <td class="center">{inv.inv_number}</td>
                        <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                        <td class="text"><b>{inv}</b></td>
                        <td class="num font-bold text-danger">({mat:,.2f})</td>
                        <td class="text">Materials supplied by client — {inv.project.project_name}</td>
                    </tr>
                    """

            # Retention & Advance tracking
            ret_a_cum = latest_inv.cumulative_retention_a if latest_inv else Decimal("0")
            ret_b_cum = latest_inv.cumulative_retention_b if latest_inv else Decimal("0")
            ret_a_rec = latest_inv.cumulative_retention_a_recovered if latest_inv else Decimal("0")
            ret_b_rec = latest_inv.cumulative_retention_b_recovered if latest_inv else Decimal("0")
            advance_rec = latest_inv.cumulative_advance_recovered if latest_inv else Decimal("0")
            advance_total = proj.total_advance_value

            # Project-level deductions from model fields
            back_charges = getattr(proj, 'back_charges', Decimal("0"))
            est_back_charges = getattr(proj, 'estimated_back_charges', Decimal("0"))
            liquidated = getattr(proj, 'liquidated_damages', Decimal("0"))

            # Net deductions = materials + retention held + advance recovered + back charges + liquidated
            net_ret_held = (ret_a_cum + ret_b_cum) - (ret_a_rec + ret_b_rec)
            proj_net_deductions = proj_materials + net_ret_held + advance_rec + back_charges + est_back_charges + liquidated

            # Update grand totals
            grand_materials += proj_materials
            grand_ret_a += ret_a_cum
            grand_ret_b += ret_b_cum
            grand_ret_a_rec += ret_a_rec
            grand_ret_b_rec += ret_b_rec
            grand_advance += advance_total
            grand_advance_rec += advance_rec
            grand_back_charges += back_charges
            grand_est_back_charges += est_back_charges
            grand_liquidated += liquidated
            grand_net_deductions += proj_net_deductions

            # Invoice-level deduction detail rows
            inv_deduction_rows = ""
            for inv in invoices:
                curr_ret_a = inv.current_retention_a
                curr_ret_b = inv.current_retention_b
                curr_adv = inv.current_advance_recovery
                curr_mat = inv.material_supplied_by_client or Decimal("0")
                curr_net_ded = curr_ret_a + curr_ret_b + curr_adv + curr_mat

                if curr_net_ded > 0:
                    inv_deduction_rows += f"""
                    <tr>
                        <td class="center">{inv.inv_number}</td>
                        <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                        <td class="num text-danger">({curr_ret_a:,.2f})</td>
                        <td class="num text-danger">({curr_ret_b:,.2f})</td>
                        <td class="num text-danger">({curr_adv:,.2f})</td>
                        <td class="num text-danger">({curr_mat:,.2f})</td>
                        <td class="num font-bold text-danger">({curr_net_ded:,.2f})</td>
                    </tr>
                    """

            project_sections += f"""
            <div class="card" style="page-break-inside: avoid; margin-bottom: 20px;">
                <div class="card-header" style="display: flex; justify-content: space-between; align-items: center;">
                    <span><span class="icon">🏗️</span> {proj.project_id_code} — {proj.project_name}</span>
                    <span class="badge badge-danger">Net Deductions: AED {proj_net_deductions:,.2f}</span>
                </div>

                <div class="meta-grid" style="margin-bottom: 12px;">
                    <div class="meta-item"><div class="meta-label">PO Amount</div><div class="meta-value">AED {proj.po_amount:,.2f}</div></div>
                    <div class="meta-item"><div class="meta-label">Advance %</div><div class="meta-value">{proj.advance_percent}%</div></div>
                    <div class="meta-item"><div class="meta-label">Retention A %</div><div class="meta-value">{proj.retention_a_percent}%</div></div>
                    <div class="meta-item"><div class="meta-label">Retention B %</div><div class="meta-value">{proj.retention_b_percent}%</div></div>
                </div>

                <!-- Materials Supplied Section -->
                <div style="margin-bottom: 16px;">
                    <div style="font-size: 11px; font-weight: 700; color: #c62828; margin-bottom: 8px; padding-bottom: 6px; border-bottom: 2px solid #ffcdd2;">
                        <span class="icon">📦</span> Materials Supplied by Client
                    </div>
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>Inv #</th>
                                <th>Date</th>
                                <th>Invoice Reference</th>
                                <th class="num">Amount</th>
                                <th>Notes</th>
                            </tr>
                        </thead>
                        <tbody>
                            {materials_rows if materials_rows else '<tr><td colspan="5" style="text-align:center; color:#999; padding:12px;">No materials recorded</td></tr>'}
                        </tbody>
                        <tfoot>
                            <tr class="total-row">
                                <td colspan="3"><b>PROJECT MATERIALS TOTAL</b></td>
                                <td class="num text-danger"><b>({proj_materials:,.2f})</b></td>
                                <td></td>
                            </tr>
                        </tfoot>
                    </table>
                </div>

                <!-- Per-Invoice Deductions Breakdown -->
                <div style="margin-bottom: 16px;">
                    <div style="font-size: 11px; font-weight: 700; color: #1a237e; margin-bottom: 8px; padding-bottom: 6px; border-bottom: 2px solid #c5cae9;">
                        <span class="icon">📋</span> Deductions by Invoice
                    </div>
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>Inv #</th>
                                <th>Date</th>
                                <th class="num">Ret A</th>
                                <th class="num">Ret B</th>
                                <th class="num">Adv Rec</th>
                                <th class="num">Materials</th>
                                <th class="num">Total Deduction</th>
                            </tr>
                        </thead>
                        <tbody>
                            {inv_deduction_rows if inv_deduction_rows else '<tr><td colspan="7" style="text-align:center; color:#999; padding:12px;">No deduction invoices</td></tr>'}
                        </tbody>
                    </table>
                </div>

                <!-- Cumulative Deductions Summary -->
                <div class="grid-3" style="margin-top: 12px;">
                    <div class="metric-card" style="border-left: 4px solid #c62828;">
                        <div class="metric-value text-danger">AED {net_ret_held:,.2f}</div>
                        <div class="metric-label">Net Retention Held</div>
                        <div style="font-size: 8px; color: #666; margin-top: 4px;">
                            A: {money(ret_a_cum - ret_a_rec):,.2f} | B: {money(ret_b_cum - ret_b_rec):,.2f}
                        </div>
                    </div>
                    <div class="metric-card" style="border-left: 4px solid #ed6c02;">
                        <div class="metric-value text-warning">AED {advance_rec:,.2f}</div>
                        <div class="metric-label">Advance Recovered</div>
                        <div style="font-size: 8px; color: #666; margin-top: 4px;">
                            of {advance_total:,.2f} total
                        </div>
                    </div>
                    <div class="metric-card" style="border-left: 4px solid #6a1b9a;">
                        <div class="metric-value" style="color: #6a1b9a;">AED {back_charges + est_back_charges + liquidated:,.2f}</div>
                        <div class="metric-label">Other Deductions</div>
                        <div style="font-size: 8px; color: #666; margin-top: 4px;">
                            Back: {back_charges:,.2f} | Est: {est_back_charges:,.2f} | Liq: {liquidated:,.2f}
                        </div>
                    </div>
                </div>
            </div>
            """

        # Grand totals summary
        grand_net_ret = (grand_ret_a + grand_ret_b) - (grand_ret_a_rec + grand_ret_b_rec)
        grand_other = grand_back_charges + grand_est_back_charges + grand_liquidated

        body = f"""
        {self._build_meta_grid({
            'Client': client.name,
            'TRN': client.vat_number or 'N/A',
            'Report Date': date.today().strftime('%d-%b-%Y'),
            'Company': company.company_name if company else 'N/A',
            'Active Projects': projects.count()
        })}

        <!-- Grand Summary Dashboard -->
        <div class="card" style="background: #1a237e; color: white; margin-bottom: 20px;">
            <div class="card-header" style="color: white; border-color: rgba(255,255,255,0.3);">
                <span class="icon">💰</span> CLIENT DEDUCTIONS SUMMARY
            </div>
            <div class="grid-4">
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {grand_materials:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Materials Supplied</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {grand_net_ret:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Net Retention Held</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {grand_advance_rec:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Advance Recovered</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {grand_other:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Other Deductions</div>
                </div>
            </div>
            <div style="text-align: center; margin-top: 16px; padding-top: 16px; border-top: 1px solid rgba(255,255,255,0.2);">
                <div style="font-size: 11px; text-transform: uppercase; letter-spacing: 1px; opacity: 0.8; margin-bottom: 8px;">Total Net Deductions Across All Projects</div>
                <div style="font-size: 24px; font-weight: 700;">AED {grand_net_deductions:,.2f}</div>
            </div>
        </div>

        {project_sections if project_sections else '<div class="card" style="text-align:center; padding:40px; color:#666;"><p>No projects found for this client.</p></div>'}

        <!-- Detailed Deductions Ledger -->
        <div class="card" style="margin-top: 20px;">
            <div class="card-header">
                <span class="icon">📊</span> CUMULATIVE DEDUCTIONS LEDGER
            </div>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Category</th>
                        <th class="num">Total Deducted</th>
                        <th class="num">Recovered / Offset</th>
                        <th class="num">Net Balance</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td><b>Materials Supplied by Client</b></td>
                        <td class="num text-danger">({grand_materials:,.2f})</td>
                        <td class="num text-success">—</td>
                        <td class="num font-bold text-danger">({grand_materials:,.2f})</td>
                        <td><span class="badge badge-danger">Permanent Deduction</span></td>
                    </tr>
                    <tr>
                        <td><b>Retention A</b></td>
                        <td class="num text-danger">({grand_ret_a:,.2f})</td>
                        <td class="num text-success">{grand_ret_a_rec:,.2f}</td>
                        <td class="num font-bold text-danger">({money(grand_ret_a - grand_ret_a_rec):,.2f})</td>
                        <td><span class="badge badge-warning">Held Until Recovered</span></td>
                    </tr>
                    <tr>
                        <td><b>Retention B</b></td>
                        <td class="num text-danger">({grand_ret_b:,.2f})</td>
                        <td class="num text-success">{grand_ret_b_rec:,.2f}</td>
                        <td class="num font-bold text-danger">({money(grand_ret_b - grand_ret_b_rec):,.2f})</td>
                        <td><span class="badge badge-warning">Held Until Recovered</span></td>
                    </tr>
                    <tr>
                        <td><b>Advance Recovery</b></td>
                        <td class="num text-danger">({grand_advance:,.2f})</td>
                        <td class="num text-success">{grand_advance_rec:,.2f}</td>
                        <td class="num font-bold text-danger">({money(grand_advance - grand_advance_rec):,.2f})</td>
                        <td><span class="badge badge-info">Recovering from Invoices</span></td>
                    </tr>
                    <tr>
                        <td><b>Back Charges / Contra-Charges</b></td>
                        <td class="num text-danger">({grand_back_charges:,.2f})</td>
                        <td class="num text-success">—</td>
                        <td class="num font-bold text-danger">({grand_back_charges:,.2f})</td>
                        <td><span class="badge badge-danger">Permanent Deduction</span></td>
                    </tr>
                    <tr>
                        <td><b>Estimated Back Charges</b></td>
                        <td class="num text-danger">({grand_est_back_charges:,.2f})</td>
                        <td class="num text-success">—</td>
                        <td class="num font-bold text-danger">({grand_est_back_charges:,.2f})</td>
                        <td><span class="badge badge-warning">Estimated / Provisional</span></td>
                    </tr>
                    <tr>
                        <td><b>Liquidated Damages</b></td>
                        <td class="num text-danger">({grand_liquidated:,.2f})</td>
                        <td class="num text-success">—</td>
                        <td class="num font-bold text-danger">({grand_liquidated:,.2f})</td>
                        <td><span class="badge badge-danger">Permanent Deduction</span></td>
                    </tr>
                </tbody>
                <tfoot>
                    <tr class="grand-total">
                        <td><b>GRAND TOTAL NET DEDUCTIONS</b></td>
                        <td class="num">—</td>
                        <td class="num">—</td>
                        <td class="num"><b>({grand_net_deductions:,.2f})</b></td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <!-- Signature Section -->
        <div class="signature-grid" style="margin-top: 40px;">
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">Finance Director</div>
            </div>
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">Technical Manager</div>
            </div>
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">General Manager</div>
            </div>
        </div>
        """

        return HttpResponse(self._report_base_wrapper(
            "CLIENT DEDUCTIONS REPORT",
            f"Complete Deductions Tracking — {client.name}",
            body,
            logo_url
        ))

# =============================================================================
# BOQ ITEM ADMIN
# =============================================================================

@admin.register(BOQItem)
class BOQItemAdmin(CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'project__company'
    search_fields = ["item_number", "description"]
    list_display = ["item_number", "description", "project", "fmt_qty", "fmt_rate", "fmt_total", "is_executed"]

    def fmt_qty(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.quantity:,.2f}</div>')
    fmt_qty.short_description = "Qty"

    def fmt_rate(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.rate:,.2f}</div>')
    fmt_rate.short_description = "Rate"

    def fmt_total(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.quantity * obj.rate:,.2f}</div>')
    fmt_total.short_description = "Total"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('get-by-project/', self.admin_site.admin_view(self.get_by_project), name='boqitem_get_by_project'),
        ]
        return custom + urls

    def get_by_project(self, request):
        from django.http import JsonResponse
        project_id = request.GET.get('project_id')
        if not project_id:
            return JsonResponse([], safe=False)

        company = self.get_active_company(request)
        qs = BOQItem.objects.filter(project_id=project_id)
        if company:
            qs = qs.filter(Q(project__company=company) | Q(project__company__isnull=True))

        items = qs.values('id', 'item_number', 'description')
        data = [{'id': item['id'], 'text': f"{item['item_number']} - {item['description'][:40]}"} for item in items]
        return JsonResponse(data, safe=False)


# =============================================================================
# PROJECT ADMIN
# =============================================================================
class BOQItemInline(TabularInlinePaginated):
    model = BOQItem
    extra = 1
    per_page = 20  # Show 20 BOQ items per page


class BOQItemInline(admin.TabularInline):
    model = BOQItem
    formset = LimitedInlineFormSet
    extra = 1
    ##max_num = 50
    fields = ["item_number", "description", "unit", "quantity", "rate"]

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'project':
            company = CompanyProfile.get_active(request)
            if company:
                kwargs['queryset'] = Project.objects.filter(company=company)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class ExpenseInlineForm(forms.ModelForm):
    class Meta:
        model = Expense
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        project = None
        if self.instance and self.instance.pk and self.instance.project:
            project = self.instance.project
        elif self.initial.get('project'):
            try:
                project = Project.objects.get(pk=self.initial['project'])
            except Project.DoesNotExist:
                pass
        if project:
            self.fields['boq_item'].queryset = BOQItem.objects.filter(project=project)
        else:
            self.fields['boq_item'].queryset = BOQItem.objects.none()
        if self.instance and self.instance.pk and self.instance.category:
            self.fields['sub_category'].queryset = SubExpense.objects.filter(parent=self.instance.category)
        else:
            self.fields['sub_category'].queryset = SubExpense.objects.none()


class ExpenseInline(admin.TabularInline):
    model = Expense
    form = ExpenseInlineForm
    formset = LimitedInlineFormSet
    extra = 0
    ##max_num = 50
    fields = ["date", "category", "sub_category", "amount", "boq_item", "description", "is_allocated"]

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        if obj:
            formset.parent_project = obj
        return formset

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'boq_item':
            if getattr(self, 'parent_project', None):
                kwargs['queryset'] = BOQItem.objects.filter(project=self.parent_project)
            else:
                kwargs['queryset'] = BOQItem.objects.none()
        elif db_field.name == 'category':
            company = CompanyProfile.get_active(request)
            if company:
                kwargs['queryset'] = ExpenseCategory.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
        elif db_field.name == 'sub_category':
            kwargs['queryset'] = SubExpense.objects.none()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(Project)
class ProjectAdmin(ProfessionalReportMixin,CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'company'
    inlines = [BOQItemInline, ExpenseInline]
    list_display = [
        "project_id_code", "action_buttons", "project_name", "client", "company", "fmt_po",
        "fmt_boq_total", "is_boq_complete", "fmt_advance",
        "fmt_ret_a_pct", "fmt_ret_b_pct"
    ]
    list_filter = ["company", "is_boq_complete"]
    readonly_fields = ["is_boq_complete"]
    search_fields = ["project_id_code", "project_name"]

    def action_buttons(self, obj):
        app_label = obj._meta.app_label
        invoices_url = reverse(f'admin:{app_label}_invoice_changelist') + f'?project__id__exact={obj.id}'
        analytics_url = reverse('admin:project_analytics', args=[obj.pk])
        cost_url = reverse('admin:project_cost_profit', args=[obj.pk])
        statement_url = reverse('admin:project_statement', args=[obj.pk])
        deductions_url = reverse('admin:project_deductions', args=[obj.pk])  # NEW

        return format_html(
            '<div style="display:flex; gap:4px; justify-content:center; flex-wrap:nowrap;">'
            '<a class="button" href="{}" style="background:#447e9b; color:white; padding:3px 8px; '
            'border-radius:3px; font-size:10px; font-weight:600; text-decoration:none; '
            'white-space:nowrap; line-height:1.4; border:none; min-width:50px; text-align:center;">Invoices</a>'
            '<a class="button" href="{}" target="_blank" style="background:#6a1b9a; color:white; padding:3px 8px; '
            'border-radius:3px; font-size:10px; font-weight:600; text-decoration:none; '
            'white-space:nowrap; line-height:1.4; border:none; min-width:50px; text-align:center;">Analytics</a>'
            '<a class="button" href="{}" target="_blank" style="background:#d32f2f; color:white; padding:3px 8px; '
            'border-radius:3px; font-size:10px; font-weight:600; text-decoration:none; '
            'white-space:nowrap; line-height:1.4; border:none; min-width:50px; text-align:center;">Cost &amp; P&amp;L</a>'
            '<a class="button" href="{}" target="_blank" style="background:#1a237e; color:white; padding:3px 8px; '
            'border-radius:3px; font-size:10px; font-weight:600; text-decoration:none; '
            'white-space:nowrap; line-height:1.4; border:none; min-width:50px; text-align:center;">Statement</a>'
            '<a class="button" href="{}" target="_blank" style="background:#6a1b9a; color:white; padding:3px 8px; '
            'border-radius:3px; font-size:10px; font-weight:600; text-decoration:none; '
            'white-space:nowrap; line-height:1.4; border:none; min-width:50px; text-align:center;">Deductions</a>'  # NEW
            '</div>',
            invoices_url, analytics_url, cost_url, statement_url, deductions_url  # Added deductions_url
        )
    action_buttons.short_description = "Actions"
    action_buttons.admin_order_field = "project_id_code"

    def _logo_bar(self, logo_url):
        if logo_url:
            return f'<div style="text-align:right; margin-bottom:6px;"><img src="{logo_url}" alt="Logo" style="max-height:120px; max-width:240px; object-fit:contain;"></div>'
        return ''
#############
    def _number_to_words(self, number):
        """Convert a number to words. Handles negative values with 'Minus' in parentheses."""
        try:
            from num2words import num2words
            is_negative = number < 0
            abs_number = abs(number)
            integer_part = int(abs_number)
            decimal_part = int((abs_number - integer_part) * 100)
            words = num2words(integer_part, lang='en').replace(',', '').title()
            if decimal_part > 0:
                words += f" and {decimal_part:02d}/100"
            if is_negative:
                words = f"(Minus {words})"
            return words
        except Exception:
            integer_part = int(abs(number))
            decimal_part = int((abs(number) - integer_part) * 100)
            result = f"{integer_part} and {decimal_part:02d}/100"
            if number < 0:
                result = f"(Minus {result})"
            return result

#############
    def cost_profit_view(self, request, pk):
        company = self.get_active_company(request)
        proj = self.get_object_or_404_scoped(request, Project, pk=pk)
        logo_url = company.logo.url if company and company.logo else ''

        latest_inv = Invoice.objects.filter(
            project=proj, is_advance_invoice=False
        ).order_by('-inv_number').first()

        boq_items = BOQItem.objects.filter(project=proj).order_by('item_number')

        rows = ""
        grand_revenue = Decimal("0")
        grand_direct_expenses = Decimal("0")
        grand_manpower = Decimal("0")
        grand_expenses = Decimal("0")
        grand_profit = Decimal("0")

        total_manpower = Decimal("0")
        payroll_filter = {'project': proj}
        if company:
            payroll_filter['payroll_record__employee__company'] = company

        for cc in PayrollCostCenter.objects.filter(**payroll_filter).select_related('payroll_record__employee'):
            emp = cc.payroll_record.employee
            days = cc.days_count
            pr = cc.payroll_record
            days_in_month = calendar.monthrange(pr.month.year, pr.month.month)[1]

            payroll_portion = money(
                (pr.total_salary_snap + pr.overtime_amount_snap) *
                Decimal(days) / Decimal(days_in_month)
            )

            annual_admin = (
                    emp.annual_benefits + emp.annual_eid_cost +
                    emp.annual_visa_cost + emp.annual_ticket_cost
            )
            admin_portion = money(annual_admin / Decimal("312") * Decimal(days))

            total_manpower += payroll_portion + admin_portion

        emp_filter = {'project': proj, 'is_active': True}
        if company:
            emp_filter['company'] = company

        for emp in Employee.objects.filter(**emp_filter):
            for pr in PayrollRecord.objects.filter(employee=emp):
                if not pr.cost_centers.filter(project=proj).exists():
                    days_in_month = calendar.monthrange(pr.month.year, pr.month.month)[1]
                    days_worked = days_in_month - pr.days_absent
                    if days_worked > 0:
                        payroll_portion = money(
                            (pr.total_salary_snap + pr.overtime_amount_snap) *
                            Decimal(days_worked) / Decimal(days_in_month)
                        )
                        annual_admin = (
                                emp.annual_benefits + emp.annual_eid_cost +
                                emp.annual_visa_cost + emp.annual_ticket_cost
                        )
                        admin_portion = money(
                            annual_admin / Decimal("312") * Decimal(days_worked)
                        )
                        total_manpower += payroll_portion + admin_portion

        boq_manpower = {}
        total_work = latest_inv.cumulative_work_done if latest_inv else Decimal("0")
        total_boq_value = sum(b.quantity * b.rate for b in boq_items)

        for boq in boq_items:
            cum_amt = InvoiceItem.objects.filter(
                boq_item=boq,
                invoice__project=proj,
                invoice__is_advance_invoice=False
            ).aggregate(total=Sum('gross_amount'))['total'] or Decimal("0")

            if total_work > 0:
                pct = cum_amt / total_work
            elif total_boq_value > 0:
                pct = (boq.quantity * boq.rate) / total_boq_value
            else:
                pct = Decimal("0")

            boq_manpower[boq.id] = money(total_manpower * pct)

        for boq in boq_items:
            inv_items = InvoiceItem.objects.filter(
                boq_item=boq,
                invoice__project=proj,
                invoice__is_advance_invoice=False
            )

            cum_qty = inv_items.aggregate(total=Sum('current_qty'))['total'] or Decimal("0")
            cum_amt = inv_items.aggregate(total=Sum('gross_amount'))['total'] or Decimal("0")
            revenue = money(cum_amt)

            direct_expenses = Expense.objects.filter(
                boq_item=boq, project=proj
            ).aggregate(total=Sum('amount'))['total'] or Decimal("0")

            payroll_allocations = boq_manpower.get(boq.id, Decimal("0"))

            total_expenses = money(direct_expenses + payroll_allocations)
            profit_loss = money(revenue - total_expenses)
            profit_pct = (profit_loss / revenue * 100) if revenue > 0 else Decimal("0")

            profit_color = "#2e7d32" if profit_loss >= 0 else "#d32f2f"
            profit_icon = "+" if profit_loss >= 0 else "-"

            grand_revenue += revenue
            grand_direct_expenses += direct_expenses
            grand_manpower += payroll_allocations
            grand_expenses += total_expenses
            grand_profit += profit_loss

            rows += f"""
                <tr>
                    <td class='col-item'>{boq.item_number}</td>
                    <td class='col-desc'>{boq.description}</td>
                    <td class='col-unit'>{boq.unit}</td>
                    <td class='col-num'>{boq.quantity:,.2f}</td>
                    <td class='col-num'>{boq.rate:,.2f}</td>
                    <td class='col-num'>{boq.quantity * boq.rate:,.2f}</td>
                    <td class='col-num'>{cum_qty:,.2f}</td>
                    <td class='col-num' style='color:#000080; font-weight:bold;'>{revenue:,.2f}</td>
                    <td class='col-num' style='color:#ed6c02;'>{direct_expenses:,.2f}</td>
                    <td class='col-num' style='color:#ed6c02;'>{payroll_allocations:,.2f}</td>
                    <td class='col-num' style='font-weight:bold;'>{total_expenses:,.2f}</td>
                    <td class='col-num' style='color:{profit_color}; font-weight:bold; font-size:11px;'>
                        {profit_icon} {abs(profit_loss):,.2f}
                    </td>
                    <td class='col-num' style='color:{profit_color};'>{profit_pct:.1f}%</td>
                </tr>
                """
        grand_profit_pct = (grand_profit / grand_revenue * 100) if grand_revenue > 0 else Decimal("0")
        grand_color = "#2e7d32" if grand_profit >= 0 else "#d32f2f"

        # Include variations in project value
        variation_amount = getattr(proj, 'variation_amount', Decimal("0"))
        po_amount = proj.po_amount + variation_amount
        total_work_done = latest_inv.cumulative_work_done if latest_inv else Decimal("0")
        balance = money(po_amount - total_work_done)
        progress_pct = (total_work_done / po_amount * 100) if po_amount > 0 else Decimal("0")

        # Variation display text
        var_display = f' (+{variation_amount:,.2f} var)' if variation_amount > 0 else ''

        html = f"""<!DOCTYPE html>
    <html><head><meta charset="UTF-8">
    <style>
        @page {{ size: A4 landscape; margin: 10mm; }}
        * {{ box-sizing: border-box; margin:0; padding:0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
        body {{ font-family: "Segoe UI", Arial, sans-serif; font-size: 9px; color: #222; padding: 10px; }}
        .logo-bar {{ text-align: right; margin-bottom: 6px; }}
        .logo-bar img {{ max-height: 120px; max-width: 240px; object-fit: contain; }}
        .report-title {{ font-size: 20px; font-weight: bold; text-align: center; color: #000080; margin-bottom: 4px; }}
        .report-subtitle {{ font-size: 12px; text-align: center; color: #666; margin-bottom: 12px; }}
        .meta-box {{ background: #f5f5f5; padding: 10px 15px; border-radius: 6px; margin-bottom: 15px; line-height: 1.5; font-size: 10px; display: flex; justify-content: space-between; }}
        .meta-left {{ flex: 1; }}
        .meta-right {{ flex: 1; text-align: right; }}
        .dashboard {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 10px; margin-bottom: 15px; }}
        .dash-card {{ border: 1px solid #ccc; border-radius: 6px; padding: 10px; text-align: center; background: #fafafa; }}
        .dash-label {{ font-size: 7px; color: #666; text-transform: uppercase; margin-bottom: 4px; }}
        .dash-value {{ font-size: 14px; font-weight: bold; color: #000080; }}
        .dash-sub {{ font-size: 8px; color: #666; margin-top: 3px; }}
        .section-header {{ font-size: 11px; font-weight: bold; color: #000080; margin: 15px 0 6px 0; border-bottom: 2px solid #000080; padding-bottom: 3px; }}
        .report-table {{ width: 100%; border-collapse: collapse; font-size: 8px; margin-top: 4px; }}
        .report-table th {{ background: #e8e8e8; border: 1px solid #999; padding: 4px 3px; font-weight: bold; text-align: center; font-size: 7.5px; }}
        .report-table td {{ border: 1px solid #ccc; padding: 3px 4px; vertical-align: top; }}
        .report-table .num {{ text-align: right; white-space: nowrap; }}
        .report-table tr:nth-child(even) {{ background: #fafafa; }}
        .col-item {{ width: 5%; text-align: center; }}
        .col-desc {{ width: 22%; text-align: left; }}
        .col-unit {{ width: 4%; text-align: center; }}
        .col-num {{ width: 7%; text-align: right; }}
        .total-row td {{ background: #e3f2fd; font-weight: bold; border-top: 2px solid #333; font-size: 9px; }}
        .grand-total-row td {{ background: {grand_color}; color: white; font-weight: bold; border-top: 3px solid #333; font-size: 11px; }}
        .legend {{ margin-top: 10px; padding: 8px; background: #f9f9f9; border-radius: 4px; font-size: 8px; }}
        .legend-item {{ display: inline-block; margin-right: 15px; }}
        .legend-color {{ display: inline-block; width: 10px; height: 10px; border-radius: 2px; margin-right: 3px; vertical-align: middle; }}
        .bar-track {{ width: 100%; height: 16px; background: #e0e0e0; border-radius: 8px; overflow: hidden; margin-top: 4px; }}
        .bar-fill {{ height: 100%; background: linear-gradient(90deg, #447e9b, #2e7d32); border-radius: 8px; }}
        @media print {{ .no-print {{ display: none; }} }}
    </style></head><body>
        {self._logo_bar(logo_url)}
        <div class="report-title">PROJECT COST & PROFITABILITY ANALYSIS</div>
        <div class="report-subtitle">{proj.project_id_code} — {proj.project_name}</div>
        <div class="meta-box">
            <div class="meta-left">
                <b>Client:</b> {proj.client.name}<br>
                <b>PO Number:</b> {proj.po_number or 'N/A'}<br>
                <b>PO Amount:</b> {proj.po_amount:,.2f}{var_display}
            </div>
            <div class="meta-right">
                <b>Report Date:</b> {date.today().strftime('%d-%b-%Y')}<br>
                <b>BOQ Items:</b> {boq_items.count()}<br>
                <b>Project Progress:</b> {progress_pct:.1f}%
            </div>
        </div>
        <div class="dashboard">
            <div class="dash-card"><div class="dash-label">Total Revenue</div><div class="dash-value" style="color:#000080;">{grand_revenue:,.2f}</div></div>
            <div class="dash-card"><div class="dash-label">Total Expenses</div><div class="dash-value" style="color:#ed6c02;">{grand_expenses:,.2f}</div></div>
            <div class="dash-card"><div class="dash-label">Net Profit / Loss</div><div class="dash-value" style="color:{grand_color};">{grand_profit:,.2f}</div></div>
            <div class="dash-card"><div class="dash-label">Profit Margin</div><div class="dash-value" style="color:{grand_color};">{grand_profit_pct:.1f}%</div></div>
            <div class="dash-card"><div class="dash-label">Balance to Complete</div><div class="dash-value" style="color:#d32f2f;">{balance:,.2f}</div><div class="dash-sub">of {po_amount:,.2f} PO{var_display}</div></div>
        </div>
        <div class="bar-track"><div class="bar-fill" style="width:{min(float(progress_pct), 100)}%;"></div></div>
        <div class="section-header">BOQ Item Cost Breakdown</div>
        <table class="report-table">
            <thead>
                <tr>
                    <th rowspan="2" class="col-item">Item</th>
                    <th rowspan="2" class="col-desc">Description</th>
                    <th rowspan="2" class="col-unit">Unit</th>
                    <th rowspan="2" class="col-num">BOQ Qty</th>
                    <th rowspan="2" class="col-num">Rate</th>
                    <th rowspan="2" class="col-num">BOQ Value</th>
                    <th colspan="2" style="background:#447e9b; color:white;">REVENUE (Work Done)</th>
                    <th colspan="3" style="background:#ed6c02; color:white;">EXPENSES</th>
                    <th colspan="2" style="background:{grand_color}; color:white;">PROFIT / LOSS</th>
                </tr>
                <tr>
                    <th class="col-num">Cum. Qty</th>
                    <th class="col-num">Amount</th>
                    <th class="col-num">Direct</th>
                    <th class="col-num">Manpower</th>
                    <th class="col-num">Total</th>
                    <th class="col-num">Amount</th>
                    <th class="col-num">%</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
            <tfoot>
                <tr class="total-row">
                    <td colspan="7"><b>GRAND TOTALS</b></td>
                    <td class='col-num'><b>{grand_revenue:,.2f}</b></td>
                    <td class='col-num' style='color:#ed6c02;'><b>{grand_direct_expenses:,.2f}</b></td>
                    <td class='col-num' style='color:#ed6c02;'><b>{grand_manpower:,.2f}</b></td>
                    <td class='col-num'><b>{grand_expenses:,.2f}</b></td>
                    <td class='col-num' style='color:{grand_color}; font-size:11px;'><b>{grand_profit:,.2f}</b></td>
                    <td class='col-num' style='color:{grand_color};'><b>{grand_profit_pct:.1f}%</b></td>
                </tr>
            </tfoot>
        </table>
        <div class="legend">
            <div class="legend-item"><span class="legend-color" style="background:#447e9b;"></span> Revenue = Cumulative work done (invoice gross amount)</div>
            <div class="legend-item"><span class="legend-color" style="background:#ed6c02;"></span> Direct Expenses = Costs linked to BOQ item + Manpower Cost allocations</div>
            <div class="legend-item"><span class="legend-color" style="background:#2e7d32;"></span> Profit = Revenue minus Expenses</div>
            <div class="legend-item"><span class="legend-color" style="background:#d32f2f;"></span> Loss = Negative profit (expenses exceed revenue)</div>
        </div>
        <script>window.onload = function() {{ window.print(); }}</script>
    </body></html>"""
        return HttpResponse(html)
#############

    def project_statement_view(self, request, pk):
            """
            Professional Project Statement of Final Dues.
            Two-column layout: Current Statement vs Project Completed.
            Supports ?format=excel for XLSX download.
            """
            from decimal import Decimal
            from django.db.models import Sum
            from datetime import date

            company = self.get_active_company(request)
            proj = self.get_object_or_404_scoped(request, Project, pk=pk)

            # Check if Excel export requested
            if request.GET.get('format') == 'excel':
                return self._project_statement_excel(request, proj, company)

            logo_url = company.logo.url if company and company.logo else ''

            invoices = Invoice.objects.filter(
                project=proj, inv_type='T'
            ).exclude(is_advance_invoice=True).order_by('inv_number')

            latest_inv = invoices.order_by('-inv_number').first()

            original_po = proj.po_amount
            amendments = getattr(proj, 'amendment_amount', Decimal("0"))
            variations = getattr(proj, 'variation_amount', Decimal("0"))
            final_contract_value = original_po + amendments + variations

            certified_work = latest_inv.cumulative_work_done if latest_inv else Decimal("0")
            total_materials = sum(
                (inv.material_supplied_by_client or Decimal("0")) for inv in invoices
            )
            back_charges = getattr(proj, 'back_charges', Decimal("0"))
            estimated_back_charges = getattr(proj, 'estimated_back_charges', Decimal("0"))
            liquidated_damages = getattr(proj, 'liquidated_damages', Decimal("0"))

            # Net Amount Payable (shown in statement)
            net_payable = final_contract_value - total_materials - back_charges - estimated_back_charges - liquidated_damages

            # Previously Paid (Net) shows amount AFTER deducting materials
            paid_invoices = invoices.filter(status='Paid')
            previously_paid = sum(
                (inv.current_certified_net_before_vat or Decimal("0")) for inv in paid_invoices
            )
            previously_paid_net = money(previously_paid - total_materials)
            advance_paid = getattr(proj, 'advance_paid', Decimal("0"))

            ret_a_pct = proj.retention_a_percent or Decimal("0")
            ret_b_pct = proj.retention_b_percent or Decimal("0")

            ret_a_cum = latest_inv.cumulative_retention_a if latest_inv else Decimal("0")
            ret_b_cum = latest_inv.cumulative_retention_b if latest_inv else Decimal("0")
            ret_a_rec = latest_inv.cumulative_retention_a_recovered if latest_inv else Decimal("0")
            ret_b_rec = latest_inv.cumulative_retention_b_recovered if latest_inv else Decimal("0")

            total_retention_held = (ret_a_cum + ret_b_cum) - (ret_a_rec + ret_b_rec)
            total_retention_deducted = ret_a_cum + ret_b_cum

            completed_ret_a = final_contract_value * ret_a_pct / Decimal("100")
            completed_ret_b = final_contract_value * ret_b_pct / Decimal("100")
            completed_retention = completed_ret_a + completed_ret_b

            # FIXED: Use net_payable (not current_net_base) for consistency
            # The statement shows "Net Amount Payable" which is based on final_contract_value
            # So Total Net Payable should also use net_payable
            current_amount_payable = net_payable - total_retention_held - previously_paid_net - advance_paid

            completed_amount_payable = net_payable - completed_retention - previously_paid_net - advance_paid

            progress_pct = (certified_work / final_contract_value * 100) if final_contract_value > 0 else Decimal("0")

            # Helper function for formatting values - DEFINED FIRST
            def _money(val):
                return Decimal(str(val)) if val is not None else Decimal("0")

            def fmt_negative(val):
                """Format negative values in parentheses with red color."""
                val = _money(val)
                if val < 0:
                    return f'<span style="color: #c62828;">({abs(val):,.2f})</span>'
                return f'{val:,.2f}'

            def fmt_negative_plain(val):
                """Format negative values for words - show minus in parentheses."""
                val = _money(val)
                if val < 0:
                    return f"(Minus {self._number_to_words(abs(val))})"
                return self._number_to_words(val)

            def fmt_negative_label(val):
                """Format label values (deductions) - always show in parentheses."""
                val = _money(val)
                if val < 0:
                    return f'<span style="color: #c62828;">({abs(val):,.2f})</span>'
                return f'({val:,.2f})'

            def fmt_grand_total(val):
                """Format grand total - ALWAYS white text, negative shown in parentheses."""
                val = _money(val)
                if val < 0:
                    return f'({abs(val):,.2f})'
                return f'{val:,.2f}'

            # Build invoice rows
            inv_rows = ""
            for inv in invoices:
                net = inv.current_certified_net_before_vat or Decimal("0")
                if inv.status == 'Paid':
                    badge = '<span class="badge badge-success">Paid</span>'
                elif inv.status == 'Approved':
                    badge = '<span class="badge badge-info">Approved</span>'
                else:
                    badge = '<span class="badge badge-warning">Draft</span>'
                inv_rows += f"""<tr>
                        <td class="center">{inv.inv_number}</td>
                        <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                        <td class="center">{badge}</td>
                        <td class="num">{inv.cumulative_work_done:,.2f}</td>
                        <td class="num">{inv.previous_work_done:,.2f}</td>
                        <td class="num font-bold">{net:,.2f}</td>
                        <td class="num">({inv.cumulative_retention_a:,.2f})</td>
                        <td class="num">({inv.cumulative_retention_b:,.2f})</td>
                        <td class="num">({inv.cumulative_advance_recovered:,.2f})</td>
                        <td class="num font-bold">{inv.total_after_vat:,.2f}</td>
                    </tr>"""

            html = f"""<!DOCTYPE html>
    <html><head><meta charset="UTF-8">
    <style>
        @page {{ size: A4 portrait; margin: 10mm; }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
        body {{ font-family: "Segoe UI", Arial, sans-serif; font-size: 9px; color: #222; line-height: 1.4; }}
        .page {{ padding: 8mm; }}
        .logo-bar {{ text-align: right; margin-bottom: 8px; }}
        .logo-bar img {{ max-height: 100px; max-width: 200px; object-fit: contain; }}
        .report-header {{ text-align: center; margin-bottom: 15px; padding-bottom: 10px; border-bottom: 3px solid #1a237e; }}
        .report-title {{ font-size: 16px; font-weight: 700; color: #1a237e; text-transform: uppercase; letter-spacing: 1px; }}
        .report-subtitle {{ font-size: 10px; color: #6c757d; margin-top: 4px; }}
        .meta-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 8px; margin-bottom: 15px; }}
        .meta-item {{ background: #f8f9fa; padding: 8px 10px; border-radius: 4px; border-left: 3px solid #1a237e; }}
        .meta-label {{ font-size: 7px; text-transform: uppercase; color: #6c757d; letter-spacing: 0.5px; margin-bottom: 2px; }}
        .meta-value {{ font-size: 10px; font-weight: 600; color: #212529; }}
        .two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin-bottom: 15px; }}
        .col-panel {{ border: 1px solid #dee2e6; border-radius: 8px; overflow: hidden; }}
        .col-header {{ background: #1a237e; color: white; padding: 8px 12px; font-size: 11px; font-weight: 700; text-align: center; }}
        .col-header.completed {{ background: #00695c; }}
        .col-body {{ padding: 10px; }}
        .stmt-table {{ width: 100%; border-collapse: collapse; font-size: 9px; }}
        .stmt-table td {{ padding: 5px 8px; border-bottom: 1px solid #e9ecef; vertical-align: middle; }}
        .stmt-table .label {{ text-align: left; font-weight: 600; color: #495057; width: 55%; }}
        .stmt-table .value {{ text-align: right; font-weight: 700; color: #212529; width: 45%; }}
        .stmt-table .deduction {{ color: #c62828; }}
        .stmt-table .total-row td {{ background: #e8eaf6; border-top: 2px solid #1a237e; font-size: 10px; }}
        .stmt-table .grand-total td {{ background: #1a237e; color: white; font-size: 11px; font-weight: 700; }}
        .stmt-table .grand-total .value {{ color: white; }}
        .inv-table {{ width: 100%; border-collapse: collapse; font-size: 8px; margin-top: 10px; }}
        .inv-table th {{ background: #1a237e; color: white; padding: 6px 4px; font-weight: 600; text-align: center; font-size: 7px; text-transform: uppercase; }}
        .inv-table td {{ border: 1px solid #dee2e6; padding: 4px; text-align: center; }}
        .inv-table .num {{ text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; }}
        .inv-table tr:nth-child(even) {{ background: #f8f9fa; }}
        .badge {{ display: inline-block; padding: 1px 6px; border-radius: 10px; font-size: 7px; font-weight: 700; text-transform: uppercase; }}
        .badge-success {{ background: #e8f5e9; color: #2e7d32; }}
        .badge-info {{ background: #e3f2fd; color: #0277bd; }}
        .badge-warning {{ background: #fff3e0; color: #f57c00; }}
        .signature-section {{ margin-top: 30px; padding-top: 20px; border-top: 2px solid #1a237e; }}
        .signature-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 40px; margin-top: 20px; }}
        .sig-block {{ text-align: center; }}
        .sig-line {{ border-top: 1px solid #333; margin-top: 50px; padding-top: 8px; font-size: 9px; font-weight: 600; }}
        .amount-words {{ text-align: center; margin: 15px 0; padding: 10px; background: #f8f9fa; border-radius: 6px; font-size: 10px; font-weight: 600; color: #1a237e; }}
        .amount-words.negative {{ color: #c62828; }}
        .section-title {{ font-size: 11px; font-weight: 700; color: #1a237e; margin: 15px 0 8px 0; padding-bottom: 5px; border-bottom: 2px solid #1a237e; }}
        .progress-bar {{ width: 100%; height: 20px; background: #e0e0e0; border-radius: 10px; overflow: hidden; margin: 8px 0; }}
        .progress-fill {{ height: 100%; background: linear-gradient(90deg, #447e9b, #2e7d32); border-radius: 10px; display: flex; align-items: center; justify-content: flex-end; padding-right: 8px; color: white; font-size: 8px; font-weight: 700; }}
        .excel-btn {{ display: inline-block; background: #217346; color: white; padding: 8px 16px; border-radius: 4px; text-decoration: none; font-size: 11px; font-weight: 600; margin-bottom: 15px; }}
        .excel-btn:hover {{ background: #1e663f; }}
    </style></head>
    <body>
        <div class="page">
            <div class="logo-bar">{f'<img src="{logo_url}" alt="Logo">' if logo_url else ''}</div>

            <!-- Excel Download Button -->
            <div style="text-align: right; margin-bottom: 10px;">
                <a href="?format=excel" class="excel-btn no-print">📊 Download Excel</a>
            </div>

            <div class="report-header">
                <div class="report-title">Statement of Final Dues</div>
                <div class="report-subtitle">Project: {proj.project_name} | Client: {proj.client.name}</div>
            </div>
            <div class="meta-grid">
                <div class="meta-item"><div class="meta-label">Project Code</div><div class="meta-value">{proj.project_id_code}</div></div>
                <div class="meta-item"><div class="meta-label">PO Number</div><div class="meta-value">{proj.po_number or 'N/A'}</div></div>
                <div class="meta-item"><div class="meta-label">Report Date</div><div class="meta-value">{date.today().strftime('%d-%b-%Y')}</div></div>
            </div>
            <div class="two-col">
                <div class="col-panel">
                    <div class="col-header">CURRENT STATEMENT</div>
                    <div class="col-body">
                        <table class="stmt-table">
                            <tr><td class="label">Original Contract Price</td><td class="value">{original_po:,.2f}</td></tr>
                            <tr><td class="label">Amendments</td><td class="value">{amendments:,.2f}</td></tr>
                            <tr><td class="label">Variations</td><td class="value">{variations:,.2f}</td></tr>
                            <tr class="total-row"><td class="label"><b>Final Contract Value</b></td><td class="value"><b>{final_contract_value:,.2f}</b></td></tr>
                            <tr><td colspan="2" style="height:8px;"></td></tr>
                            <tr><td class="label deduction">Materials Supplied by Client</td><td class="value deduction">{fmt_negative_label(total_materials)}</td></tr>
                            <tr><td class="label deduction">Back Charges / Contra-Charges</td><td class="value deduction">{fmt_negative_label(back_charges)}</td></tr>
                            <tr><td class="label deduction">Estimated Back Charges</td><td class="value deduction">{fmt_negative_label(estimated_back_charges)}</td></tr>
                            <tr><td class="label deduction">Liquidated Damages</td><td class="value deduction">{fmt_negative_label(liquidated_damages)}</td></tr>
                            <tr class="total-row"><td class="label"><b>Net Amount Payable</b></td><td class="value"><b>{fmt_negative(net_payable)}</b></td></tr>
                            <tr><td colspan="2" style="height:8px;"></td></tr>
                            <tr><td class="label">Previously Paid (Advance)</td><td class="value">{fmt_negative_label(advance_paid)}</td></tr>
                            <tr><td class="label">Previously Paid (Net of Materials)</td><td class="value">{fmt_negative_label(previously_paid_net)}</td></tr>
                            <tr><td class="label deduction">Retention A ({ret_a_pct}%)</td><td class="value deduction">{fmt_negative_label(ret_a_cum)}</td></tr>
                            <tr><td class="label deduction">Retention B ({ret_b_pct}%)</td><td class="value deduction">{fmt_negative_label(ret_b_cum)}</td></tr>
                            <tr class="grand-total"><td class="label">TOTAL NET PAYABLE</td><td class="value">{fmt_grand_total(current_amount_payable)}</td></tr>
                        </table>
                        <div style="margin-top: 10px; padding: 8px; background: #f5f5f5; border-radius: 4px; font-size: 8px; color: #666;">
                            <b>Progress:</b> {certified_work:,.2f} of {final_contract_value:,.2f} ({progress_pct:.1f}%)
                        </div>
                        <div class="progress-bar"><div class="progress-fill" style="width: {min(float(progress_pct), 100)}%;">{progress_pct:.1f}%</div></div>
                    </div>
                </div>
                <div class="col-panel">
                    <div class="col-header completed">IF PROJECT COMPLETED</div>
                    <div class="col-body">
                        <table class="stmt-table">
                            <tr><td class="label">Original Contract Price</td><td class="value">{original_po:,.2f}</td></tr>
                            <tr><td class="label">Amendments</td><td class="value">{amendments:,.2f}</td></tr>
                            <tr><td class="label">Variations</td><td class="value">{variations:,.2f}</td></tr>
                            <tr class="total-row"><td class="label"><b>Final Contract Value</b></td><td class="value"><b>{final_contract_value:,.2f}</b></td></tr>
                            <tr><td colspan="2" style="height:8px;"></td></tr>
                            <tr><td class="label deduction">Materials Supplied by Client</td><td class="value deduction">{fmt_negative_label(total_materials)}</td></tr>
                            <tr><td class="label deduction">Back Charges / Contra-Charges</td><td class="value deduction">{fmt_negative_label(back_charges)}</td></tr>
                            <tr><td class="label deduction">Estimated Back Charges</td><td class="value deduction">{fmt_negative_label(estimated_back_charges)}</td></tr>
                            <tr><td class="label deduction">Liquidated Damages</td><td class="value deduction">{fmt_negative_label(liquidated_damages)}</td></tr>
                            <tr class="total-row"><td class="label"><b>Net Amount Payable</b></td><td class="value"><b>{fmt_negative(net_payable)}</b></td></tr>
                            <tr><td colspan="2" style="height:8px;"></td></tr>
                            <tr><td class="label">Previously Paid (Advance)</td><td class="value">{fmt_negative_label(advance_paid)}</td></tr>
                            <tr><td class="label">Previously Paid (Net of Materials)</td><td class="value">{fmt_negative_label(previously_paid_net)}</td></tr>
                            <tr><td class="label deduction">Retention A ({ret_a_pct}%)</td><td class="value deduction">{fmt_negative_label(completed_ret_a)}</td></tr>
                            <tr><td class="label deduction">Retention B ({ret_b_pct}%)</td><td class="value deduction">{fmt_negative_label(completed_ret_b)}</td></tr>
                            <tr class="grand-total"><td class="label">TOTAL NET PAYABLE</td><td class="value">{fmt_grand_total(completed_amount_payable)}</td></tr>
                        </table>
                        <div style="margin-top: 10px; padding: 8px; background: #e8f5e9; border-radius: 4px; font-size: 8px; color: #2e7d32;"><b>100% Complete</b> — Full contract value realized</div>
                        <div class="progress-bar"><div class="progress-fill" style="width: 100%;">100%</div></div>
                    </div>
                </div>
            </div>
            <div class="section-title">Retention Summary</div>
            <div class="two-col" style="margin-bottom: 15px;">
                <div class="col-panel">
                    <div class="col-header">CURRENT RETENTION STATUS</div>
                    <div class="col-body">
                        <table class="stmt-table">
                            <tr><td class="label">Retention A Deducted (Cum)</td><td class="value">{ret_a_cum:,.2f}</td></tr>
                            <tr><td class="label">Retention A Recovered</td><td class="value addition">{fmt_negative_label(ret_a_rec)}</td></tr>
                            <tr class="total-row"><td class="label"><b>Retention A Held</b></td><td class="value"><b>{fmt_negative(ret_a_cum - ret_a_rec)}</b></td></tr>
                            <tr><td colspan="2" style="height:5px;"></td></tr>
                            <tr><td class="label">Retention B Deducted (Cum)</td><td class="value">{ret_b_cum:,.2f}</td></tr>
                            <tr><td class="label">Retention B Recovered</td><td class="value addition">{fmt_negative_label(ret_b_rec)}</td></tr>
                            <tr class="total-row"><td class="label"><b>Retention B Held</b></td><td class="value"><b>{fmt_negative(ret_b_cum - ret_b_rec)}</b></td></tr>
                            <tr><td colspan="2" style="height:5px;"></td></tr>
                            <tr class="grand-total"><td class="label">TOTAL RETENTION HELD</td><td class="value">{fmt_grand_total(total_retention_held)}</td></tr>
                        </table>
                    </div>
                </div>
                <div class="col-panel">
                    <div class="col-header completed">UPON PROJECT COMPLETION</div>
                    <div class="col-body">
                        <table class="stmt-table">
                            <tr><td class="label">Retention A ({ret_a_pct}% of contract)</td><td class="value">{completed_ret_a:,.2f}</td></tr>
                            <tr><td class="label">Retention A Already Recovered</td><td class="value addition">{fmt_negative_label(ret_a_rec)}</td></tr>
                            <tr class="total-row"><td class="label"><b>Retention A Remaining</b></td><td class="value"><b>{fmt_negative(completed_ret_a - ret_a_rec)}</b></td></tr>
                            <tr><td colspan="2" style="height:5px;"></td></tr>
                            <tr><td class="label">Retention B ({ret_b_pct}% of contract)</td><td class="value">{completed_ret_b:,.2f}</td></tr>
                            <tr><td class="label">Retention B Already Recovered</td><td class="value addition">{fmt_negative_label(ret_b_rec)}</td></tr>
                            <tr class="total-row"><td class="label"><b>Retention B Remaining</b></td><td class="value"><b>{fmt_negative(completed_ret_b - ret_b_rec)}</b></td></tr>
                            <tr><td colspan="2" style="height:5px;"></td></tr>
                            <tr class="grand-total"><td class="label">TOTAL RETENTION TO RECOVER</td><td class="value">{fmt_grand_total(completed_retention - total_retention_deducted + total_retention_held)}</td></tr>
                        </table>
                    </div>
                </div>
            </div>
            <div class="section-title">Tax Invoice History (Excl. VAT)</div>
            <table class="inv-table">
                <thead><tr><th>Invoice #</th><th>Date</th><th>Status</th><th>Cum. Work</th><th>Prev. Work</th><th>Current Net</th><th>Ret A</th><th>Ret B</th><th>Adv Rec</th><th>Payable</th></tr></thead>
                <tbody>{inv_rows if inv_rows else '<tr><td colspan="10" style="text-align:center; color:#999; padding:20px;">No tax invoices found</td></tr>'}</tbody>
                <tfoot><tr style="background: #e8eaf6; font-weight: bold; border-top: 2px solid #1a237e;">
                    <td colspan="3"><b>TOTALS</b></td>
                    <td class="num">{certified_work:,.2f}</td>
                    <td class="num">—</td>
                    <td class="num">{sum((inv.current_certified_net_before_vat or Decimal("0")) for inv in invoices):,.2f}</td>
                    <td class="num">{fmt_negative_label(ret_a_cum)}</td>
                    <td class="num">{fmt_negative_label(ret_b_cum)}</td>
                    <td class="num">{fmt_negative_label(latest_inv.cumulative_advance_recovered if latest_inv else Decimal("0"))}</td>
                    <td class="num">{sum((inv.total_after_vat or Decimal("0")) for inv in invoices):,.2f}</td>
                </tr></tfoot>
            </table>
            <div class="amount-words {'negative' if current_amount_payable < 0 else ''}">
                Amount in Words: AED {fmt_negative_plain(current_amount_payable)} Dirhams Only
            </div>
            <div class="signature-section">
                <div style="text-align: center; font-size: 10px; margin-bottom: 15px; color: #1a237e; font-weight: 700;">CERTIFICATION</div>
                <div style="font-size: 9px; line-height: 1.6; margin-bottom: 20px; text-align: center;">
                    We, <b>{company.company_name if company else "PROCON GENERAL CONTRACTING L.L.C"}</b> hereby confirm that the final amount due, 
                    <b>AED {fmt_negative(current_amount_payable)}</b> is the full and final settlement due to us against the captioned Project.
                    <br>We further confirm that upon payment of the Final Amount due, the Client shall have honored all of its obligations under this Contract.
                </div>
                <div class="signature-grid">
                    <div class="sig-block"><div class="sig-line">For Contractor</div><div style="font-size: 8px; color: #666; margin-top: 4px;">Name & Title</div></div>
                    <div class="sig-block"><div class="sig-line">For Client</div><div style="font-size: 8px; color: #666; margin-top: 4px;">Name & Title</div></div>
                </div>
            </div>
            <script>window.onload = function() {{ window.print(); }}</script>
        </div>
    </body></html>"""
            return HttpResponse(html)

    def _project_statement_excel(self, request, proj, company):
            """Generate Excel version of the project statement."""
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Project Statement"

            # Styles
            header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            subheader_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
            subheader_font = Font(bold=True, color="1A237E", size=10)
            grand_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
            grand_font = Font(bold=True, color="FFFFFF", size=11)
            deduction_font = Font(color="C62828")
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")
            thin_border = Border(
                bottom=Side(style='thin', color='CCCCCC')
            )

            # Helper to format numbers
            def fmt(val):
                return float(val) if val else 0.0

            # --- HEADER ---
            ws.merge_cells('A1:D1')
            ws['A1'] = "STATEMENT OF FINAL DUES"
            ws['A1'].font = Font(bold=True, size=16, color="1A237E")
            ws['A1'].alignment = center_align

            ws.merge_cells('A2:D2')
            ws['A2'] = f"Project: {proj.project_name}"
            ws['A2'].alignment = center_align

            ws.merge_cells('A3:D3')
            ws[
                'A3'] = f"Client: {proj.client.name} | PO: {proj.po_number or 'N/A'} | Date: {date.today().strftime('%d-%b-%Y')}"
            ws['A3'].alignment = center_align
            ws.row_dimensions[3].height = 25

            row = 5

            # --- CURRENT STATEMENT COLUMN ---
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = "CURRENT STATEMENT"
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].alignment = center_align
            ws.merge_cells(f'C{row}:D{row}')
            ws[f'C{row}'] = "IF PROJECT COMPLETED"
            ws[f'C{row}'].fill = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")
            ws[f'C{row}'].font = header_font
            ws[f'C{row}'].alignment = center_align
            row += 1

            # Fetch data
            invoices = Invoice.objects.filter(project=proj, inv_type='T').exclude(is_advance_invoice=True).order_by(
                'inv_number')
            latest_inv = invoices.order_by('-inv_number').first()

            original_po = proj.po_amount
            amendments = getattr(proj, 'amendment_amount', Decimal("0"))
            variations = getattr(proj, 'variation_amount', Decimal("0"))
            final_contract_value = original_po + amendments + variations
            certified_work = latest_inv.cumulative_work_done if latest_inv else Decimal("0")
            total_materials = sum((inv.material_supplied_by_client or Decimal("0")) for inv in invoices)
            back_charges = getattr(proj, 'back_charges', Decimal("0"))
            estimated_back_charges = getattr(proj, 'estimated_back_charges', Decimal("0"))
            liquidated_damages = getattr(proj, 'liquidated_damages', Decimal("0"))
            net_payable = final_contract_value - total_materials - back_charges - estimated_back_charges - liquidated_damages

            paid_invoices = invoices.filter(status='Paid')
            previously_paid = sum((inv.current_certified_net_before_vat or Decimal("0")) for inv in paid_invoices)
            advance_paid = getattr(proj, 'advance_paid', Decimal("0"))

            ret_a_pct = proj.retention_a_percent or Decimal("0")
            ret_b_pct = proj.retention_b_percent or Decimal("0")
            ret_a_cum = latest_inv.cumulative_retention_a if latest_inv else Decimal("0")
            ret_b_cum = latest_inv.cumulative_retention_b if latest_inv else Decimal("0")
            ret_a_rec = latest_inv.cumulative_retention_a_recovered if latest_inv else Decimal("0")
            ret_b_rec = latest_inv.cumulative_retention_b_recovered if latest_inv else Decimal("0")
            total_retention_held = (ret_a_cum + ret_b_cum) - (ret_a_rec + ret_b_rec)
            total_retention_deducted = ret_a_cum + ret_b_cum

            completed_ret_a = final_contract_value * ret_a_pct / Decimal("100")
            completed_ret_b = final_contract_value * ret_b_pct / Decimal("100")
            completed_retention = completed_ret_a + completed_ret_b

            current_amount_payable = certified_work - total_materials - total_retention_held - previously_paid - advance_paid
            completed_amount_payable = final_contract_value - total_materials - completed_retention - previously_paid - advance_paid

            # Data rows
            data_rows = [
                ("Original Contract Price", original_po, original_po),
                ("Amendments", amendments, amendments),
                ("Variations", variations, variations),
                ("FINAL CONTRACT VALUE", final_contract_value, final_contract_value),
                ("", "", ""),
                ("Materials Supplied by Client", -total_materials, -total_materials),
                ("Back Charges / Contra-Charges", -back_charges, -back_charges),
                ("Estimated Back Charges", -estimated_back_charges, -estimated_back_charges),
                ("Liquidated Damages", -liquidated_damages, -liquidated_damages),
                ("NET AMOUNT PAYABLE", net_payable, net_payable),
                ("", "", ""),
                ("Previously Paid (Advance)", -advance_paid, -advance_paid),
                ("Previously Paid (Net)", -previously_paid, -previously_paid),
                (f"Retention A ({ret_a_pct}%)", -ret_a_cum, -completed_ret_a),
                (f"Retention B ({ret_b_pct}%)", -ret_b_cum, -completed_ret_b),
                ("TOTAL NET PAYABLE", current_amount_payable, completed_amount_payable),
            ]

            for label, current_val, completed_val in data_rows:
                if label == "":
                    row += 1
                    continue

                is_total = label in ["FINAL CONTRACT VALUE", "NET AMOUNT PAYABLE", "TOTAL NET PAYABLE"]
                is_deduction = current_val < 0 and label not in ["TOTAL NET PAYABLE", "NET AMOUNT PAYABLE",
                                                                 "FINAL CONTRACT VALUE"]

                ws[f'A{row}'] = label
                ws[f'A{row}'].font = bold_font if is_total else (deduction_font if is_deduction else Font())
                ws[f'A{row}'].border = thin_border

                ws[f'B{row}'] = fmt(current_val)
                ws[f'B{row}'].number_format = '#,##0.00'
                ws[f'B{row}'].alignment = right_align
                ws[f'B{row}'].font = bold_font if is_total else (deduction_font if is_deduction else Font())
                ws[f'B{row}'].border = thin_border
                if is_total:
                    ws[f'A{row}'].fill = subheader_fill
                    ws[f'B{row}'].fill = subheader_fill

                ws[f'C{row}'] = label
                ws[f'C{row}'].font = bold_font if is_total else (deduction_font if is_deduction else Font())
                ws[f'C{row}'].border = thin_border

                ws[f'D{row}'] = fmt(completed_val)
                ws[f'D{row}'].number_format = '#,##0.00'
                ws[f'D{row}'].alignment = right_align
                ws[f'D{row}'].font = bold_font if is_total else (deduction_font if is_deduction else Font())
                ws[f'D{row}'].border = thin_border
                if is_total:
                    ws[f'C{row}'].fill = subheader_fill
                    ws[f'D{row}'].fill = subheader_fill

                row += 1

            # Grand total row - DON'T merge, just write to individual cells
            ws[f'A{row}'] = "TOTAL NET PAYABLE"
            ws[f'A{row}'].fill = grand_fill
            ws[f'A{row}'].font = grand_font
            ws[f'A{row}'].alignment = center_align
            ws[f'B{row}'] = fmt(current_amount_payable)
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'B{row}'].alignment = right_align
            ws[f'B{row}'].fill = grand_fill
            ws[f'B{row}'].font = grand_font

            ws[f'C{row}'] = "TOTAL NET PAYABLE (COMPLETED)"
            ws[f'C{row}'].fill = grand_fill
            ws[f'C{row}'].font = grand_font
            ws[f'C{row}'].alignment = center_align
            ws[f'D{row}'] = fmt(completed_amount_payable)
            ws[f'D{row}'].number_format = '#,##0.00'
            ws[f'D{row}'].alignment = right_align
            ws[f'D{row}'].fill = grand_fill
            ws[f'D{row}'].font = grand_font
            row += 2

            # --- INVOICE HISTORY ---
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = "TAX INVOICE HISTORY (Excl. VAT)"
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].alignment = center_align
            row += 1

            # Invoice headers
            inv_headers = ["Invoice #", "Date", "Status", "Cum. Work", "Prev. Work", "Current Net", "Ret A", "Ret B",
                           "Adv Rec", "Payable"]
            for col_idx, header in enumerate(inv_headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = subheader_fill
                cell.font = subheader_font
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

            for inv in invoices:
                net = inv.current_certified_net_before_vat or Decimal("0")
                ws.cell(row=row, column=1, value=inv.inv_number)
                ws.cell(row=row, column=2, value=inv.date.strftime('%d-%b-%Y'))
                ws.cell(row=row, column=3, value=inv.status)
                ws.cell(row=row, column=4, value=fmt(inv.cumulative_work_done)).number_format = '#,##0.00'
                ws.cell(row=row, column=5, value=fmt(inv.previous_work_done)).number_format = '#,##0.00'
                ws.cell(row=row, column=6, value=fmt(net)).number_format = '#,##0.00'
                ws.cell(row=row, column=7, value=fmt(-inv.cumulative_retention_a)).number_format = '#,##0.00'
                ws.cell(row=row, column=8, value=fmt(-inv.cumulative_retention_b)).number_format = '#,##0.00'
                ws.cell(row=row, column=9, value=fmt(-inv.cumulative_advance_recovered)).number_format = '#,##0.00'
                ws.cell(row=row, column=10, value=fmt(inv.total_after_vat)).number_format = '#,##0.00'
                row += 1

            # Column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 18
            for col in range(1, 11):
                ws.column_dimensions[get_column_letter(col)].width = 15

            # Response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response[
                'Content-Disposition'] = f'attachment; filename="Project_Statement_{proj.project_id_code}_{date.today().strftime("%Y%m%d")}.xlsx"'

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response.write(output.getvalue())
            return response

#############
    def analytics_view(self, request, pk):
        company = self.get_active_company(request)
        proj = self.get_object_or_404_scoped(request, Project, pk=pk)
        invoices = Invoice.objects.filter(project=proj).order_by('inv_number')
        if company:
            invoices = invoices.filter(Q(project__company=company) | Q(project__company__isnull=True))
        boq_items = BOQItem.objects.filter(project=proj).order_by('item_number')
        logo_url = company.logo.url if company and company.logo else ''

        certified_invoices = invoices.filter(
            Q(inv_type='T') | Q(inv_type='P', status='Approved')
        ).exclude(is_advance_invoice=True)

        draft_proforma_invoices = invoices.filter(
            inv_type='P', status='Draft'
        ).exclude(is_advance_invoice=True)

        inv_rows = ""
        inv_total_net = Decimal("0")
        inv_total_vat = Decimal("0")
        inv_total_payable = Decimal("0")

        for inv in certified_invoices:
            net = inv.current_certified_net_before_vat
            vat = inv.vat_amount
            payable = inv.total_after_vat
            inv_total_net += net
            inv_total_vat += vat
            inv_total_payable += payable
            if inv.inv_type == 'T':
                coll_date = inv.collection_date.strftime('%d-%b-%Y') if inv.collection_date else '—'
            else:
                coll_date = 'N/A (Proforma)'
            inv_rows += f"""<tr>
                <td>{inv}</td>
                <td>{inv.get_inv_type_display()}</td>
                <td>{inv.status}</td>
                <td>{inv.date}</td>
                <td class='num'>{net:,.2f}</td>
                <td class='num'>{vat:,.2f}</td>
                <td class='num'>{payable:,.2f}</td>
                <td>{coll_date}</td>
            </tr>"""

        inv_totals_row = f"""<tr class='total-row'>
            <td colspan="4"><b>TOTAL CERTIFIED INVOICES</b></td>
            <td class='num'><b>{inv_total_net:,.2f}</b></td>
            <td class='num'><b>{inv_total_vat:,.2f}</b></td>
            <td class='num'><b>{inv_total_payable:,.2f}</b></td>
            <td></td>
        </tr>"""

        draft_rows = ""
        draft_total_net = Decimal("0")
        for inv in draft_proforma_invoices:
            gross = inv.current_gross_total
            draft_total_net += gross
            draft_rows += f"""<tr>
                <td>{inv}</td>
                <td>{inv.get_inv_type_display()}</td>
                <td>{inv.status}</td>
                <td>{inv.date}</td>
                <td class='num'>{gross:,.2f}</td>
                <td class='num'>—</td>
                <td class='num'>—</td>
                <td>—</td>
            </tr>"""

        draft_totals_row = ""
        draft_section = ""
        if draft_rows:
            draft_totals_row = f"""<tr class='total-row' style="background:#fff3e0;">
                <td colspan="4"><b>DRAFT PROFORMA TOTAL (Not Certified)</b></td>
                <td class='num'><b>{draft_total_net:,.2f}</b></td>
                <td class='num'>—</td>
                <td class='num'>—</td>
                <td></td>
            </tr>"""
            draft_section = f"""<div class="section-header">Draft Proforma Invoices <span class="draft-badge">NOT CERTIFIED</span></div>
    <table class="report-table">
        <thead>
            <tr><th>Invoice #</th><th>Type</th><th>Status</th><th>Date</th><th class='num'>Gross Amount</th><th class='num'>VAT</th><th class='num'>Payable</th><th>Collection Date</th></tr>
        </thead>
        <tbody>{draft_rows}</tbody>
        <tfoot>{draft_totals_row}</tfoot>
    </table>
    """
        else:
            draft_section = '<div class="section-header">Draft Proforma Invoices</div><p style="color:#999; font-size:11px;">No draft proforma invoices.</p>'

        boq_rows = ""
        boq_total = Decimal("0")
        for b in boq_items:
            line_total = b.quantity * b.rate
            boq_total += line_total
            boq_rows += f"""<tr>
                <td>{b.item_number}</td>
                <td>{b.description[:1500]}</td>
                <td>{b.unit}</td>
                <td class='num'>{b.quantity:,.2f}</td>
                <td class='num'>{b.rate:,.2f}</td>
                <td class='num'>{line_total:,.2f}</td>
            </tr>"""

        latest_certified = certified_invoices.order_by('-inv_number').first()
        certified_work = latest_certified.certified_work_done if latest_certified else Decimal("0")
        ret_a = latest_certified.cumulative_retention_a if latest_certified else Decimal("0")
        ret_b = latest_certified.cumulative_retention_b if latest_certified else Decimal("0")
        ret_a_rec = latest_certified.cumulative_retention_a_recovered if latest_certified else Decimal("0")
        ret_b_rec = latest_certified.cumulative_retention_b_recovered if latest_certified else Decimal("0")
        adv_rec = latest_certified.cumulative_advance_recovered if latest_certified else Decimal("0")
        net_inv = latest_certified.certified_net_invoiced_cumulative if latest_certified else Decimal("0")

        # Include variations in contract value for progress calculation
        variation_amount = getattr(proj, 'variation_amount', Decimal("0"))
        po = proj.po_amount + variation_amount
        progress_pct = (certified_work / po * 100) if po > 0 else Decimal("0")
        balance = money(po - certified_work)

        cash_collected = Decimal("0")
        for inv in invoices.filter(inv_type='T', status='Paid').exclude(is_advance_invoice=True):
            cash_collected += inv.current_certified_net_before_vat

        balance_cash = money(po - cash_collected)

        html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
    @page {{ size: A4 portrait; margin: 10mm; }}
    * {{ box-sizing: border-box; margin:0; padding:0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    body {{ font-family: "Segoe UI", Arial, sans-serif; font-size: 10px; color: #222; padding: 10px; }}
    .logo-bar {{ text-align: right; margin-bottom: 6px; }}
    .logo-bar img {{ max-height: 120px; max-width: 240px; object-fit: contain; }}
    .report-title {{ font-size: 18px; font-weight: bold; text-align: center; color: #000080; margin-bottom: 4px; }}
    .report-subtitle {{ font-size: 12px; text-align: center; color: #666; margin-bottom: 15px; }}
    .meta-box {{ background: #f5f5f5; padding: 10px 15px; border-radius: 6px; margin-bottom: 20px; line-height: 1.6; font-size: 10px; }}
    .section-header {{ font-size: 12px; font-weight: bold; color: #000080; margin: 20px 0 8px 0; border-bottom: 2px solid #000080; padding-bottom: 4px; }}
    .report-table {{ width: 100%; border-collapse: collapse; font-size: 9px; margin-top: 6px; }}
    .report-table th {{ background: #e8e8e8; border: 1px solid #999; padding: 5px; text-align: left; font-weight: bold; }}
    .report-table td {{ border: 1px solid #ccc; padding: 5px; }}
    .report-table .num {{ text-align: right; white-space: nowrap; }}
    .report-table tr:nth-child(even) {{ background: #fafafa; }}
    .total-row td {{ background: #e3f2fd; font-weight: bold; border-top: 2px solid #333; }}
    .dashboard {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-bottom: 20px; }}
    .dash-card {{ border: 1px solid #ccc; border-radius: 8px; padding: 12px; text-align: center; background: #fafafa; }}
    .dash-label {{ font-size: 8px; color: #666; text-transform: uppercase; margin-bottom: 6px; }}
    .dash-value {{ font-size: 16px; font-weight: bold; color: #000080; }}
    .dash-sub {{ font-size: 9px; color: #666; margin-top: 4px; }}
    .bar-track {{ width: 100%; height: 22px; background: #e0e0e0; border-radius: 11px; overflow: hidden; margin-top: 6px; }}
    .bar-fill {{ height: 100%; background: linear-gradient(90deg, #447e9b, #2e7d32); border-radius: 11px; }}
    .progress-label {{ text-align: center; font-weight: bold; font-size: 11px; margin-top: 8px; }}
    .two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; }}
    .panel {{ border: 1px solid #ddd; border-radius: 8px; padding: 12px; background: white; }}
    .panel-title {{ font-size: 10px; font-weight: bold; color: #000080; margin-bottom: 10px; border-bottom: 1px solid #eee; padding-bottom: 6px; }}
    .panel-row {{ display: flex; justify-content: space-between; font-size: 9px; margin-bottom: 5px; }}
    .panel-row b {{ color: #333; }}
    .draft-badge {{ display: inline-block; background: #ed6c02; color: white; padding: 2px 6px; border-radius: 3px; font-size: 8px; font-weight: bold; }}
</style></head><body>
    {self._logo_bar(logo_url)}
    <div class="report-title">PROJECT ANALYTICS REPORT</div>
    <div class="report-subtitle">{proj.project_id_code} — {proj.project_name}</div>
    <div class="meta-box">
        <b>Client:</b> {proj.client.name} &nbsp;|&nbsp;
        <b>PO Number:</b> {proj.po_number or 'N/A'} &nbsp;|&nbsp;
        <b>PO Date:</b> {proj.po_date or 'N/A'} &nbsp;|&nbsp;
        <b>BOQ Complete:</b> {'Yes' if proj.is_boq_complete else 'No'}
    </div>
    <div class="dashboard">
        <div class="dash-card">
            <div class="dash-label">PO Amount</div>
            <div class="dash-value">{po:,.2f}</div>
        </div>
        <div class="dash-card">
            <div class="dash-label">Certified Work</div>
            <div class="dash-value" style="color:#2e7d32;">{certified_work:,.2f}</div>
            <div class="dash-sub">{progress_pct:.1f}% complete</div>
        </div>
        <div class="dash-card">
            <div class="dash-label">Balance to PO</div>
            <div class="dash-value" style="color:#d32f2f;">{balance:,.2f}</div>
        </div>
        <div class="dash-card">
            <div class="dash-label">Net Invoiced (Certified)</div>
            <div class="dash-value">{net_inv:,.2f}</div>
        </div>
        <div class="dash-card">
            <div class="dash-label">Cash Collected</div>
            <div class="dash-value" style="color:#2e7d32;">{cash_collected:,.2f}</div>
            <div class="dash-sub">Paid Tax Invoices (excl. VAT)</div>
        </div>
        <div class="dash-card">
            <div class="dash-label">Balance Cash</div>
            <div class="dash-value" style="color:#d32f2f;">{balance_cash:,.2f}</div>
            <div class="dash-sub">PO minus Cash Collected</div>
        </div>
        <div class="dash-card">
            <div class="dash-label">Retention Held</div>
            <div class="dash-value" style="color:#ed6c02;">{money(ret_a + ret_b - ret_a_rec - ret_b_rec):,.2f}</div>
            <div class="dash-sub">A: {money(ret_a - ret_a_rec):,.2f} &nbsp; B: {money(ret_b - ret_b_rec):,.2f}</div>
        </div>
        <div class="dash-card">
            <div class="dash-label">Advance Status</div>
            <div class="dash-value">{adv_rec:,.2f}</div>
            <div class="dash-sub">of {proj.total_advance_value:,.2f} recovered</div>
        </div>
        <div class="dash-card">
            <div class="dash-label">Invoices</div>
            <div class="dash-value">{certified_invoices.count()}</div>
            <div class="dash-sub">{draft_proforma_invoices.count()} draft proforma</div>
        </div>
    </div>
    <div class="bar-track">
        <div class="bar-fill" style="width:{min(float(progress_pct), 100)}%;"></div>
    </div>
    <div class="progress-label">Project Progress: {progress_pct:.1f}%</div>
    <div class="two-col" style="margin-top:20px;">
        <div class="panel">
            <div class="panel-title">Retention Summary</div>
            <div class="panel-row"><span>Retention A Deducted (Cum):</span><b>{ret_a:,.2f}</b></div>
            <div class="panel-row"><span>Retention A Recovered:</span><b style="color:green;">{ret_a_rec:,.2f}</b></div>
            <div class="panel-row"><span>Retention A Held:</span><b>{money(ret_a - ret_a_rec):,.2f}</b></div>
            <div style="height:1px;background:#eee;margin:8px 0;"></div>
            <div class="panel-row"><span>Retention B Deducted (Cum):</span><b>{ret_b:,.2f}</b></div>
            <div class="panel-row"><span>Retention B Recovered:</span><b style="color:green;">{ret_b_rec:,.2f}</b></div>
            <div class="panel-row"><span>Retention B Held:</span><b>{money(ret_b - ret_b_rec):,.2f}</b></div>
        </div>
        <div class="panel">
            <div class="panel-title">Contract Terms</div>
            <div class="panel-row"><span>Advance Percent:</span><b>{proj.advance_percent}%</b></div>
            <div class="panel-row"><span>Retention A Percent:</span><b>{proj.retention_a_percent}%</b></div>
            <div class="panel-row"><span>Retention B Percent:</span><b>{proj.retention_b_percent}%</b></div>
            <div class="panel-row"><span>Total Advance Value:</span><b>{proj.total_advance_value:,.2f}</b></div>
            <div style="height:1px;background:#eee;margin:8px 0;"></div>
            <div class="panel-row"><span>BOQ Items:</span><b>{boq_items.count()}</b></div>
            <div class="panel-row"><span>BOQ Total:</span><b>{proj.boq_total_value:,.2f}</b></div>
            <div class="panel-row"><span>Variations:</span><b style="color:#6a1b9a;">{variation_amount:,.2f}</b></div>
            <div class="panel-row"><span>Total Invoices:</span><b>{invoices.count()}</b></div>
        </div>
    </div>
    <div class="section-header">Invoice History (Tax + Approved Proforma)</div>
    <table class="report-table">
        <thead>
            <tr><th>Invoice #</th><th>Type</th><th>Status</th><th>Date</th><th class='num'>Net</th><th class='num'>VAT</th><th class='num'>Payable</th><th>Collection Date</th></tr>
        </thead>
        <tbody>{inv_rows}</tbody>
        <tfoot>{inv_totals_row}</tfoot>
    </table>
    {draft_section}
    <div class="section-header">Bill of Quantities</div>
    <table class="report-table">
        <thead>
            <tr><th>Item</th><th>Description</th><th>Unit</th><th class='num'>Qty</th><th class='num'>Rate</th><th class='num'>Total</th></tr>
        </thead>
        <tbody>{boq_rows}</tbody>
        <tfoot>
            <tr class="total-row">
                <td colspan="5"><b>BOQ TOTAL</b></td>
                <td class='num'><b>{boq_total:,.2f}</b></td>
            </tr>
        </tfoot>
    </table>
</body></html>"""
        return HttpResponse(html)
################################################

    # =============================================================================
    # PROJECT DEDUCTIONS REPORT (New - Add to ProjectAdmin)
    # =============================================================================

    def deductions_view(self, request, pk):
        """
        Professional Project Deductions Report - detailed tracking of all deductions
        for a single project with invoice-level granularity.
        """
        company = self.get_active_company(request)
        proj = self.get_object_or_404_scoped(request, Project, pk=pk)
        logo_url = company.logo.url if company and company.logo else ''

        invoices = Invoice.objects.filter(
            project=proj, inv_type='T'
        ).exclude(is_advance_invoice=True).order_by('inv_number')

        latest_inv = invoices.order_by('-inv_number').first()

        # === BUILD INVOICE-LEVEL DEDUCTION TABLE ===
        inv_rows = ""
        cumulative_materials = Decimal("0")
        cumulative_ret_a = Decimal("0")
        cumulative_ret_b = Decimal("0")
        cumulative_adv = Decimal("0")
        cumulative_total_ded = Decimal("0")

        for inv in invoices:
            mat = inv.material_supplied_by_client or Decimal("0")
            ret_a = inv.current_retention_a
            ret_b = inv.current_retention_b
            adv = inv.current_advance_recovery
            total_ded = mat + ret_a + ret_b + adv

            cumulative_materials += mat
            cumulative_ret_a += ret_a
            cumulative_ret_b += ret_b
            cumulative_adv += adv
            cumulative_total_ded += total_ded

            status_badge = '<span class="badge badge-success">Paid</span>' if inv.status == 'Paid' else \
                '<span class="badge badge-warning">Pending</span>' if inv.status == 'Draft' else \
                    '<span class="badge badge-info">Approved</span>'

            inv_rows += f"""
            <tr>
                <td class="center">{inv.inv_number}</td>
                <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                <td class="text"><b>{inv}</b></td>
                <td class="center">{status_badge}</td>
                <td class="num">{inv.current_gross_total:,.2f}</td>
                <td class="num text-danger">({ret_a:,.2f})</td>
                <td class="num text-danger">({ret_b:,.2f})</td>
                <td class="num text-danger">({adv:,.2f})</td>
                <td class="num text-danger">({mat:,.2f})</td>
                <td class="num font-bold text-danger">({total_ded:,.2f})</td>
                <td class="num font-bold text-primary">{inv.current_net_before_vat:,.2f}</td>
            </tr>
            """

        # === CUMULATIVE SUMMARY DATA ===
        ret_a_cum = latest_inv.cumulative_retention_a if latest_inv else Decimal("0")
        ret_b_cum = latest_inv.cumulative_retention_b if latest_inv else Decimal("0")
        ret_a_rec = latest_inv.cumulative_retention_a_recovered if latest_inv else Decimal("0")
        ret_b_rec = latest_inv.cumulative_retention_b_recovered if latest_inv else Decimal("0")
        advance_rec = latest_inv.cumulative_advance_recovered if latest_inv else Decimal("0")
        advance_total = proj.total_advance_value

        back_charges = getattr(proj, 'back_charges', Decimal("0"))
        est_back_charges = getattr(proj, 'estimated_back_charges', Decimal("0"))
        liquidated = getattr(proj, 'liquidated_damages', Decimal("0"))

        net_ret_held = (ret_a_cum + ret_b_cum) - (ret_a_rec + ret_b_rec)
        total_other_ded = back_charges + est_back_charges + liquidated
        grand_total_deductions = cumulative_materials + net_ret_held + advance_rec + total_other_ded

        # === MATERIALS DETAIL TABLE ===
        materials_rows = ""
        for inv in invoices:
            mat = inv.material_supplied_by_client or Decimal("0")
            if mat > 0:
                materials_rows += f"""
                <tr>
                    <td class="center">{inv.inv_number}</td>
                    <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                    <td class="text"><b>{inv}</b></td>
                    <td class="num font-bold text-danger">({mat:,.2f})</td>
                    <td class="text">Materials supplied by client — {inv.project.project_name}</td>
                </tr>
                """

        # === RETENTION RECOVERY TRACKING ===
        recovery_invoices = Invoice.objects.filter(
            project=proj,
            retention_recovery__in=['A', 'B']
        ).order_by('inv_number')

        recovery_rows = ""
        for inv in recovery_invoices:
            if inv.retention_recovery == 'A':
                rec_type = "Retention A"
                rec_amount = inv.current_retention_a_recovery
                cum_rec = inv.cumulative_retention_a_recovered
            else:
                rec_type = "Retention B"
                rec_amount = inv.current_retention_b_recovery
                cum_rec = inv.cumulative_retention_b_recovered

            recovery_rows += f"""
            <tr style="background: #e8f5e9;">
                <td class="center">{inv.inv_number}</td>
                <td class="center">{inv.date.strftime('%d-%b-%Y')}</td>
                <td class="text"><b>{rec_type} Recovery</b></td>
                <td class="num font-bold text-success">{rec_amount:,.2f}</td>
                <td class="num text-success">{cum_rec:,.2f}</td>
            </tr>
            """

        body = f"""
        {self._build_meta_grid({
            'Project': f"{proj.project_id_code} — {proj.project_name}",
            'Client': proj.client.name,
            'PO Amount': f"AED {proj.po_amount:,.2f}",
            'Report Date': date.today().strftime('%d-%b-%Y'),
            'Company': company.company_name if company else 'N/A'
        })}

        <!-- Executive Summary -->
        <div class="card" style="background: #1a237e; color: white; margin-bottom: 20px;">
            <div class="card-header" style="color: white; border-color: rgba(255,255,255,0.3);">
                <span class="icon">📊</span> DEDUCTIONS EXECUTIVE SUMMARY
            </div>
            <div class="grid-4">
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {cumulative_materials:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Materials Supplied</div>
                    <div style="font-size: 8px; color: rgba(255,255,255,0.6); margin-top: 4px;">Client-supplied materials (VAT excl.)</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {net_ret_held:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Net Retention Held</div>
                    <div style="font-size: 8px; color: rgba(255,255,255,0.6); margin-top: 4px;">
                        A: {money(ret_a_cum - ret_a_rec):,.2f} | B: {money(ret_b_cum - ret_b_rec):,.2f}
                    </div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {advance_rec:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Advance Recovered</div>
                    <div style="font-size: 8px; color: rgba(255,255,255,0.6); margin-top: 4px;">of {advance_total:,.2f} total advance</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.1); border: none; color: white;">
                    <div class="metric-value" style="color: white;">AED {total_other_ded:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.8);">Other Deductions</div>
                    <div style="font-size: 8px; color: rgba(255,255,255,0.6); margin-top: 4px;">Back charges, LD, estimates</div>
                </div>
            </div>
            <div style="text-align: center; margin-top: 16px; padding-top: 16px; border-top: 1px solid rgba(255,255,255,0.2);">
                <div style="font-size: 11px; text-transform: uppercase; letter-spacing: 1px; opacity: 0.8; margin-bottom: 8px;">Grand Total Deductions</div>
                <div style="font-size: 28px; font-weight: 700;">AED {grand_total_deductions:,.2f}</div>
                <div style="font-size: 10px; opacity: 0.7; margin-top: 4px;">
                    from PO Amount of AED {proj.po_amount:,.2f} 
                    ({(grand_total_deductions / proj.po_amount * 100) if proj.po_amount > 0 else Decimal('0'):.1f}% of contract)
                </div>
            </div>
        </div>

        <!-- Invoice-Level Deductions Table -->
        <div class="card" style="margin-bottom: 20px;">
            <div class="card-header">
                <span class="icon">📋</span> DEDUCTIONS BY INVOICE
            </div>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Inv #</th>
                        <th>Date</th>
                        <th>Reference</th>
                        <th>Status</th>
                        <th class="num">Gross Work</th>
                        <th class="num">Ret A</th>
                        <th class="num">Ret B</th>
                        <th class="num">Adv Rec</th>
                        <th class="num">Materials</th>
                        <th class="num">Total Deduction</th>
                        <th class="num">Net Payable</th>
                    </tr>
                </thead>
                <tbody>
                    {inv_rows if inv_rows else '<tr><td colspan="11" style="text-align:center; color:#999; padding:20px;">No tax invoices found</td></tr>'}
                </tbody>
                <tfoot>
                    <tr class="total-row">
                        <td colspan="4"><b>CUMULATIVE TOTALS</b></td>
                        <td class="num"><b>{sum(inv.current_gross_total for inv in invoices):,.2f}</b></td>
                        <td class="num text-danger"><b>({cumulative_ret_a:,.2f})</b></td>
                        <td class="num text-danger"><b>({cumulative_ret_b:,.2f})</b></td>
                        <td class="num text-danger"><b>({cumulative_adv:,.2f})</b></td>
                        <td class="num text-danger"><b>({cumulative_materials:,.2f})</b></td>
                        <td class="num font-bold text-danger"><b>({cumulative_total_ded:,.2f})</b></td>
                        <td class="num font-bold text-primary"><b>{sum(inv.current_net_before_vat for inv in invoices):,.2f}</b></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <!-- Materials Detail Section -->
        <div class="card" style="margin-bottom: 20px; page-break-inside: avoid;">
            <div class="card-header" style="color: #c62828; border-color: #ffcdd2;">
                <span class="icon">📦</span> MATERIALS SUPPLIED BY CLIENT — DETAIL
            </div>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Inv #</th>
                        <th>Date</th>
                        <th>Invoice Reference</th>
                        <th class="num">Amount (VAT excl.)</th>
                        <th>Description / Notes</th>
                    </tr>
                </thead>
                <tbody>
                    {materials_rows if materials_rows else '<tr><td colspan="5" style="text-align:center; color:#999; padding:20px;">No materials recorded for this project</td></tr>'}
                </tbody>
                <tfoot>
                    <tr class="total-row">
                        <td colspan="3"><b>TOTAL MATERIALS DEDUCTED</b></td>
                        <td class="num font-bold text-danger"><b>({cumulative_materials:,.2f})</b></td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
            <div style="margin-top: 10px; padding: 10px; background: #ffebee; border-radius: 4px; font-size: 9px; color: #c62828;">
                <b>Note:</b> Materials supplied by client are deducted from the gross work done before VAT calculation. 
                The VAT on these materials (input VAT) is separately accounted for and does not form part of the contractor's output VAT liability.
            </div>
        </div>

        <!-- Retention & Advance Tracking -->
        <div class="grid-2" style="margin-bottom: 20px;">
            <div class="card">
                <div class="card-header" style="color: #ed6c02; border-color: #ffe0b2;">
                    <span class="icon">🔒</span> RETENTION TRACKING
                </div>
                <table class="data-table" style="font-size: 9px;">
                    <tbody>
                        <tr>
                            <td style="width: 60%;"><b>Retention A Deducted (Cumulative)</b></td>
                            <td class="num text-danger">({ret_a_cum:,.2f})</td>
                        </tr>
                        <tr>
                            <td><b>Retention A Recovered</b></td>
                            <td class="num text-success">{ret_a_rec:,.2f}</td>
                        </tr>
                        <tr style="background: #fff3e0;">
                            <td><b>Retention A Net Held</b></td>
                            <td class="num font-bold text-warning">({money(ret_a_cum - ret_a_rec):,.2f})</td>
                        </tr>
                        <tr><td colspan="2" style="height: 8px;"></td></tr>
                        <tr>
                            <td><b>Retention B Deducted (Cumulative)</b></td>
                            <td class="num text-danger">({ret_b_cum:,.2f})</td>
                        </tr>
                        <tr>
                            <td><b>Retention B Recovered</b></td>
                            <td class="num text-success">{ret_b_rec:,.2f}</td>
                        </tr>
                        <tr style="background: #fff3e0;">
                            <td><b>Retention B Net Held</b></td>
                            <td class="num font-bold text-warning">({money(ret_b_cum - ret_b_rec):,.2f})</td>
                        </tr>
                        <tr><td colspan="2" style="height: 8px;"></td></tr>
                        <tr class="total-row">
                            <td><b>TOTAL NET RETENTION HELD</b></td>
                            <td class="num font-bold text-danger">({net_ret_held:,.2f})</td>
                        </tr>
                    </tbody>
                </table>
            </div>

            <div class="card">
                <div class="card-header" style="color: #1565c0; border-color: #bbdefb;">
                    <span class="icon">💰</span> ADVANCE TRACKING
                </div>
                <table class="data-table" style="font-size: 9px;">
                    <tbody>
                        <tr>
                            <td style="width: 60%;"><b>Total Advance Value</b></td>
                            <td class="num">{advance_total:,.2f}</td>
                        </tr>
                        <tr>
                            <td><b>Advance Paid (Project Level)</b></td>
                            <td class="num">{getattr(proj, 'advance_paid', Decimal('0')):,.2f}</td>
                        </tr>
                        <tr style="background: #e3f2fd;">
                            <td><b>Advance Recovered (Cumulative)</b></td>
                            <td class="num font-bold text-success">{advance_rec:,.2f}</td>
                        </tr>
                        <tr><td colspan="2" style="height: 8px;"></td></tr>
                        <tr class="total-row">
                            <td><b>ADVANCE BALANCE REMAINING</b></td>
                            <td class="num font-bold text-primary">{money(advance_total - advance_rec):,.2f}</td>
                        </tr>
                    </tbody>
                </table>
            </div>
        </div>

        <!-- Retention Recovery History -->
        <div class="card" style="margin-bottom: 20px; page-break-inside: avoid;">
            <div class="card-header" style="color: #2e7d32; border-color: #c8e6c9;">
                <span class="icon">✅</span> RETENTION RECOVERY HISTORY
            </div>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Inv #</th>
                        <th>Date</th>
                        <th>Type</th>
                        <th class="num">Amount Recovered</th>
                        <th class="num">Cumulative Recovered</th>
                    </tr>
                </thead>
                <tbody>
                    {recovery_rows if recovery_rows else '<tr><td colspan="5" style="text-align:center; color:#999; padding:20px;">No retention recovery invoices yet</td></tr>'}
                </tbody>
            </table>
        </div>

        <!-- Other Deductions -->
        <div class="card" style="margin-bottom: 20px; page-break-inside: avoid;">
            <div class="card-header" style="color: #6a1b9a; border-color: #e1bee7;">
                <span class="icon">⚠️</span> OTHER DEDUCTIONS & CHARGES
            </div>
            <table class="data-table">
                <thead>
                    <tr>
                        <th>Category</th>
                        <th class="num">Amount</th>
                        <th>Status</th>
                        <th>Notes</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td><b>Back Charges / Contra-Charges</b></td>
                        <td class="num text-danger">({back_charges:,.2f})</td>
                        <td><span class="badge badge-danger">Permanent</span></td>
                        <td>Deducted from final settlement</td>
                    </tr>
                    <tr>
                        <td><b>Estimated Back Charges</b></td>
                        <td class="num text-danger">({est_back_charges:,.2f})</td>
                        <td><span class="badge badge-warning">Provisional</span></td>
                        <td>Estimated pending final assessment</td>
                    </tr>
                    <tr>
                        <td><b>Liquidated Damages</b></td>
                        <td class="num text-danger">({liquidated:,.2f})</td>
                        <td><span class="badge badge-danger">Permanent</span></td>
                        <td>As per contract penalty clauses</td>
                    </tr>
                </tbody>
                <tfoot>
                    <tr class="total-row">
                        <td><b>TOTAL OTHER DEDUCTIONS</b></td>
                        <td class="num font-bold text-danger"><b>({total_other_ded:,.2f})</b></td>
                        <td colspan="2"></td>
                    </tr>
                </tfoot>
            </table>
        </div>

        <!-- Final Summary -->
        <div class="card" style="background: #fafafa; border: 2px solid #1a237e;">
            <div class="card-header">
                <span class="icon">📊</span> DEDUCTIONS IMPACT ON CONTRACT
            </div>
            <table class="data-table" style="font-size: 10px;">
                <tbody>
                    <tr>
                        <td style="width: 50%;"><b>Original Contract Value (PO Amount)</b></td>
                        <td class="num font-bold">{proj.po_amount:,.2f}</td>
                    </tr>
                    <tr>
                        <td><b>Amendments</b></td>
                        <td class="num">{getattr(proj, 'amendment_amount', Decimal('0')):,.2f}</td>
                    </tr>
                    <tr>
                        <td><b>Variations</b></td>
                        <td class="num">{getattr(proj, 'variation_amount', Decimal('0')):,.2f}</td>
                    </tr>
                    <tr style="background: #e8eaf6;">
                        <td><b>Final Contract Value</b></td>
                        <td class="num font-bold">{proj.po_amount + getattr(proj, 'amendment_amount', Decimal('0')) + getattr(proj, 'variation_amount', Decimal('0')):,.2f}</td>
                    </tr>
                    <tr><td colspan="2" style="height: 8px;"></td></tr>
                    <tr>
                        <td><b>Less: Materials Supplied by Client</b></td>
                        <td class="num text-danger">({cumulative_materials:,.2f})</td>
                    </tr>
                    <tr>
                        <td><b>Less: Net Retention Held</b></td>
                        <td class="num text-danger">({net_ret_held:,.2f})</td>
                    </tr>
                    <tr>
                        <td><b>Less: Advance Recovered</b></td>
                        <td class="num text-danger">({advance_rec:,.2f})</td>
                    </tr>
                    <tr>
                        <td><b>Less: Back Charges & Liquidated Damages</b></td>
                        <td class="num text-danger">({total_other_ded:,.2f})</td>
                    </tr>
                    <tr style="background: #ffebee;">
                        <td><b>TOTAL DEDUCTIONS</b></td>
                        <td class="num font-bold text-danger">({grand_total_deductions:,.2f})</td>
                    </tr>
                    <tr><td colspan="2" style="height: 8px;"></td></tr>
                    <tr style="background: #e8f5e9;">
                        <td><b>NET CONTRACT VALUE AFTER DEDUCTIONS</b></td>
                        <td class="num font-bold text-success">
                            {proj.po_amount + getattr(proj, 'amendment_amount', Decimal('0')) + getattr(proj, 'variation_amount', Decimal('0')) - grand_total_deductions:,.2f}
                        </td>
                    </tr>
                </tbody>
            </table>
        </div>

        <!-- Signatures -->
        <div class="signature-grid" style="margin-top: 40px;">
            <div class="signature-block">
                <div class="signature-line">Finance Director</div>
            </div>
            <div class="signature-block">
                <div class="signature-line">Technical Manager</div>
            </div>
            <div class="signature-block">
                <div class="signature-line">General Manager</div>
            </div>
        </div>
        """

        return HttpResponse(self._report_base_wrapper(
            "PROJECT DEDUCTIONS REPORT",
            f"Complete Deductions Tracking — {proj.project_id_code}",
            body,
            logo_url
        ))
    ######################
    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('<int:pk>/analytics/', self.admin_site.admin_view(self.analytics_view), name='project_analytics'),
            path('<int:pk>/cost-profit/', self.admin_site.admin_view(self.cost_profit_view),
                 name='project_cost_profit'),
            path('<int:pk>/statement/', self.admin_site.admin_view(self.project_statement_view),
                 name='project_statement'),
            path('<int:pk>/deductions/', self.admin_site.admin_view(self.deductions_view), name='project_deductions'),
            # NEW
        ]
        return custom + urls

    def fmt_po(self, obj):
        return mark_safe(f'<div style="text-align: right; font-weight: bold;">{obj.po_amount:,.2f}</div>')
    fmt_po.short_description = 'PO Amount'
    fmt_po.admin_order_field = 'po_amount'

    def fmt_boq_total(self, obj):
        val = obj.boq_total_value
        color = "green" if obj.is_boq_complete else "red"
        return mark_safe(f'<span style="color: {color}; font-weight: bold; float: right;">{val:,.2f}</span>')
    fmt_boq_total.short_description = "BOQ Total"

    def fmt_advance(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.advance_percent:,.2f}%</div>')
    fmt_advance.short_description = "Adv %"
    fmt_advance.admin_order_field = "advance_percent"

    def fmt_ret_a_pct(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.retention_a_percent:,.2f}%</div>')
    fmt_ret_a_pct.short_description = "Ret A %"
    fmt_ret_a_pct.admin_order_field = "retention_a_percent"

    def fmt_ret_b_pct(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.retention_b_percent:,.2f}%</div>')
    fmt_ret_b_pct.short_description = "Ret B %"
    fmt_ret_b_pct.admin_order_field = "retention_b_percent"



# =============================================================================
# INVOICE ADMIN
# =============================================================================

class InvoiceItemInline(admin.TabularInline):
    model = InvoiceItem
    extra = 0
    readonly_fields = ["rate", "fmt_prev_amt", "fmt_prev_pct", "fmt_gross", "fmt_cum_pct", "fmt_cum_amt", "fmt_ret_a",
                       "fmt_ret_b"]
    fields = [
        "boq_item", "billing_method", "rate",
        "fmt_prev_pct", "fmt_prev_amt",
        "current_percentage", "current_qty", "fmt_gross",
        "fmt_cum_pct", "fmt_cum_amt", "fmt_ret_a", "fmt_ret_b"
    ]

    def fmt_prev_pct(self, obj):
        return mark_safe(f'<div style="text-align:right;">{obj.prev_percentage:,.2f}%</div>')
    fmt_prev_pct.short_description = "Prev %"

    def fmt_prev_amt(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.prev_amount:,.2f}</div>')
    fmt_prev_amt.short_description = "Prev Amt"

    def fmt_gross(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.gross_amount:,.2f}</div>')
    fmt_gross.short_description = "Curr Amt"

    def fmt_cum_pct(self, obj):
        return mark_safe(f'<div style="text-align:right;">{obj.prev_percentage + obj.current_percentage:,.2f}%</div>')
    fmt_cum_pct.short_description = "Cum. %"

    def fmt_cum_amt(self, obj):
        return mark_safe(
            f'<div style="text-align:right;font-weight:bold;">{obj.prev_amount + obj.gross_amount:,.2f}</div>')
    fmt_cum_amt.short_description = "Cum. Amt"

    def fmt_ret_a(self, obj):
        return mark_safe(f'<div style="text-align:right;color:#ed6c02;">{obj.retention_a_amount:,.2f}</div>')
    fmt_ret_a.short_description = "Ret A"

    def fmt_ret_b(self, obj):
        return mark_safe(f'<div style="text-align:right;color:#ed6c02;">{obj.retention_b_amount:,.2f}</div>')
    fmt_ret_b.short_description = "Ret B"

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'boq_item':
            if hasattr(request, '_invoice_project_id'):
                kwargs['queryset'] = BOQItem.objects.filter(project_id=request._invoice_project_id)
            else:
                kwargs['queryset'] = BOQItem.objects.none()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

class VariationOrderInline(admin.TabularInline):
    model = VariationOrder
    extra = 0
    fields = ['description', 'amount']

@admin.register(Invoice)
class InvoiceAdmin(CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'project__company'
    inlines = [InvoiceItemInline, VariationOrderInline ]
    list_display = [
        "fmt_inv_str", "project", "inv_type", "status", "is_final_invoice", "retention_recovery",
        "ui_cumulative_work", "ui_total_invoiced", "ui_previously_invoiced", "ui_subtotal_no_vat",
        "ui_vat", "ui_total_before_deductions", "ui_payable", "print_button", "ui_variation_total",
        "ui_total_with_variations",
    ]
    list_filter = ["status", "inv_type", "is_final_invoice", "project", "retention_recovery"]
    search_fields = ["inv_number", "project__project_id_code"]
    readonly_fields = [
        "ui_cumulative_work", "ui_previous_work", "ui_current_work",
        "ui_advance_recovery", "ui_retention_a", "ui_retention_b",
        "ui_retention_a_recovery", "ui_retention_b_recovery",
        "ui_total_invoiced", "ui_previously_invoiced", "ui_subtotal_no_vat",
        "ui_vat", "ui_total_before_deductions", "ui_payable",
        "ui_advance_balance", "ui_final_advance_recovery", "ui_remaining_advance",
        "ui_variation_total", "ui_total_with_variations",
    ]

    fieldsets = (
        ("Invoice Details", {
            "fields": ("project", ("inv_type", "status"), ("date", "inv_number"), "revision",
                       ("is_advance_invoice", "is_final_invoice", "retention_recovery"),
                       ("collection_date", "payment_date"))
        }),
        ("Calculated Billing Summary", {
            "fields": (
                "ui_cumulative_work", "ui_total_invoiced", "ui_previously_invoiced",
                "ui_subtotal_no_vat", "vat_percent", "ui_vat", "ui_total_before_deductions",
                "material_supplied_by_client", "ui_payable"
            )
        }),
        ("Advance & Retention Details", {
            "fields": (
                "ui_advance_balance", "ui_final_advance_recovery", "ui_remaining_advance",
                "ui_advance_recovery", "ui_retention_a", "ui_retention_b",
                "ui_retention_a_recovery", "ui_retention_b_recovery"
            ),
            "classes": ["collapse"]
        }),
    )

    def ui_variation_total(self, obj):
        total = obj.variation_orders.aggregate(total=Sum('amount'))['total'] or Decimal("0")
        if total > 0:
            return mark_safe(f'<div style="text-align:right;font-weight:bold;color:#6a1b9a;">{total:,.2f}</div>')
        return mark_safe('<span style="color:#999;">—</span>')

    ui_variation_total.short_description = "Variations"

    def ui_total_with_variations(self, obj):
        base = obj.current_net_before_vat
        var_total = obj.variation_orders.aggregate(total=Sum('amount'))['total'] or Decimal("0")
        total = base + var_total
        return mark_safe(f'<div style="text-align:right;font-weight:bold;color:#1a237e;">{total:,.2f}</div>')

    ui_total_with_variations.short_description = "Total + Variations"

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if obj and obj.project:
            request._invoice_project_id = obj.project_id
        return form

    def ui_cumulative_work(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.cumulative_work_done:,.2f}</div>')
    ui_cumulative_work.short_description = "Total Certified Work"

    def ui_total_invoiced(self, obj):
        return mark_safe(
            f'<div style="text-align:right;font-weight:bold;">{obj.net_total_invoiced_cumulative:,.2f}</div>')
    ui_total_invoiced.short_description = "Total Invoiced (Cum)"

    def ui_previously_invoiced(self, obj):
        return mark_safe(f'<div style="text-align:right;color:#666;">({obj.previous_net_total_invoiced:,.2f})</div>')
    ui_previously_invoiced.short_description = "Prev Invoiced"

    def ui_subtotal_no_vat(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.current_net_before_vat:,.2f}</div>')
    ui_subtotal_no_vat.short_description = "Sub Total No VAT"

    def ui_vat(self, obj):
        return mark_safe(f'<div style="text-align:right;">{obj.vat_amount:,.2f}</div>')
    ui_vat.short_description = "VAT"

    def ui_total_before_deductions(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.total_with_vat:,.2f}</div>')
    ui_total_before_deductions.short_description = "Total Before Ded."

    def ui_payable(self, obj):
        return mark_safe(
            f'<div style="text-align:right;"><b style="color:#d32f2f;">{obj.total_after_vat:,.2f}</b></div>')
    ui_payable.short_description = "Payable"

    def ui_previous_work(self, obj):
        return "{:,.2f}".format(obj.previous_work_done)

    def ui_current_work(self, obj):
        return "{:,.2f}".format(obj.current_gross_total)

    def ui_advance_recovery(self, obj):
        return "({:,.2f})".format(obj.current_advance_recovery)

    def ui_retention_a(self, obj):
        return "({:,.2f})".format(obj.current_retention_a)

    def ui_retention_b(self, obj):
        return "({:,.2f})".format(obj.current_retention_b)

    def ui_retention_a_recovery(self, obj):
        if obj.retention_recovery == 'A':
            val = obj.current_retention_a_recovery
            return format_html("<b style='color:green;'>{:,.2f}</b>", val)
        return "0.00"
    ui_retention_a_recovery.short_description = "Ret A Recovery"

    def ui_retention_b_recovery(self, obj):
        if obj.retention_recovery == 'B':
            val = obj.current_retention_b_recovery
            return format_html("<b style='color:green;'>{:,.2f}</b>", val)
        return "0.00"
    ui_retention_b_recovery.short_description = "Ret B Recovery"

    def ui_advance_balance(self, obj):
        if obj.was_advance_taken:
            remaining = money(obj.project.total_advance_value - obj.previous_advance_recovered)
            if obj.is_final_invoice and remaining > 0:
                return mark_safe(
                    f'<div style="text-align:right;color:#ed6c02;font-weight:bold;">'
                    f'{remaining:,.2f} <span style="font-size:9px;">(will be recovered)</span></div>'
                )
            return mark_safe(f'<div style="text-align:right;">{remaining:,.2f}</div>')
        return mark_safe('<span style="color:#999;">—</span>')

    ui_advance_balance.short_description = "Advance Balance (Remaining)"

    def ui_final_advance_recovery(self, obj):
        if obj.is_final_invoice and obj.was_advance_taken:
            recovery = obj.current_advance_recovery
            return mark_safe(
                f'<div style="text-align:right;color:#d32f2f;font-weight:bold;">'
                f'({recovery:,.2f}) <span style="font-size:9px;">FINAL RECOVERY</span></div>'
            )
        elif obj.was_advance_taken:
            return mark_safe(f'<div style="text-align:right;">({obj.current_advance_recovery:,.2f})</div>')
        return mark_safe('<span style="color:#999;">—</span>')

    ui_final_advance_recovery.short_description = "Current Advance Recovery"

    def ui_remaining_advance(self, obj):
        if obj.was_advance_taken:
            remaining = obj.advance_balance_remaining
            if remaining == 0:
                return mark_safe(
                    f'<div style="text-align:right;color:#2e7d32;font-weight:bold;">'
                    f'{remaining:,.2f} <span style="font-size:9px;">✓ FULLY RECOVERED</span></div>'
                )
            return mark_safe(f'<div style="text-align:right;color:#ed6c02;">{remaining:,.2f}</div>')
        return mark_safe('<span style="color:#999;">—</span>')

    ui_remaining_advance.short_description = "Advance Balance After This Invoice"

    def fmt_inv_str(self, obj):
        badge = ""
        if obj.is_final_invoice:
            badge = '<span style="background:#d32f2f;color:white;padding:1px 4px;border-radius:3px;font-size:8px;margin-left:4px;">FINAL</span>'
        return format_html(
            '<div style="font-weight:bold;width:120px;">{} {}</div>',
            str(obj), mark_safe(badge)
        )
    fmt_inv_str.short_description = "Invoice ID"

    def print_button(self, obj):
        url = reverse('admin:invoice_print', args=[obj.pk])
        return format_html(
            '<a class="button" href="{}" target="_blank" style="background:#447e9b; color:white; padding: 2px 8px; border-radius: 4px;">Print</a>',
            url)
    print_button.short_description = "Report"

    def get_urls(self):
        urls = super().get_urls()
        return [path('<int:pk>/print/', self.admin_site.admin_view(self.print_view), name='invoice_print')] + urls

    def print_view(self, request, pk):
        company = self.get_active_company(request)
        inv = self.get_object_or_404_scoped(request, Invoice, pk=pk)
        logo_url = company.logo.url if company and company.logo else ''
        header_title = "PROFORMA INVOICE" if inv.inv_type == "P" else "TAX INVOICE"

        all_boq = BOQItem.objects.filter(project=inv.project).order_by('item_number')
        items_map = {item.boq_item.id: item for item in inv.items.all()}
        rows = ""
        totals = {"prev": Decimal("0"), "curr": Decimal("0"), "cum": Decimal("0")}

        for boq in all_boq:
            item = items_map.get(boq.id)
            if item:
                p_amt, p_pct = item.prev_amount, item.prev_percentage
                c_amt, c_pct = item.gross_amount, item.current_percentage
            else:
                prev_data = InvoiceItem.objects.filter(
                    invoice__project=inv.project,
                    invoice__date__lt=inv.date,
                    boq_item=boq
                ).aggregate(p_amt=Sum('gross_amount'))
                p_amt = prev_data['p_amt'] or Decimal("0")
                p_pct = (p_amt / (boq.quantity * boq.rate) * 100) if (boq.quantity * boq.rate) > 0 else Decimal("0")
                c_pct, c_amt = Decimal("0"), Decimal("0")

            totals["prev"] += p_amt
            totals["curr"] += c_amt
            totals["cum"] += (p_amt + c_amt)

            rows += f"""<tr>
                <td class='col-item'>{boq.item_number}</td>
                <td class='col-desc'>{boq.description}</td>
                <td class='col-unit'>{boq.unit}</td>
                <td class='col-num'>{boq.quantity:,.2f}</td>
                <td class='col-num'>{boq.rate:,.0f}</td>
                <td class='col-num'>{p_pct:,.0f}%</td>
                <td class='col-num'>{p_amt:,.2f}</td>
                <td class='col-num'>{c_pct:,.0f}%</td>
                <td class='col-num'>{c_amt:,.2f}</td>
                <td class='col-num'>{(p_pct + c_pct):,.0f}%</td>
                <td class='col-num'>{(p_amt + c_amt):,.2f}</td>
            </tr>"""

        # ─────────────────────────────────────────────────────────────
        # VARIATION ORDERS — calculated early for use in all pages
        # ─────────────────────────────────────────────────────────────
        variation_rows = ""
        var_total = Decimal("0")
        for vo in inv.variation_orders.all():
            var_total += vo.amount
            variation_rows += f"""
                <tr style='background:#0000;'>
                    <td colspan='6' class='col-label'>Variation: {vo.description[:60]}{'...' if len(vo.description) > 60 else ''}</td>
                    <td></td><td></td><td></td><td></td>
                    <td class='col-num' style='color:#6a1b9a; font-weight:bold;'>{vo.amount:,.2f}</td>
                </tr>
                """

        # DETECT RETENTION RECOVERY
        is_retention_recovery = inv.retention_recovery in ['A', 'B']
        recovery_amount = Decimal("0")
        recovery_type = ""

        if is_retention_recovery:
            if inv.retention_recovery == 'A':
                recovery_amount = inv.current_retention_a_recovery
                recovery_type = "Retention A Recovery"
            else:
                recovery_amount = inv.current_retention_b_recovery
                recovery_type = "Retention B Recovery"

        # KEY FIX: For retention recovery, the "net before VAT" is the recovery amount
        # For normal invoices, it's gross minus deductions PLUS variations
        if is_retention_recovery:
            net_before_vat = recovery_amount
        else:
            # FIX: Include var_total in net_before_vat
            net_before_vat = inv.current_gross_total - inv.current_advance_recovery - inv.current_retention_a - inv.current_retention_b + var_total

        is_credit_note = net_before_vat < 0

        # VAT Output: calculated on net_before_vat (which now includes variations)
        vat_rate = inv.vat_percent or Decimal("5")
        vat_output = net_before_vat * vat_rate / Decimal("100")

        # Materials & VAT Input
        materials_exclusive = inv.materials_exclusive
        vat_input = money(materials_exclusive * vat_rate / Decimal("100"))

        # Net VAT = Output - Input (can be negative)
        net_vat = vat_output - vat_input

        # Payable = Net Before VAT - Materials + Net VAT
        payable_amount = net_before_vat - materials_exclusive + net_vat

        # Formatting helpers for negative values
        def fmt_val(val):
            if val < 0:
                return f'<span style="color: #d15353;">({abs(val):,.2f})</span>'
            return f'{val:,.2f}'

        def fmt_val_bold(val):
            if val < 0:
                return f'<span style="color: #c62828; font-weight: bold;">({abs(val):,.2f})</span>'
            return f'<b>{val:,.2f}</b>'

        # Styling based on credit note
        if is_credit_note:
            header_title = "CREDIT NOTE" if inv.inv_type == "T" else "PROFORMA CREDIT NOTE"
            net_color = '#c62828'
            vat_color = '#c62828'
            vat_sign = ''
            grand_bg = '#c62828'
        else:
            net_color = '#000000'
            vat_color = '#1565c0'
            vat_sign = '+'
            grand_bg = '#1a237e'

        # Build variation summary lines for detail page footer
        variation_footer = ""
        if var_total > 0:
            variation_footer += f"""
            <tr style='background:##0000;'>
                <td colspan='6' class='col-label'><b>Total Variations</b></td>
                <td></td><td></td><td></td><td></td>
                <td class='col-num' style='color:#6a1b9a; font-weight:bold;'>{var_total:,.2f}</td>
            </tr>
            """

        # ─────────────────────────────────────────────────────────────
        # DETAIL PAGE FOOTER ROWS — FIX: variations BEFORE net total
        # ─────────────────────────────────────────────────────────────
        footer_rows = f"""
            <tr class='total-row'>
                <td colspan='6' class='col-label'>GROSS WORK DONE</td>
                <td class='col-num'>{totals['prev']:,.2f}</td>
                <td></td><td class='col-num'>{totals['curr']:,.2f}</td>
                <td></td><td class='col-num'>{totals['cum']:,.2f}</td>
            </tr>
        """

        # Only show advance/retention deduction lines for non-recovery invoices
        if not is_retention_recovery:
            footer_rows += f"""
            <tr>
                <td colspan='6' class='col-label'>Advance Recovery ({inv.project.advance_percent}%)</td>
                <td class='col-num'>({inv.previous_advance_recovered:,.2f})</td>
                <td></td><td class='col-num'>({inv.current_advance_recovery:,.2f})</td>
                <td></td><td class='col-num'>({inv.cumulative_advance_recovered:,.2f})</td>
            </tr>
            <tr>
                <td colspan='6' class='col-label'>Retention A ({inv.project.retention_a_percent}%)</td>
                <td class='col-num'>({inv.previous_retention_a:,.2f})</td>
                <td></td><td class='col-num'>({inv.current_retention_a:,.2f})</td>
                <td></td><td class='col-num'>({inv.cumulative_retention_a:,.2f})</td>
            </tr>
            <tr>
                <td colspan='6' class='col-label'>Retention B ({inv.project.retention_b_percent}%)</td>
                <td class='col-num'>({inv.previous_retention_b:,.2f})</td>
                <td></td><td class='col-num'>({inv.current_retention_b:,.2f})</td>
                <td></td><td class='col-num'>({inv.cumulative_retention_b:,.2f})</td>
            </tr>
            """
        else:
            footer_rows += f"""
            <tr style='background:#e8f5e9;'>
                <td colspan='6' class='col-label'><b>{recovery_type}</b></td>
                <td class='col-num'>—</td>
                <td></td><td class='col-num'><b style='color:#2e7d32;'>{recovery_amount:,.2f}</b></td>
                <td></td><td class='col-num'><b>{recovery_amount:,.2f}</b></td>
            </tr>
            """

        # Show retention recovery in footer if applicable (existing logic)
        if inv.retention_recovery == 'A' and not is_retention_recovery:
            footer_rows += f"""
            <tr style='background:#e8f5e9;'>
                <td colspan='6' class='col-label'><b>Retention A Recovery</b></td>
                <td class='col-num'>({inv.previous_retention_a_recovered:,.2f})</td>
                <td></td><td class='col-num'><b>{inv.current_retention_a_recovery:,.2f}</b></td>
                <td></td><td class='col-num'>({inv.cumulative_retention_a_recovered:,.2f})</td>
            </tr>
            """

        if inv.retention_recovery == 'B' and not is_retention_recovery:
            footer_rows += f"""
            <tr style='background:#e8f5e9;'>
                <td colspan='6' class='col-label'><b>Retention B Recovery</b></td>
                <td class='col-num'>({inv.previous_retention_b_recovered:,.2f})</td>
                <td></td><td class='col-num'><b>{inv.current_retention_b_recovery:,.2f}</b></td>
                <td></td><td class='col-num'>({inv.cumulative_retention_b_recovered:,.2f})</td>
            </tr>
            """

        # FIX: Add variation rows BEFORE the NET BEFORE VAT line
        footer_rows += variation_rows + variation_footer

        # Cumulative net invoiced for display in footer
        if is_retention_recovery:
            cum_net_display = inv.previous_net_total_invoiced + recovery_amount
        else:
            # FIX: Include var_total in cumulative net display
            cum_net_display = inv.previous_net_total_invoiced + net_before_vat

        footer_rows += f"""
            <tr class='grand-total-row'>
                <td colspan='6' class='col-label'><b>NET BEFORE VAT</b></td>
                <td class='col-num'>{inv.previous_net_total_invoiced:,.2f}</td>
                <td></td><td class='col-num'>{fmt_val_bold(net_before_vat)}</td>
                <td></td><td class='col-num'>{fmt_val(cum_net_display)}</td>
            </tr>
            <tr style='background:#e3f2fd;'>
                <td colspan='6' class='col-label' style='color:{vat_color};'><b>VAT OUTPUT ({vat_rate}%)</b></td>
                <td class='col-num'>—</td>
                <td></td><td class='col-num' style='color:{vat_color}; font-weight:bold;'>{vat_sign}{fmt_val(vat_output)}</td>
                <td></td><td class='col-num' style='color:{vat_color};'>{vat_sign}{fmt_val(vat_output)}</td>
            </tr>
            <tr>
                <td colspan='6' class='col-label'>Materials Supplied by Client (VAT excl.)</td>
                <td class='col-num'>—</td>
                <td></td><td class='col-num'>({materials_exclusive:,.2f})</td>
                <td></td><td class='col-num'>({materials_exclusive:,.2f})</td>
            </tr>
            <tr style='background:#fff3e0;'>
                <td colspan='6' class='col-label' style='color:#e65100;'><b>VAT INPUT ON MATERIALS ({vat_rate}%)</b></td>
                <td class='col-num'>—</td>
                <td></td><td class='col-num' style='color:#e65100; font-weight:bold;'>{vat_input:,.2f}</td>
                <td></td><td class='col-num' style='color:#e65100;'>{vat_input:,.2f}</td>
            </tr>
            <tr class='grand-total-row' style='background:{grand_bg}; color:white;'>
                <td colspan='6' class='col-label'><b>NET VAT PAYABLE</b></td>
                <td class='col-num'>—</td>
                <td></td><td class='col-num'><b>{fmt_val(net_vat)}</b></td>
                <td></td><td class='col-num'>{fmt_val(net_vat)}</td>
            </tr>
            <tr class='grand-total-row' style='background:#000080; color:white;'>
                <td colspan='6' class='col-label'><b>TOTAL AMOUNT INCLUDING VAT</b></td>
                <td class='col-num'>{inv.previous_net_total_invoiced:,.2f}</td>
                <td></td><td class='col-num'><b>{fmt_val(payable_amount)}</b></td>
                <td></td><td class='col-num'>{fmt_val(inv.previous_net_total_invoiced + payable_amount)}</td>
            </tr>
        """

        # ─────────────────────────────────────────────────────────────
        # SHARED HELPERS
        # ─────────────────────────────────────────────────────────────
        logo_bar_html = f'<div class="logo-bar"><img src="{logo_url}" alt="Logo"></div>' if logo_url else ''

        if company and company.bank:
            bank_html = company.bank
        else:
            bank_html = (
                '<div class="bank-row"><b>Bank:</b> WIO Bank</div>'
                '<div class="bank-row"><b>Account Name:</b> PROCON GENERAL CONTRACTING LLC</div>'
                '<div class="bank-row"><b>Account No:</b> 9635743367</div>'
                '<div class="bank-row"><b>IBAN:</b> AE390860000009635743367</div>'
                '<div class="bank-row"><b>SWIFT:</b> WIOBAEADXXX</div>'
            )

        # Amount in words for cover page
        amount_words = self._number_to_words(abs(payable_amount))
        if payable_amount < 0:
            amount_words = f"(Minus {amount_words})"

        # ─────────────────────────────────────────────────────────────
        # COVER PAGE
        # ─────────────────────────────────────────────────────────────
        credit_badge = ''
        if is_credit_note:
            credit_badge = '<div style="background:#c62828; color:white; padding:4px 12px; border-radius:4px; font-size:11px; font-weight:bold; display:inline-block; margin-bottom:10px;">CREDIT NOTE</div>'

        cover_page = (
                '<div class="page cover-page">'
                + logo_bar_html +
                '<div class="cover-content">'
                + credit_badge +
                '<div class="cover-ref">'
                f'<div class="cover-ref-line"><span class="cover-label">REF :</span>{inv.project.po_number or "N/A"}</div>'
                f'<div class="cover-ref-line"><span class="cover-label">Date :</span>{inv.date.strftime("%d-%b-%Y")}</div>'
                '</div>'
                '<div class="cover-to">'
                '<div class="cover-to-line">To :</div>'
                f'<div class="cover-to-line">{inv.project.client.name}</div>'
                f'<div class="cover-to-line">{getattr(inv.project.client, "address", "") or ""}</div>'
                f'<div class="cover-to-line">{getattr(inv.project.client, "city", "") or ""}, {getattr(inv.project.client, "country", "UAE") or "UAE"}</div>'
                '</div>'
                '<div class="cover-attn">'
                f'<div class="cover-attn-line">Attn : {getattr(inv.project.client, "contact_person", "Finance Director") or "Finance Director"}</div>'
                '</div>'
                '<div class="cover-project">'
                f'<div class="cover-project-line"><span class="cover-label">Project :</span>{inv.project.project_name}</div>'
                '</div>'
                '<div class="cover-subject">'
                f'<div class="cover-subject-line"><span class="cover-label">Subject :</span>{header_title} NO : {inv}</div>'
                '</div>'
                '<div class="cover-body">'
                '<p>Dear Sir/Madam,</p>'
        )

        if is_credit_note:
            cover_page += f'<p>Please find attached {header_title} <b>{inv}</b> for a credit amount of <b style="color:#c62828;">AED ({abs(payable_amount):,.2f})</b> ({amount_words})</p>'
            cover_page += '<p>This credit note is issued due to deductions/adjustments exceeding the current period gross work value.</p>'
        else:
            cover_page += f'<p>Please find attached {header_title} <b>{inv}</b> for an amount of <b>AED {payable_amount:,.2f}</b> ({amount_words})</p>'

        cover_page += (
            '</div>'
            '<div class="cover-closing">'
            '<p>Thank you,</p>'
            '<p>Yours faithfully</p>'
            f'<p>For <b>{company.company_name if company else "PROCON GENERAL CONTRACTING LLC"}</b></p>'
            '<div class="cover-signature-space"></div>'
            f'<div class="cover-signature-name">{company.contact_person if company and hasattr(company, "contact_person") and company.contact_person else "Eng. Sherif Hemaya"}</div>'
            f'<div class="cover-signature-title">{company.contact_title if company and hasattr(company, "contact_title") and company.contact_title else "General Manager"}</div>'
            '</div>'
            '</div>'
            '</div>'
        )

        # ─────────────────────────────────────────────────────────────
        # SUMMARY PAGE — FIX: include variations in net calculation
        # ─────────────────────────────────────────────────────────────
        summary_net_style = 'color:#c62828;' if is_credit_note else ''
        summary_vat_style = 'color:#c62828;' if is_credit_note else 'color:#1565c0;'
        summary_payable_style = 'color:#c62828;' if is_credit_note else ''

        # Build variation rows for summary page
        summary_variation_rows = ""
        if var_total > 0:
            for vo in inv.variation_orders.all():
                summary_variation_rows += f"""
                    <tr style="background:#0000;">
                        <td class="summary-label">Variation: {vo.description[:55]}{'...' if len(vo.description) > 55 else ''}</td>
                        <td class="summary-num" style="color:#6a1b9a; font-weight:bold;">{vo.amount:,.2f}</td>
                    </tr>
                """
            ###
            #summary_variation_rows += f"""
            #        <tr style="background:#0000;">
            #            <td class="summary-label"><b>Total Variations</b></td>
            #           <td class="summary-num" style="color:#6a1b9a; font-weight:bold;">{var_total:,.2f}</td>
            #       </tr>
            #"""
            ###
        # Build summary rows dynamically based on invoice type
        if is_retention_recovery:
            summary_deduction_rows = f"""
                    <tr>
                        <td class="summary-label"><b>{recovery_type}</b></td>
                        <td class="summary-num" style="color:#2e7d32; font-weight:bold;">{recovery_amount:,.2f}</td>
                    </tr>
                """
        else:
            summary_deduction_rows = f"""
                    <tr>
                        <td class="summary-label">Advance Recovery</td>
                        <td class="summary-num">({inv.current_advance_recovery:,.2f})</td>
                    </tr>
                    <tr>
                        <td class="summary-label">Retention A ({inv.project.retention_a_percent}%)</td>
                        <td class="summary-num">({inv.current_retention_a:,.2f})</td>
                    </tr>
                    <tr>
                        <td class="summary-label">Retention B ({inv.project.retention_b_percent}%)</td>
                        <td class="summary-num">({inv.current_retention_b:,.2f})</td>
                    </tr>
                """

        summary_page = (
                '<div class="page summary-page">'
                + logo_bar_html +
                f'<div class="invoice-title">{header_title}</div>'
                '<div class="summary-meta">'
                f'<div class="summary-meta-row"><b>Invoice Number:</b> {inv}</div>'
                f'<div class="summary-meta-row"><b>Date:</b> {inv.date.strftime("%d-%b-%Y")}</div>'
                f'<div class="summary-meta-row"><b>Valuation Date:</b> {(getattr(inv, "valuation_date", None).strftime("%d-%b-%Y") if getattr(inv, "valuation_date", None) else inv.date.strftime("%d-%b-%Y"))}</div>'
                f'<div class="summary-meta-row"><b>Project Number:</b> {inv.project.project_id_code}</div>'
                '</div>'
                '<div class="parties-row">'
                '<div class="party-block">'
                f'<div class="party-name">{company.company_name if company else "PROCON GENERAL CONTRACTING LLC"}</div>'
                f'<div class="party-detail"><b>TRN:</b> {company.trn_number if company and company.trn_number else "N/A"}</div>'
                '</div>'
                '<div class="party-block" style="text-align:right;">'
                f'<div class="party-name">{inv.project.client.name}</div>'
                f'<div class="party-detail"><b>TRN:</b> {inv.project.client.vat_number or "N/A"}</div>'
                '</div>'
                '</div>'
                f'<div class="project-name"><b>Project:</b> {inv.project.project_name}</div>'
                '<div class="section-title">PROGRESS PAYMENT SUMMARY</div>'
                '<table class="summary-table">'
                '<tbody>'
        )

        if not is_retention_recovery:
            summary_page += f"""
                    <tr>
                        <td class="summary-label">Total Work Done</td>
                        <td class="summary-num">{inv.cumulative_work_done:,.2f}</td>
                    </tr>
                    <tr>
                        <td class="summary-label">Previously Invoiced</td>
                        <td class="summary-num">({inv.previous_net_total_invoiced:,.2f})</td>
                    </tr>
                    <tr>
                        <td class="summary-label">Current Gross Work</td>
                        <td class="summary-num">{inv.current_gross_total:,.2f}</td>
                    </tr>
                """
        else:
            summary_page += f"""
                    <tr>
                        <td class="summary-label">Total Work Done</td>
                        <td class="summary-num">{inv.cumulative_work_done:,.2f}</td>
                    </tr>
                    <tr>
                        <td class="summary-label">Previously Invoiced</td>
                        <td class="summary-num">({inv.previous_net_total_invoiced:,.2f})</td>
                    </tr>
                    <tr style="background:#e8f5e9;">
                        <td class="summary-label"><b>{recovery_type}</b></td>
                        <td class="summary-num" style="color:#2e7d32; font-weight:bold;">{recovery_amount:,.2f}</td>
                    </tr>
                """

        # FIX: Add variations BEFORE deductions in summary, and include in net
        summary_page += summary_variation_rows
        summary_page += summary_deduction_rows

        summary_page += f"""
                    <tr class="summary-subtotal">
                        <td class="summary-label"><b>Net Payable Amount without VAT</b></td>
                        <td class="summary-num" style="{summary_net_style}"><b>{fmt_val(net_before_vat)}</b></td>
                    </tr>
                    <tr style="background:#e3f2fd;">
                        <td class="summary-label" style="{summary_vat_style}"><b>VAT Output ({vat_rate}%)</b></td>
                        <td class="summary-num" style="{summary_vat_style} font-weight:bold;">{vat_sign}{fmt_val(vat_output)}</td>
                    </tr>
                    <tr>
                        <td class="summary-label">Materials Supplied by Client (VAT excl.)</td>
                        <td class="summary-num deduction">({materials_exclusive:,.2f})</td>
                    </tr>
                    <tr style="background:#fff3e0;">
                        <td class="summary-label" style="color:#e65100;"><b>VAT Input on Materials ({vat_rate}%)</b></td>
                        <td class="summary-num" style="color:#e65100; font-weight:bold;">({vat_input:,.2f})</td>
                    </tr>
                    <tr class="summary-subtotal" style="background:#e8eaf6;">
                        <td class="summary-label"><b>Net VAT Payable</b></td>
                        <td class="summary-num"><b>{fmt_val(net_vat)}</b></td>
                    </tr>
                    <tr class="summary-grand">
                        <td class="summary-label"><b>Net Payable Amount</b></td>
                        <td class="summary-num" style="{summary_payable_style}"><b>{fmt_val(payable_amount)}</b></td>
                    </tr>
                    </tbody>
                    </table>
                    <div class="summary-words">
                        <b>AED {amount_words}</b>
                    </div>
                    <div class="bank-details">
                        <div class="bank-title">Payment to be made by Bank transfer to:</div>
                        <div class="bank-content">
                        {bank_html}
                        </div>
                    </div>
                    <div class="summary-signatures">
                        <div class="sig-block">
                            <div class="sig-line"></div>
                            <div class="sig-name">Technical Manager</div>
                        </div>
                        <div class="sig-block">
                            <div class="sig-line"></div>
                            <div class="sig-name">Finance Director</div>
                        </div>
                    </div>
                    </div>
            """

        # ─────────────────────────────────────────────────────────────
        # DETAIL PAGE
        # ─────────────────────────────────────────────────────────────
        detail_page = (
                '<div class="page detail-page">'
                + logo_bar_html +
                f'<div class="invoice-title">{header_title}</div>'
                '<div class="invoice-meta">'
                f'<div class="invoice-meta-row">Invoice Number: {inv}</div>'
                '<br>'
                f'<div class="invoice-meta-row">Date : {inv.date}</div>'
                '<br>'
                f'<div class="invoice-meta-row">PO Number : {getattr(inv.project, "po_number", "N/A")}</div>'
                '</div>'
                '<div class="parties-row">'
                '<div class="party-block">'
                f'<div class="party-name">{company.company_name if company else ""}</div>'
                f'<div class="party-detail"><strong>TRN:</strong> {company.trn_number if company and company.trn_number else "N/A"}</div>'
                '<br>'
                f'<div class="party-name">{inv.project.client.name}</div>'
                f'<div class="party-detail"><strong>TRN:</strong> {inv.project.client.vat_number or "N/A"}</div>'
                '</div>'
                '</div>'
                f'<div class="project-name"><strong>Project:</strong> {inv.project.project_name}</div>'
                '<div class="section-title">PROGRESS PAYMENT</div>'
                '<table class="boq-table">'
                '<thead>'
                '<tr>'
                '<th rowspan="2" class="col-item">Item</th>'
                '<th rowspan="2" class="col-desc">Description</th>'
                '<th rowspan="2" class="col-unit">Unit</th>'
                '<th rowspan="2" class="col-num">BOQ<br>Qty</th>'
                '<th rowspan="2" class="col-num">Rate</th>'
                '<th colspan="2">Previous</th>'
                '<th colspan="2">Current</th>'
                '<th colspan="2">Cumulative</th>'
                '</tr>'
                '<tr>'
                '<th class="col-num">%</th>'
                '<th class="col-num">Amt</th>'
                '<th class="col-num">%</th>'
                '<th class="col-num">Amt</th>'
                '<th class="col-num">%</th>'
                '<th class="col-num">Amt</th>'
                '</tr>'
                '</thead>'
                f'<tbody>{rows}</tbody>'
                f'<tfoot>{footer_rows}</tfoot>'
                '</table>'
                '<div class="summary-wrapper">'
                '<div class="summary-box bank">'
                '<div class="summary-box-title">Bank Account Details:</div>'
                '<div class="bank-content">'
                f'{company.bank if company and company.bank else ""}'
                '</div>'
                '</div>'
                '<div class="summary-box totals">'
                f'<div class="summary-row"><span>Total Invoiced (Cumulative):</span><span>{inv.net_total_invoiced_cumulative:,.2f}</span></div>'
                f'<div class="summary-row"><span>Previously Invoiced:</span><span>({inv.previous_net_total_invoiced:,.2f})</span></div>'
        )

        if is_retention_recovery:
            detail_page += f"""
                <div class="summary-row" style="background:#e8f5e9; padding:2px 0;"><span><b>{recovery_type}:</b></span><span style="color:#2e7d32; font-weight:bold;">{recovery_amount:,.2f}</span></div>
            """
        else:
            detail_page += f"""
                <div class="summary-row"><span>Current Gross Work:</span><span>{inv.current_gross_total:,.2f}</span></div>
                <div class="summary-row"><span>Advance Recovery:</span><span>({inv.current_advance_recovery:,.2f})</span></div>
                <div class="summary-row"><span>Retention A:</span><span>({inv.current_retention_a:,.2f})</span></div>
                <div class="summary-row"><span>Retention B:</span><span>({inv.current_retention_b:,.2f})</span></div>
            """

        # FIX: Add variations to detail page summary box
        if var_total > 0:
            detail_page += f"""
                <div class="summary-row" style="background:#0000; padding:2px 0;"><span><b>Variations:</b></span><span style="color:#6a1b9a; font-weight:bold;">{var_total:,.2f}</span></div>
            """

        detail_page += f"""
                <div class="summary-row border-top" style="font-weight:bold;"><span>Net Before VAT:</span>
                <span style="{summary_net_style}">{fmt_val(net_before_vat)}</span></div>
                <div class="summary-row" style="{summary_vat_style}"><span>VAT Output ({vat_rate}%):</span><span>{vat_sign}{fmt_val(vat_output)}</span></div>
                <div class="summary-row"><span>Materials Supplied by Client:</span><span>({materials_exclusive:,.2f})</span></div>
                <div class="summary-row" style="color:#e65100;"><span>VAT Input on Materials ({vat_rate}%):</span><span>({vat_input:,.2f})</span></div>
                <div class="summary-row" style="font-weight:bold; background:#e8eaf6; padding:4px 0;"><span>Net VAT Payable:</span>
                <span>{fmt_val(net_vat)}</span></div>
                <div class="summary-row border-top red-text"><span>Payable Amount:</span>
                <span style="{summary_payable_style}">{fmt_val(payable_amount)}</span></div>
                </div>
                </div>
                </div>
        """

        html = f"""<!DOCTYPE html>
    <html>
    <head>
    <meta charset="UTF-8">
    <style>
        @page {{ size: A4 portrait; margin: 12mm; }}
        * {{ box-sizing: border-box; -webkit-print-color-adjust: exact; print-color-adjust: exact; margin: 0; padding: 0; }}
        body {{ font-family: "Segoe UI", Arial, Helvetica, sans-serif; font-size: 9px; line-height: 1.3; color: #222; }}

        .page {{ page-break-after: always; break-after: page; position: relative; padding: 8mm; }}
        .page:last-child {{ page-break-after: auto; }}

        .logo-bar {{ text-align: right; margin-bottom: 12px; }}
        .logo-bar img {{ max-height: 120px; max-width: 240px; object-fit: contain; }}

        .cover-page {{ display: flex; flex-direction: column; min-height: calc(100vh - 20mm); }}
        .cover-content {{ margin-top: 20px; }}
        .cover-ref {{ margin-bottom: 25px; }}
        .cover-ref-line {{ font-size: 10px; margin-bottom: 5px; }}
        .cover-label {{ font-weight: bold; display: inline-block; min-width: 50px; }}
        .cover-to {{ margin-bottom: 15px; }}
        .cover-to-line {{ font-size: 10px; margin-bottom: 3px; }}
        .cover-attn {{ margin-bottom: 12px; font-size: 10px; }}
        .cover-project {{ margin-bottom: 12px; font-size: 10px; }}
        .cover-subject {{ margin-bottom: 25px; font-size: 10px; }}
        .cover-body {{ margin-bottom: 30px; font-size: 10px; line-height: 1.7; }}
        .cover-body p {{ margin-bottom: 10px; }}
        .cover-closing {{ margin-top: auto; font-size: 10px; }}
        .cover-closing p {{ margin-bottom: 6px; }}
        .cover-signature-space {{ height: 50px; }}
        .cover-signature-name {{ font-weight: bold; margin-top: 8px; }}
        .cover-signature-title {{ color: #666; }}

        .summary-meta {{ margin-bottom: 12px; }}
        .summary-meta-row {{ font-size: 10px; font-weight: bold; margin-bottom: 3px; }}
        .parties-row {{ display: flex; justify-content: space-between; margin: 12px 0; gap: 15px; }}
        .party-block {{ flex: 1; }}
        .party-name {{ font-size: 12px; font-weight: bold; margin-bottom: 2px; }}
        .party-detail {{ font-size: 9px; margin-bottom: 2px; }}
        .project-name {{ font-size: 9px; margin: 8px 0 12px 0; }}
        .section-title {{ text-align: center; font-size: 13px; font-weight: bold; margin: 12px 0 10px 0; color: #111; border-bottom: 2px solid #000080; padding-bottom: 5px; }}

        .summary-table {{ width: 100%; border-collapse: collapse; margin-top: 8px; font-size: 10px; table-layout: fixed; }}
        .summary-table td {{ padding: 6px 10px; border-bottom: 1px solid #ddd; vertical-align: middle; }}
        .summary-table tr:last-child td {{ border-bottom: 2px solid #000080; }}
        .summary-table .summary-label {{ text-align: left; width: 65%; }}
        .summary-table .summary-num {{ text-align: right; font-weight: bold; width: 35%; white-space: nowrap; overflow-wrap: break-word; word-break: break-all; }}
        .summary-table .deduction {{ color: #c62828; }}
        .summary-subtotal td {{ background: #f5f5f5; }}
        .summary-grand td {{ background: #000080; color: white; font-size: 11px; }}
        .summary-grand .summary-label {{ color: white; }}
        .summary-grand .summary-num {{ color: white; }}

        .summary-words {{ text-align: center; margin: 12px 0; font-size: 10px; color: #333; }}

        .bank-details {{ margin-top: 20px; border: 1px solid #ccc; padding: 10px; border-radius: 3px; }}
        .bank-title {{ font-weight: bold; font-size: 9px; margin-bottom: 6px; color: #000080; }}
        .bank-content {{ font-family: "Courier New", monospace; font-size: 9px; line-height: 1.5; }}
        .bank-row {{ margin-bottom: 2px; }}

        .summary-signatures {{ display: flex; justify-content: space-between; margin-top: 30px; gap: 50px; }}
        .sig-block {{ flex: 1; text-align: center; }}
        .sig-line {{ border-top: 1px solid #333; margin-top: 40px; padding-top: 6px; }}
        .sig-name {{ font-size: 9px; font-weight: bold; }}

        .invoice-title {{ font-size: 14px; font-weight: bold; text-align: center; margin: 6px 0 10px 0; color: #000080; letter-spacing: 1px; }}
        .invoice-meta {{ margin-bottom: 8px; }}
        .invoice-meta-row {{ font-size: 10px; font-weight: bold; margin-bottom: 3px; }}

        .boq-table {{ width: 100%; border-collapse: collapse; margin-top: 5px; font-size: 7.5px; table-layout: fixed; }}
        .boq-table thead {{ display: table-header-group; }}
        .boq-table th {{ background: #e8e8e8; border: 1px solid #666; padding: 3px 2px; font-weight: bold; text-align: center; font-size: 7px; word-wrap: break-word; line-height: 1.2; }}
        .boq-table td {{ border: 1px solid #666; padding: 2px 3px; vertical-align: top; }}
        .boq-table .col-item {{ width: 5%; text-align: center; }}
        .boq-table .col-desc {{ width: 28%; text-align: left; }}
        .boq-table .col-unit {{ width: 5%; text-align: center; }}
        .boq-table .col-num {{ width: 7%; text-align: right; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
        .boq-table .col-label {{ text-align: right; font-weight: bold; padding-right: 6px; }}
        .boq-table td.col-desc {{ font-size: 7px; line-height: 1.2; word-wrap: break-word; overflow-wrap: break-word; hyphens: auto; }}
        .boq-table .total-row td {{ font-weight: bold; background: #f0f0f0; border-top: 2px solid #333; }}
        .boq-table .grand-total-row td {{ font-weight: bold; background: #f5f5f5; border-top: 2px solid #333; border-bottom: 2px solid #333; }}
        .boq-table tr {{ page-break-inside: avoid; }}

        .summary-wrapper {{ display: flex; justify-content: space-between; margin-top: 12px; gap: 12px; page-break-inside: avoid; }}
        .summary-box {{ border: 1px solid #666; padding: 6px 10px; }}
        .summary-box.bank {{ flex: 1; max-width: 48%; }}
        .summary-box.totals {{ flex: 1; max-width: 48%; margin-left: auto; }}
        .summary-box-title {{ font-weight: bold; margin-bottom: 4px; text-decoration: underline; font-size: 8px; }}
        .summary-box .summary-row {{ display: flex; justify-content: space-between; margin-bottom: 3px; font-size: 8px; }}
        .summary-box .summary-row.border-top {{ border-top: 1px solid #333; margin-top: 4px; padding-top: 4px; }}
        .red-text {{ color: #000080; font-weight: bold; font-size: 1.3em; }}
        .page-break-avoid {{ page-break-inside: avoid; }}
    </style>
    </head>
    <body>
        {cover_page}
        {summary_page}
        {detail_page}
        <script>window.onload = function() {{ window.print(); }}</script>
    </body>
    </html>"""
        return HttpResponse(html)

    def _number_to_words(self, number):
        """Convert a number to words. Handles negative values with 'Minus' in parentheses."""
        try:
            from num2words import num2words
            is_negative = number < 0
            abs_number = abs(number)
            integer_part = int(abs_number)
            decimal_part = int((abs_number - integer_part) * 100)
            words = num2words(integer_part, lang='en').replace(',', '').title()
            if decimal_part > 0:
                words += f" and {decimal_part:02d}/100"
            if is_negative:
                words = f"(Minus {words})"
            return words
        except Exception:
            integer_part = int(abs(number))
            decimal_part = int((abs(number) - integer_part) * 100)
            result = f"{integer_part} and {decimal_part:02d}/100"
            if number < 0:
                result = f"(Minus {result})"
            return result

    def logo_preview(self, obj):
        if obj.logo:
            return format_html('<img src="{}" style="max-height:120px; max-width:240px;" />', obj.logo.url)
        return "—"
    logo_preview.short_description = "Logo Preview"


# =============================================================================
# EXPENSE ADMIN
# =============================================================================

class SubExpenseInline(admin.TabularInline):
    model = SubExpense
    extra = 1


@admin.register(ExpenseCategory)
class ExpenseCategoryAdmin(CompanyScopedAdminMixin,admin.ModelAdmin):
    company_field_path = 'company'
    list_display = ["name", "sub_expense_count", "default_supplier", "company"]
    list_filter = ["company", "default_supplier"]
    autocomplete_fields = ["default_supplier"]
    inlines = [SubExpenseInline]
    search_fields = ["name"]

    def sub_expense_count(self, obj):
        return obj.sub_expenses.count()
    sub_expense_count.short_description = "Sub-Expenses"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('subexpense/get-by-category/', self.admin_site.admin_view(self.get_subexpenses),
                 name='subexpense_get_by_category'),
        ]
        return custom + urls

    def get_subexpenses(self, request):
        from django.http import JsonResponse
        category_id = request.GET.get('category_id')
        if not category_id:
            return JsonResponse([], safe=False)
        items = SubExpense.objects.filter(parent_id=category_id).values('id', 'name')
        data = [{'id': item['id'], 'text': item['name']} for item in items]
        return JsonResponse(data, safe=False)


@admin.register(Expense)
class ExpenseAdmin(CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'project__company'
    list_display = [
        "date", "project", "category", "sub_category",
        "fmt_supplier", "fmt_amount", "boq_item",
        "is_allocated", "is_auto_generated"
    ]
    list_filter = [
        "project__company", "project", "category", "supplier",
        "date", "is_allocated", "is_auto_generated"
    ]
    search_fields = ["description", "reference_number", "supplier__name"]
    autocomplete_fields = ["project", "boq_item", "supplier", "supplier_invoice"]
    readonly_fields = ["is_auto_generated", "supplier_invoice_link"]

    fieldsets = (
        ("Expense Details", {
            "fields": (
                ("project", "date"),
                ("category", "sub_category"),
                ("boq_item",),
                ("amount",),
                ("description", "reference_number"),
                ("is_allocated",),
            )
        }),
        ("Supplier Linkage", {
            "fields": (
                ("supplier", "supplier_invoice"),
                ("supplier_invoice_link",),
                ("is_auto_generated",),
            ),
            "description": "Link to supplier invoice (auto-populated when created from AP)",
            "classes": ("collapse",)
        }),
    )

    class Media:
        js = ('admin/js/vendor/jquery/jquery.min.js', 'admin/js/jquery.init.js', 'admin/js/expense_boq_filter.js')

    def fmt_supplier(self, obj):
        if obj.supplier:
            return mark_safe(
                f'<div style="font-size:10px;"><b>{obj.supplier.name}</b></div>'
            )
        return mark_safe('<span style="color:#999;">—</span>')
    fmt_supplier.short_description = "Supplier"

    def supplier_invoice_link(self, obj):
        """Display a clickable link to the linked supplier invoice."""
        if obj.supplier_invoice:
            url = reverse('admin:billing_supplierinvoice_change', args=[obj.supplier_invoice.pk])
            return mark_safe(
                f'<a href="{url}" target="_blank" style="color:#1a237e; font-weight:bold;">'
                f'📄 {obj.supplier_invoice.supplier_inv_number} (AED {obj.supplier_invoice.total_amount:,.2f})'
                f'</a>'
            )
        return mark_safe('<span style="color:#999;">No linked invoice</span>')
    supplier_invoice_link.short_description = "Linked Supplier Invoice"

    def fmt_amount(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.amount:,.2f}</div>')
    fmt_amount.short_description = "Amount"

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if obj and obj.category:
            form.base_fields['sub_category'].queryset = SubExpense.objects.filter(parent=obj.category)
        else:
            form.base_fields['sub_category'].queryset = SubExpense.objects.none()
        if obj and obj.project:
            form.base_fields['boq_item'].queryset = BOQItem.objects.filter(project=obj.project)
        else:
            form.base_fields['boq_item'].queryset = BOQItem.objects.none()
        return form

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        company = self.get_active_company(request)
        if db_field.name == "boq_item":
            if request.resolver_match and request.resolver_match.kwargs.get('object_id'):
                obj = self.get_object(request, request.resolver_match.kwargs['object_id'])
                if obj and obj.project:
                    kwargs["queryset"] = BOQItem.objects.filter(project=obj.project)
                else:
                    kwargs["queryset"] = BOQItem.objects.none()
            else:
                kwargs["queryset"] = BOQItem.objects.none()
        elif db_field.name == "sub_category":
            if request.resolver_match and request.resolver_match.kwargs.get('object_id'):
                obj = self.get_object(request, request.resolver_match.kwargs['object_id'])
                if obj and obj.category:
                    kwargs["queryset"] = SubExpense.objects.filter(parent=obj.category)
                else:
                    kwargs["queryset"] = SubExpense.objects.none()
            else:
                kwargs["queryset"] = SubExpense.objects.none()
        elif db_field.name == "supplier":
            if company:
                kwargs["queryset"] = Supplier.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
        elif db_field.name == "supplier_invoice":
            if company:
                kwargs["queryset"] = SupplierInvoice.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
        else:
            return super().formfield_for_foreignkey(db_field, request, **kwargs)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


# =============================================================================
# EMPLOYEE ADMIN
# =============================================================================

class EmployeeTransferInline(admin.TabularInline):
    model = EmployeeTransfer
    extra = 0
    fields = ["to_project", "from_date", "to_date", "days_count", "notes"]
    readonly_fields = ["days_count"]

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'to_project':
            company = CompanyProfile.get_active(request)
            if company:
                kwargs['queryset'] = Project.objects.filter(company=company)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class EmployeeAdminForm(forms.ModelForm):
    class Meta:
        model = Employee
        fields = '__all__'
        labels = {
            'employee_id': 'Employee ID',
            'name': 'Name',
            'employee_type': 'Employee Type',
            'payment_type': 'Payment Type',
            'project': 'Project',
            'is_head_office': 'Is Head Office',
            'is_active': 'Is Active',
            'date_joined': 'Date Joined',
            'basic_salary': 'Basic Salary',
            'housing_allowance': 'Housing Allowance',
            'transport_allowance': 'Transport Allowance',
            'other_allowances': 'Other Allowances',
            'annual_benefits': 'Annual Benefits',
            'annual_eid_cost': 'Annual Housing Cost',
            'annual_visa_cost': 'Annual Resident Permit Cost',
            'annual_ticket_cost': 'Annual Tickets',
            'total_salary': 'Total Salary',
            'monthly_admin_cost': 'Monthly Admin Cost',
            'daily_cost': 'Daily Cost',
            'hourly_rate_ot': 'Hourly Rate OT',
            'bank_name': 'Bank Name',
            'routing_number': 'Routing Number',
            'iban': 'IBAN',
        }


@admin.register(Employee)
class EmployeeAdmin(CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'company'
    form = EmployeeAdminForm
    list_display = [
        "employee_id", "name", "employee_type", "payment_type",
        "cost_center", "company", "fmt_total_salary", "fmt_total_package", "fmt_eos",
        "fmt_daily_cost", "fmt_hourly_rate", "fmt_bank_info", "is_active", "transfer_status"
    ]
    list_filter = ["company", "employee_type", "payment_type", "is_head_office", "is_active", "project"]
    search_fields = ["name", "employee_id"]
    inlines = [EmployeeTransferInline]
    readonly_fields = ["total_salary", "monthly_admin_cost", "daily_cost", "hourly_rate_ot", "display_eos"]
    fieldsets = (
        ("Employee Information", {
            "fields": (
                ("employee_id", "name"),
                ("employee_type", "payment_type"),
                ("company", "project", "is_head_office"),
                ("is_active", "date_joined"),
            )
        }),
        ("Salary Components", {
            "fields": (
                ("basic_salary", "housing_allowance"),
                ("transport_allowance", "other_allowances"),
            )
        }),
        ("Annual Administrative Costs", {
            "fields": (
                ("annual_benefits", "annual_eid_cost"),
                ("annual_visa_cost", "annual_ticket_cost"),
            ),
            "description": "These are summed and divided by 12 to produce the monthly admin cost."
        }),
        ("Computed Employment Costs (Read-Only)", {
            "fields": (
                ("total_salary", "monthly_admin_cost"),
                ("daily_cost", "hourly_rate_ot"),
                ("display_eos",),
            ),
            "description": (
                "Daily Cost = (Total Salary + Admin Cost) / 30. "
                "Hourly OT Rate = Total Salary / 30 / 8 (Site workers only). "
                "EOSB = End of Service Benefits (21 days/years 1-3, 30 days/year 4+)."
            ),
        }),
        ("Bank Details", {
            "fields": (
                ("bank_name", "routing_number"),
                ("iban",),
            ),
            "description": "Bank information for Bank Transfer or WPS Agency payments."
        }),
    )

    def save_model(self, request, obj, form, change):
        if obj.project and not obj.company_id:
            obj.company = obj.project.company
        elif not obj.company_id:
            obj.company = CompanyProfile.get_active(request)
        super().save_model(request, obj, form, change)

    def display_eos(self, obj):
        import logging
        logger = logging.getLogger(__name__)
        try:
            val = obj.eos_amount
            if val is None:
                return mark_safe('<span style="color:#999;">0.00</span>')
            val = money(val)
            if val > 0:
                return mark_safe(f'<div style="text-align:right;font-weight:bold;color:#d32f2f;">{val:,.2f}</div>')
            else:
                return mark_safe('<span style="color:#999;">0.00</span>')
        except Exception as e:
            logger.error(f"DEBUG EOS ERROR for {obj}: {e}", exc_info=True)
            return mark_safe(f'<span style="color:red; font-weight:bold;">ERROR: {str(e)[:40]}</span>')
    display_eos.short_description = "EOSB (End of Service Benefit)"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.prefetch_related('payroll_records')

    def cost_center(self, obj):
        if obj.is_head_office:
            return mark_safe('<b style="color:#000080;">HEAD OFFICE</b>')
        return obj.project.project_id_code if obj.project else mark_safe('<span style="color:#999;">—</span>')
    cost_center.short_description = "Cost Center"

    def fmt_bank_info(self, obj):
        if obj.bank_name:
            return mark_safe(
                f'<div style="font-size:10px;"><b>{obj.bank_name}</b><br><span style="color:#666;">IBAN: {obj.iban or "—"}</span></div>')
        return mark_safe('<span style="color:#999;">—</span>')
    fmt_bank_info.short_description = "Bank Info"

    def fmt_total_salary(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.total_salary:,.2f}</div>')
    fmt_total_salary.short_description = "Total Salary"

    def fmt_total_package(self, obj):
        total = obj.total_salary + obj.monthly_admin_cost
        return mark_safe(f'<div style="text-align:right;font-weight:bold;color:#000080;">{total:,.2f}</div>')
    fmt_total_package.short_description = "Total Package"

    def fmt_eos(self, obj):
        try:
            val = obj.eos_amount
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error accessing eos_amount for {obj}: {e}")
            val = Decimal("0.00")
        if val and val > 0:
            return mark_safe(
                f'<div style="text-align:right;font-weight:bold;color:#d32f2f;">{val:,.2f}</div>'
            )
        return mark_safe('<span style="color:#999;">—</span>')
    fmt_eos.short_description = "EOS"

    def fmt_daily_cost(self, obj):
        return mark_safe(f'<div style="text-align:right;color:#2e7d32;font-weight:bold;">{obj.daily_cost:,.2f}</div>')
    fmt_daily_cost.short_description = "Daily Cost"

    def fmt_hourly_rate(self, obj):
        if obj.hourly_rate_ot > 0:
            return mark_safe(f'<div style="text-align:right;">{obj.hourly_rate_ot:,.2f}</div>')
        return mark_safe('<span style="color:#999;">—</span>')
    fmt_hourly_rate.short_description = "Hourly Rate"

    def transfer_status(self, obj):
        active_transfers = obj.transfers.filter(
            from_date__lte=date.today(),
            to_date__gte=date.today()
        ).select_related('to_project')
        if active_transfers.exists():
            t = active_transfers.first()
            return mark_safe(f'<b style="color:#ed6c02;">to {t.to_project.project_id_code}</b>')
        return mark_safe('<span style="color:#999;">—</span>')
    transfer_status.short_description = "Transfer"

    def changelist_view(self, request, extra_context=None):
        from django.db.models import Q
        company = self.get_active_company(request)
        complete_projects = Project.objects.filter(is_boq_complete=True, company=company)
        for proj in complete_projects:
            workers = Employee.objects.filter(project=proj, is_active=True)
            if workers.exists():
                messages.warning(
                    request,
                    mark_safe(
                        f"<b>PROJECT COMPLETE:</b> {proj.project_id_code} has {workers.count()} active worker(s). "
                        f"<a href='{reverse('admin:billing_employee_changelist')}?project__id__exact={proj.id}' "
                        f"style='color:#d32f2f; text-decoration:underline; font-weight:bold;'>Transfer Workers</a>"
                    )
                )
        return super().changelist_view(request, extra_context)

# =============================================================================
# EMPLOYEE TRANSFER ADMIN (standalone for tree navigation)
# =============================================================================

@admin.register(EmployeeTransfer)
class EmployeeTransferAdmin(CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'employee__company'
    list_display = ["employee", "to_project", "from_date", "to_date", "days_count", "notes"]
    list_filter = ["employee__company", "to_project", "from_date"]
    search_fields = ["employee__name", "employee__employee_id"]
    autocomplete_fields = ["employee", "to_project"]

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        company = self.get_active_company(request)
        if company:
            if db_field.name == 'to_project':
                kwargs['queryset'] = Project.objects.filter(company=company)
            elif db_field.name == 'employee':
                kwargs['queryset'] = Employee.objects.filter(company=company)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

# =============================================================================
# PAYROLL ADMIN
# =============================================================================

class PayrollAllocationInline(admin.TabularInline):
    model = PayrollAllocation
    extra = 0
    readonly_fields = [
        "project", "boq_item", "salary_allocated", "admin_cost_allocated",
        "total_allocated", "project_work_done_pct", "boq_item_work_done_pct", "created_at"
    ]
    can_delete = False

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        company = CompanyProfile.get_active(request)
        if company:
            if db_field.name == 'project':
                kwargs['queryset'] = Project.objects.filter(company=company)
            elif db_field.name == 'boq_item':
                kwargs['queryset'] = BOQItem.objects.filter(project__company=company)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class PayrollCostCenterInline(admin.TabularInline):
    model = PayrollCostCenter
    extra = 0
    fields = ["project", "from_date", "to_date", "days_count", "overtime_hours", "bonus", "prorated_salary",
              "fmt_overtime_amount", "notes"]
    readonly_fields = ["days_count", "prorated_salary", "fmt_overtime_amount"]

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        if 'overtime_hours' in formset.form.base_fields:
            formset.form.base_fields['overtime_hours'].widget = forms.NumberInput(attrs={'step': '1', 'min': '0'})
        return formset

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        company = CompanyProfile.get_active(request)
        if company and db_field.name == 'project':
            kwargs['queryset'] = Project.objects.filter(company=company)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def fmt_overtime_amount(self, obj):
        try:
            ot_hours = obj.overtime_hours
        except Exception:
            ot_hours = Decimal("0")
        if ot_hours and ot_hours > 0 and obj.payroll_record and obj.payroll_record.employee:
            rate = obj.payroll_record.employee.hourly_rate_ot
            if rate > 0:
                amt = money(ot_hours * rate)
                return mark_safe(f'<div style="text-align:right;font-weight:bold;">{amt:,.2f}</div>')
        return mark_safe('<span style="color:#999;">—</span>')
    fmt_overtime_amount.short_description = "OT Amount"

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        company = CompanyProfile.get_active(request)
        if company:
            qs = qs.filter(project__company=company)
        return qs

class PayrollRecordAdminForm(forms.ModelForm):
    class Meta:
        model = PayrollRecord
        fields = '__all__'
        exclude = ['overtime_hours']


@admin.register(PayrollRecord)
class PayrollRecordAdmin(CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'employee__company'
    form = PayrollRecordAdminForm
    inlines = [PayrollCostCenterInline, PayrollAllocationInline]
    list_display = [
        "employee", "month", "fmt_total_salary", "fmt_overtime", "fmt_days_absent",
        "fmt_absence", "fmt_advance", "fmt_other_ded", "fmt_net_salary",
        "timesheet_button", "labor_cost_button", "allocation_status",
    ]
    list_filter = ["month", "is_allocated", "employee__employee_type", "employee__payment_type", "employee__company"]
    search_fields = ["employee__name", "employee__employee_id"]
    actions = ["allocate_selected", "recalculate_selected"]
    date_hierarchy = "month"

    def fmt_total_salary(self, obj):
        return mark_safe(f'<div style="text-align:right;">{obj.total_salary_snap:,.2f}</div>')
    fmt_total_salary.short_description = "Total Salary"

    def fmt_overtime(self, obj):
        if obj.employee.employee_type == 'Site':
            try:
                cc_ot_hours = obj.cost_centers.aggregate(total=Sum('overtime_hours'))['total'] or Decimal("0")
            except Exception:
                cc_ot_hours = Decimal("0")
            if cc_ot_hours > 0 and obj.employee.hourly_rate_ot > 0:
                ot_amount = money(cc_ot_hours * obj.employee.hourly_rate_ot)
                return mark_safe(f'<div style="text-align:right;">{cc_ot_hours}h / {ot_amount:,.2f}</div>')
            return mark_safe('<span style="color:#999;">—</span>')
        return mark_safe('<span style="color:#999;">—</span>')
    fmt_overtime.short_description = "OT (Hrs/Amt)"

    def fmt_days_absent(self, obj):
        days = obj.days_absent
        if days > 0:
            return mark_safe(f'<div style="text-align:right;color:#d32f2f;font-weight:bold;">{days}</div>')
        return mark_safe('<div style="text-align:right;color:#2e7d32;">0</div>')
    fmt_days_absent.short_description = "Abs Days"

    def fmt_absence(self, obj):
        return mark_safe(f'<div style="text-align:right;color:#d32f2f;">({obj.absence_deduction_snap:,.2f})</div>')
    fmt_absence.short_description = "Absence"

    def fmt_advance(self, obj):
        return mark_safe(f'<div style="text-align:right;color:#d32f2f;">({obj.salary_advance:,.2f})</div>')
    fmt_advance.short_description = "Advance"

    def fmt_other_ded(self, obj):
        return mark_safe(f'<div style="text-align:right;color:#d32f2f;">({obj.other_deduction:,.2f})</div>')
    fmt_other_ded.short_description = "Other Ded."

    def fmt_net_salary(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.net_salary_snap:,.2f}</div>')
    fmt_net_salary.short_description = "Net Salary"

    def allocation_status(self, obj):
        if obj.is_allocated:
            return mark_safe('<b style="color:#2e7d32;">ALLOCATED</b>')
        return mark_safe('<b style="color:#d32f2f;">PENDING</b>')
    allocation_status.short_description = "Status"

    def timesheet_button(self, obj):
        url = reverse('admin:payroll_timesheet', args=[obj.pk])
        return format_html(
            '<a class="button" href="{}" target="_blank" style="background:#0288d1; color:white; padding: 2px 8px; border-radius: 4px; font-size:10px;">Time Sheet</a>',
            url)
    timesheet_button.short_description = "Sheet"

    def labor_cost_button(self, obj):
        url = reverse('admin:payroll_labor_cost') + f'?month={obj.month.strftime("%Y-%m")}'
        return format_html(
            '<a class="button" href="{}" target="_blank" style="background:#6a1b9a; color:white; padding: 2px 8px; border-radius: 4px; font-size:10px;">Labor Cost</a>',
            url)
    labor_cost_button.short_description = "Cost Rpt"

    @admin.action(description="Allocate selected payroll to projects / BOQ items")
    def allocate_selected(self, request, queryset):
        from .payroll import allocate_payroll
        done = 0
        skipped = 0
        for rec in queryset:
            if rec.is_allocated:
                skipped += 1
                continue
            if allocate_payroll(rec):
                done += 1
            else:
                skipped += 1
        self.message_user(request, f"Allocated: {done} | Skipped/Failed: {skipped}")

    @admin.action(description="Recalculate snaps & net salary for selected")
    def recalculate_selected(self, request, queryset):
        for rec in queryset:
            rec.save()
        self.message_user(request, f"Recalculated {queryset.count()} record(s).")


    def changelist_view(self, request, extra_context=None):
        today = date.today()
        first_day = today.replace(day=1)
        last_month = (first_day - timedelta(days=1)).replace(day=1)
        company = self.get_active_company(request)
        unallocated = PayrollRecord.objects.filter(month=last_month, is_allocated=False, employee__company=company).count()
        if unallocated > 0:
            messages.warning(
                request,
                mark_safe(
                    f"<b>PAYROLL ALERT:</b> {unallocated} record(s) for <b>{last_month.strftime('%b %Y')}</b> "
                    f"are not yet allocated to projects. "
                    f"<a href='allocate/' style='color:#d32f2f; text-decoration:underline; font-weight:bold;'>Allocate Now</a>"
                )
            )
        messages.info(
            request,
            mark_safe(
                f'<a href="{reverse("admin:payroll_labor_cost")}?month={last_month.strftime("%Y-%m")}" '
                f'target="_blank" style="color:#6a1b9a; font-weight:bold;">📊 View Monthly Labor Cost Report</a>'
            )
        )
        return super().changelist_view(request, extra_context)

    def get_urls(self):
        urls = super().get_urls()
        custom = [
                path("allocate/", self.admin_site.admin_view(self.allocate_view), name="payroll_allocate"),
                path("reports/staff/", self.admin_site.admin_view(self.staff_report), name="payroll_staff_report"),
                path("reports/wps/", self.admin_site.admin_view(self.wps_report), name="payroll_wps_report"),
                path("reports/cash/", self.admin_site.admin_view(self.cash_report), name="payroll_cash_report"),
                path("reports/labor-cost/", self.admin_site.admin_view(self.labor_cost_report),
                     name="payroll_labor_cost"),
                path('<int:pk>/timesheet/', self.admin_site.admin_view(self.timesheet_view), name='payroll_timesheet'),
            ]
        return custom + urls

    def timesheet_view(self, request, pk):
            company = self.get_active_company(request)
            payroll = self.get_object_or_404_scoped(request, PayrollRecord, pk=pk)
            emp = payroll.employee
            logo_url = company.logo.url if company and company.logo else ''

            month_start = payroll.month
            if month_start.month == 12:
                next_month = date(month_start.year + 1, 1, 1)
            else:
                next_month = date(month_start.year, month_start.month + 1, 1)
            month_end = next_month - timedelta(days=1)
            days_in_month = month_end.day

            cost_centers = list(PayrollCostCenter.objects.filter(
                payroll_record=payroll
            ).select_related('project'))

            day_entries = []
            for day_num in range(1, days_in_month + 1):
                day_date = month_start.replace(day=day_num)
                cc_for_day = None
                for cc in cost_centers:
                    if cc.from_date <= day_date <= cc.to_date:
                        cc_for_day = cc
                        break

                if cc_for_day:
                    is_last_day = (day_date == cc_for_day.to_date)
                    day_entries.append({
                        'date': day_date,
                        'project': cc_for_day.project,
                        'status': 'Present',
                        'status_color': '#2e7d32',
                        'ot_hours': cc_for_day.overtime_hours if is_last_day else Decimal("0"),
                        'bonus': cc_for_day.bonus if is_last_day else Decimal("0"),
                        'is_last_day': is_last_day,
                        'is_weekend': day_date.weekday() >= 5,
                    })
                else:
                    if emp.project:
                        day_entries.append({
                            'date': day_date,
                            'project': emp.project,
                            'status': 'Present',
                            'status_color': '#2e7d32',
                            'ot_hours': Decimal("0"),
                            'bonus': Decimal("0"),
                            'is_last_day': False,
                            'is_weekend': day_date.weekday() >= 5,
                        })
                    else:
                        day_entries.append({
                            'date': day_date,
                            'project': None,
                            'status': 'Absent',
                            'status_color': '#d32f2f',
                            'ot_hours': Decimal("0"),
                            'bonus': Decimal("0"),
                            'is_last_day': False,
                            'is_weekend': day_date.weekday() >= 5,
                        })

            rows = ""
            total_ot = Decimal("0")
            total_bonus = Decimal("0")
            present_days = 0
            absent_days = 0
            for entry in day_entries:
                day_name = entry['date'].strftime('%a')
                date_str = entry['date'].strftime('%d-%b-%Y')
                proj_name = entry['project'].project_name if entry['project'] else '—'
                proj_code = entry['project'].project_id_code if entry['project'] else ''
                ot = money(entry['ot_hours'])
                bonus = money(entry['bonus'])
                total_ot += ot
                total_bonus += bonus
                if entry['status'] == 'Present':
                    present_days += 1
                else:
                    absent_days += 1

                last_day_style = "background:#e8f5e9; font-weight:bold;" if entry['is_last_day'] else ""
                weekend_style = "background:#f5f5f5;" if entry['is_weekend'] and not entry['is_last_day'] else ""
                row_style = last_day_style or weekend_style

                rows += f"""<tr style="{row_style}">
                    <td style="text-align:center; font-weight:bold;">{day_name}</td>
                    <td>{date_str}</td>
                    <td><b>{proj_code}</b> — {proj_name}</td>
                    <td style="text-align:center; color:{entry['status_color']}; font-weight:bold;">{entry['status']}</td>
                    <td style="text-align:right; font-weight:bold; color:#ed6c02;">{ot:,.2f}</td>
                    <td style="text-align:right; font-weight:bold; color:#2e7d32;">{bonus:,.2f}</td>
                </tr>"""

            exact_ot = payroll.overtime_hours
            exact_bonus = Decimal("0")
            for cc in cost_centers:
                exact_ot += cc.overtime_hours
                exact_bonus += cc.bonus

            hourly_rate = emp.hourly_rate_ot if emp.employee_type == 'Site' else Decimal("0")
            total_ot_amount = money(exact_ot * hourly_rate) if hourly_rate > 0 else Decimal("0")
            total_extra = money(total_ot_amount + exact_bonus)

            basic = payroll.basic_salary_snap
            housing = payroll.housing_allowance_snap
            transport = payroll.transport_allowance_snap
            other = payroll.other_allowances_snap
            total_salary = payroll.total_salary_snap
            gross = money(total_salary + total_ot_amount + exact_bonus)
            absence_ded = payroll.absence_deduction_snap
            advance = payroll.salary_advance
            other_ded = payroll.other_deduction
            total_deductions = money(absence_ded + advance + other_ded)
            net = payroll.net_salary_snap

            html = f"""<!DOCTYPE html>
    <html><head><meta charset="UTF-8">
    <style>
        @page {{ size: A4 portrait; margin: 10mm; }}
        * {{ box-sizing: border-box; margin:0; padding:0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
        body {{ font-family: "Segoe UI", Arial, sans-serif; font-size: 10px; color: #222; padding: 10px; }}
        .logo-bar {{ text-align: right; margin-bottom: 6px; }}
        .logo-bar img {{ max-height: 120px; max-width: 240px; object-fit: contain; }}
        .report-title {{ font-size: 18px; font-weight: bold; text-align: center; color: #000080; margin-bottom: 4px; }}
        .report-subtitle {{ font-size: 12px; text-align: center; color: #666; margin-bottom: 15px; }}
        .meta-box {{ background: #f5f5f5; padding: 10px 15px; border-radius: 6px; margin-bottom: 15px; line-height: 1.6; font-size: 10px; }}
        .meta-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-bottom: 15px; }}
        .meta-card {{ border: 1px solid #ccc; border-radius: 6px; padding: 8px; background: #fafafa; }}
        .meta-label {{ font-size: 8px; color: #666; text-transform: uppercase; margin-bottom: 3px; }}
        .meta-value {{ font-size: 12px; font-weight: bold; color: #000080; }}
        .report-table {{ width: 100%; border-collapse: collapse; font-size: 9px; margin-top: 6px; }}
        .report-table th {{ background: #e8e8e8; border: 1px solid #999; padding: 6px; text-align: center; font-weight: bold; font-size: 8px; }}
        .report-table td {{ border: 1px solid #ccc; padding: 5px; }}
        .report-table tr:nth-child(even) {{ background: #fafafa; }}
        .total-row td {{ background: #e3f2fd; font-weight: bold; border-top: 2px solid #333; }}
        .summary-box {{ margin-top: 15px; padding: 12px; background: #000080; color: white; text-align: center; font-size: 14px; border-radius: 6px; }}
        .signature-grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px; margin-top: 30px; }}
        .signature-block {{ text-align: center; }}
        .signature-line {{ border-top: 1px solid #333; margin-top: 40px; padding-top: 5px; font-size: 10px; }}
        .payroll-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 15px; margin: 15px 0; }}
        .payroll-panel {{ border: 1px solid #ccc; border-radius: 8px; padding: 12px; background: white; }}
        .panel-title {{ font-size: 11px; font-weight: bold; color: #000080; margin-bottom: 10px; border-bottom: 2px solid #000080; padding-bottom: 6px; text-align: center; }}
        .pay-row {{ display: flex; justify-content: space-between; font-size: 10px; margin-bottom: 5px; padding: 3px 0; border-bottom: 1px dotted #eee; }}
        .pay-row.total {{ font-weight: bold; font-size: 11px; border-top: 2px solid #333; border-bottom: none; margin-top: 5px; padding-top: 8px; color: #000080; }}
        .pay-row.deduction {{ color: #d32f2f; }}
        .pay-row.net {{ font-size: 13px; font-weight: bold; color: #000080; border-top: 2px solid #000080; margin-top: 8px; padding-top: 10px; }}
        .highlight-green {{ color: #2e7d32; font-weight: bold; }}
        .highlight-red {{ color: #d32f2f; font-weight: bold; }}
        .highlight-orange {{ color: #ed6c02; font-weight: bold; }}
        @media print {{ .no-print {{ display: none; }} }}
    </style></head><body>
        {self._logo_bar(logo_url)}
        <div class="report-title">WORKER TIME SHEET & PAYROLL SUMMARY</div>
        <div class="report-subtitle">{emp.name} — {month_start.strftime('%B %Y')}</div>
        <div class="meta-grid">
            <div class="meta-card"><div class="meta-label">Employee ID</div><div class="meta-value">{emp.employee_id}</div></div>
            <div class="meta-card"><div class="meta-label">Employee Type</div><div class="meta-value">{emp.get_employee_type_display()}</div></div>
            <div class="meta-card"><div class="meta-label">Payment Type</div><div class="meta-value">{emp.get_payment_type_display()}</div></div>
            <div class="meta-card"><div class="meta-label">Basic Salary</div><div class="meta-value">{emp.basic_salary:,.2f}</div></div>
            <div class="meta-card"><div class="meta-label">Hourly OT Rate</div><div class="meta-value">{hourly_rate:,.2f}</div></div>
            <div class="meta-card"><div class="meta-label">Daily Rate</div><div class="meta-value">{emp.daily_rate:,.2f}</div></div>
        </div>
        <div class="meta-box">
            <b>Bank Name:</b> {emp.bank_name or 'N/A'} &nbsp;|&nbsp;
            <b>IBAN:</b> {emp.iban or 'N/A'} &nbsp;|&nbsp;
            <b>Routing:</b> {emp.routing_number or 'N/A'}
        </div>
        <table class="report-table">
            <thead>
                <tr>
                    <th style="width:8%;">Day</th>
                    <th style="width:15%;">Date</th>
                    <th style="width:35%;">Project</th>
                    <th style="width:12%;">Status</th>
                    <th style="width:15%;">OT Hours</th>
                    <th style="width:15%;">Bonus</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
            <tfoot>
                <tr class="total-row">
                    <td colspan="3"><b>TOTALS</b></td>
                    <td style="text-align:center;">
                        <b style="color:#2e7d32;">{present_days} Pres</b> / 
                        <b style="color:#d32f2f;">{absent_days} Abs</b>
                    </td>
                    <td style="text-align:right;"><b>{exact_ot:,.2f}h</b></td>
                    <td style="text-align:right;"><b>{exact_bonus:,.2f}</b></td>
                </tr>
            </tfoot>
        </table>
        <div class="payroll-grid">
            <div class="payroll-panel">
                <div class="panel-title">EARNINGS</div>
                <div class="pay-row"><span>Basic Salary</span><span>{basic:,.2f}</span></div>
                <div class="pay-row"><span>Housing Allowance</span><span>{housing:,.2f}</span></div>
                <div class="pay-row"><span>Transport Allowance</span><span>{transport:,.2f}</span></div>
                <div class="pay-row"><span>Other Allowances</span><span>{other:,.2f}</span></div>
                <div class="pay-row total"><span>TOTAL SALARY</span><span>{total_salary:,.2f}</span></div>
                <div style="height:8px;"></div>
                <div class="pay-row"><span>Total OT ({exact_ot:,.2f}h @ {hourly_rate:,.2f})</span><span class="highlight-green">{total_ot_amount:,.2f}</span></div>
                <div class="pay-row"><span>Bonus</span><span class="highlight-green">{exact_bonus:,.2f}</span></div>
                <div class="pay-row total"><span>GROSS PAY</span><span class="highlight-green">{gross:,.2f}</span></div>
            </div>
            <div class="payroll-panel">
                <div class="panel-title">DEDUCTIONS & NET PAY</div>
                <div class="pay-row deduction">
                    <span>Absence Deduction ({absent_days} days @ {emp.daily_rate:,.2f})</span>
                    <span>({absence_ded:,.2f})</span>
                </div>
                <div class="pay-row deduction"><span>Salary Advance</span><span>({advance:,.2f})</span></div>
                <div class="pay-row deduction"><span>Other Deductions</span><span>({other_ded:,.2f})</span></div>
                <div class="pay-row total deduction"><span>TOTAL DEDUCTIONS</span><span>({total_deductions:,.2f})</span></div>
                <div style="height:15px;"></div>
                <div class="pay-row net"><span>NET SALARY</span><span>{net:,.2f}</span></div>
                <div style="height:8px;"></div>
                <div class="pay-row" style="font-size:9px; color:#666;">
                    <span>Days Present: {present_days}</span>
                    <span>Days Absent: {absent_days}</span>
                </div>
            </div>
        </div>
        <div class="summary-box">
            <b>Total OT Hours:</b> {exact_ot:,.2f}h &nbsp;|&nbsp;
            <b>Total OT Amount:</b> {total_ot_amount:,.2f} &nbsp;|&nbsp;
            <b>Total Bonus:</b> {exact_bonus:,.2f} &nbsp;|&nbsp;
            <b>Total Extra:</b> {total_extra:,.2f} &nbsp;|&nbsp;
            <b>Net Salary:</b> {net:,.2f}
        </div>
        <div class="signature-grid">
            <div class="signature-block"><div class="signature-line">Worker Signature</div></div>
            <div class="signature-block"><div class="signature-line">Site Engineer</div></div>
            <div class="signature-block"><div class="signature-line">HR Manager</div></div>
        </div>
        <script>window.onload = function() {{ window.print(); }}</script>
    </body></html>"""
            return HttpResponse(html)

    def labor_cost_report(self, request):
            from datetime import datetime
            company = self.get_active_company(request)
            month_str = request.GET.get("month")
            if month_str:
                try:
                    month = datetime.strptime(month_str, "%Y-%m").date().replace(day=1)
                except ValueError:
                    month = date.today().replace(day=1)
            else:
                month = date.today().replace(day=1)

            last_day = calendar.monthrange(month.year, month.month)[1]
            month_end = month.replace(day=last_day)

            logo_url = company.logo.url if company and company.logo else ''

            records = PayrollRecord.objects.filter(
                month=month, employee__company=company
            ).select_related('employee').prefetch_related('cost_centers', 'cost_centers__project')

            projects = {}
            for rec in records:
                emp = rec.employee
                cc_projects = set()

                for cc in rec.cost_centers.all():
                    proj = cc.project
                    cc_projects.add(proj.id)
                    if proj not in projects:
                        projects[proj] = []

                    cc_start = max(cc.from_date, month)
                    cc_end = min(cc.to_date, month_end)
                    days_in_month = (cc_end - cc_start).days + 1 if cc_start <= cc_end else 0

                    ot_amount = money(cc.overtime_hours * emp.hourly_rate_ot) if emp.hourly_rate_ot > 0 else Decimal(
                        "0")

                    projects[proj].append({
                        'employee': emp,
                        'days': days_in_month,
                        'daily_rate': rec.daily_rate,
                        'daily_cost': rec.daily_cost,
                        'salary_cost': cc.prorated_salary,
                        'admin_cost': cc.prorated_admin_cost,
                        'ot_hours': cc.overtime_hours,
                        'ot_amount': ot_amount,
                        'bonus': cc.bonus,
                        'total_cost': money(cc.prorated_salary + cc.prorated_admin_cost + ot_amount + cc.bonus),
                        'record': rec,
                    })

                if emp.project and emp.project.id not in cc_projects:
                    proj = emp.project
                    if proj not in projects:
                        projects[proj] = []

                    days_in_month = last_day - rec.days_absent
                    if days_in_month > 0:
                        daily_rate = rec.daily_rate
                        daily_cost = rec.daily_cost
                        salary_cost = money(daily_rate * days_in_month)
                        admin_cost = money((emp.monthly_admin_cost / Decimal("30")) * days_in_month)

                        projects[proj].append({
                            'employee': emp,
                            'days': days_in_month,
                            'daily_rate': daily_rate,
                            'daily_cost': daily_cost,
                            'salary_cost': salary_cost,
                            'admin_cost': admin_cost,
                            'ot_hours': Decimal("0"),
                            'ot_amount': Decimal("0"),
                            'bonus': Decimal("0"),
                            'total_cost': money(salary_cost + admin_cost),
                            'record': rec,
                        })

            sorted_projects = sorted(projects.items(), key=lambda x: x[0].project_id_code)

            project_cards = ""
            grand_total = Decimal("0")
            grand_salary = Decimal("0")
            grand_admin = Decimal("0")
            grand_ot = Decimal("0")
            grand_bonus = Decimal("0")

            for proj, employees in sorted_projects:
                if not employees:
                    continue

                rows = ""
                proj_total = Decimal("0")
                proj_salary = Decimal("0")
                proj_admin = Decimal("0")
                proj_ot = Decimal("0")
                proj_bonus = Decimal("0")

                for entry in employees:
                    rows += f"""
                    <tr>
                        <td style="text-align:center;"><b>{entry['employee'].employee_id}</b></td>
                        <td>{entry['employee'].name}</td>
                        <td style="text-align:center;">{entry['days']}</td>
                        <td class='num'>{entry['daily_rate']:,.2f}</td>
                        <td class='num'>{entry['daily_cost']:,.2f}</td>
                        <td class='num'>{entry['salary_cost']:,.2f}</td>
                        <td class='num'>{entry['admin_cost']:,.2f}</td>
                        <td class='num'>{entry['ot_hours']:,.2f}</td>
                        <td class='num'>{entry['ot_amount']:,.2f}</td>
                        <td class='num'>{entry['bonus']:,.2f}</td>
                        <td class='num' style="font-weight:bold; color:#000080;">{entry['total_cost']:,.2f}</td>
                    </tr>
                    """
                    proj_total += entry['total_cost']
                    proj_salary += entry['salary_cost']
                    proj_admin += entry['admin_cost']
                    proj_ot += entry['ot_amount']
                    proj_bonus += entry['bonus']

                grand_total += proj_total
                grand_salary += proj_salary
                grand_admin += proj_admin
                grand_ot += proj_ot
                grand_bonus += proj_bonus

                project_cards += f"""
                <div class="project-card" style="margin-bottom:20px; border:1px solid #ccc; border-radius:8px; overflow:hidden; page-break-inside:avoid;">
                    <div style="background:#000080; color:white; padding:8px 12px; font-size:11px; font-weight:bold;">
                        {proj.project_id_code} — {proj.project_name}
                        <span style="float:right; font-weight:normal;">{len(employees)} worker(s)</span>
                    </div>
                    <table class="report-table" style="margin:0;">
                        <thead>
                            <tr style="background:#e8e8e8;">
                                <th style="width:10%;">Emp ID</th>
                                <th style="width:18%;">Name</th>
                                <th style="width:7%;">Days</th>
                                <th class='num' style="width:9%;">Daily Rate</th>
                                <th class='num' style="width:9%;">Daily Cost</th>
                                <th class='num' style="width:10%;">Salary</th>
                                <th class='num' style="width:10%;">Admin</th>
                                <th class='num' style="width:7%;">OT Hrs</th>
                                <th class='num' style="width:9%;">OT Amt</th>
                                <th class='num' style="width:8%;">Bonus</th>
                                <th class='num' style="width:10%;">Total</th>
                            </tr>
                        </thead>
                        <tbody>{rows}</tbody>
                        <tfoot>
                            <tr style="background:#e3f2fd; font-weight:bold; border-top:2px solid #333;">
                                <td colspan="5"><b>PROJECT TOTAL</b></td>
                                <td class='num'>{proj_salary:,.2f}</td>
                                <td class='num'>{proj_admin:,.2f}</td>
                                <td></td>
                                <td class='num'>{proj_ot:,.2f}</td>
                                <td class='num'>{proj_bonus:,.2f}</td>
                                <td class='num' style="color:#000080; font-size:11px;">{proj_total:,.2f}</td>
                            </tr>
                        </tfoot>
                    </table>
                </div>
                """

            month_options = ""
            for i in range(0, 12):
                m = (date.today().replace(day=1) - timedelta(days=i * 30)).replace(day=1)
                selected = "selected" if m == month else ""
                month_options += f'<option value="{m.strftime("%Y-%m")}" {selected}>{m.strftime("%B %Y")}</option>'

            html = f"""<!DOCTYPE html>
    <html><head><meta charset="UTF-8">
    <style>
        @page {{ size: A4 landscape; margin: 10mm; }}
        * {{ box-sizing: border-box; margin:0; padding:0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
        body {{ font-family: "Segoe UI", Arial, sans-serif; font-size: 9px; color: #222; padding: 10px; }}
        .logo-bar {{ text-align: right; margin-bottom: 6px; }}
        .logo-bar img {{ max-height: 120px; max-width: 240px; object-fit: contain; }}
        .report-title {{ font-size: 20px; font-weight: bold; text-align: center; color: #000080; margin-bottom: 4px; }}
        .report-subtitle {{ font-size: 12px; text-align: center; color: #666; margin-bottom: 12px; }}
        .meta-box {{ background: #f5f5f5; padding: 10px 15px; border-radius: 6px; margin-bottom: 15px; line-height: 1.5; font-size: 10px; display:flex; justify-content:space-between; align-items:center; }}
        .report-table {{ width: 100%; border-collapse: collapse; font-size: 8.5px; }}
        .report-table th {{ background: #e8e8e8; border: 1px solid #999; padding: 5px 4px; font-weight: bold; text-align: center; }}
        .report-table td {{ border: 1px solid #ccc; padding: 4px 5px; vertical-align: middle; }}
        .report-table .num {{ text-align: right; white-space: nowrap; }}
        .report-table tr:nth-child(even) {{ background: #fafafa; }}
        .grand-box {{ margin-top: 20px; padding: 15px; background: #000080; color: white; text-align: center; font-size: 16px; border-radius: 8px; }}
        .grand-grid {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 15px; margin-top: 8px; font-size: 11px; }}
        .grand-item {{ text-align: center; }}
        .grand-label {{ font-size: 8px; text-transform: uppercase; opacity: 0.9; margin-bottom: 4px; }}
        .month-form {{ display: flex; gap: 10px; align-items: center; }}
        .month-form select {{ padding: 6px 10px; border-radius: 4px; border: 1px solid #ccc; font-size: 12px; }}
        .month-form button {{ padding: 6px 16px; background: #000080; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: bold; }}
        @media print {{ .no-print {{ display: none; }} }}
    </style></head><body>
        {self._logo_bar(logo_url)}
        <div class="report-title">MONTHLY PROJECT LABOR COST REPORT</div>
        <div class="report-subtitle">Worker Cost Breakdown by Project</div>
        <div class="meta-box">
            <div>
                <b>Company:</b> {company.company_name if company else 'N/A'} &nbsp;|&nbsp;
                <b>Month:</b> {month.strftime('%B %Y')} &nbsp;|&nbsp;
                <b>Projects:</b> {len(sorted_projects)} &nbsp;|&nbsp;
                <b>Generated:</b> {date.today().strftime('%d-%b-%Y')}
            </div>
            <div class="month-form no-print">
                <form method="get">
                    <select name="month">{month_options}</select>
                    <button type="submit">View Report</button>
                </form>
            </div>
        </div>
        {project_cards if sorted_projects else '<div style="text-align:center; padding:40px; color:#999; font-size:14px;">No payroll records found for this month.</div>'}
        <div class="grand-box">
            <div style="font-size:13px; margin-bottom:8px; opacity:0.9;">GRAND TOTAL ACROSS ALL PROJECTS</div>
            <div class="grand-grid">
                <div class="grand-item"><div class="grand-label">Total Salary Cost</div><div style="font-weight:bold;">{grand_salary:,.2f}</div></div>
                <div class="grand-item"><div class="grand-label">Total Admin Cost</div><div style="font-weight:bold;">{grand_admin:,.2f}</div></div>
                <div class="grand-item"><div class="grand-label">Total OT Amount</div><div style="font-weight:bold;">{grand_ot:,.2f}</div></div>
                <div class="grand-item"><div class="grand-label">Total Bonus</div><div style="font-weight:bold;">{grand_bonus:,.2f}</div></div>
                <div class="grand-item"><div class="grand-label">Grand Total Cost</div><div style="font-weight:bold; font-size:14px;">{grand_total:,.2f}</div></div>
            </div>
        </div>
        <script>window.onload = function() {{ setTimeout(function() {{ window.print(); }}, 500); }}</script>
    </body></html>"""
            return HttpResponse(html)

    def _logo_bar(self, logo_url):
            if logo_url:
                return f'<div style="text-align:right; margin-bottom:6px;"><img src="{logo_url}" alt="Logo" style="max-height:120px; max-width:240px; object-fit:contain;"></div>'
            return ''
    def _payroll_report_wrapper(self, request, title, headers, rows, totals, payment_method):
            company = self.get_active_company(request)
            logo_url = company.logo.url if company and company.logo else ''
            logo_bar_html = f'<div style="text-align:right; margin-bottom:6px;"><img src="{logo_url}" alt="Logo" style="max-height:120px; max-width:240px; object-fit:contain;"></div>' if logo_url else ''

            header_cells = "".join(f"<th>{h}</th>" for h in headers)
            total_row = ""
            if "basic" in totals:
                total_row = f"""<tr class='total-row'>
                    <td colspan='2'><b>TOTAL</b></td>
                    <td class='num'>{totals['basic']:,.2f}</td>
                    <td class='num'>{totals['housing']:,.2f}</td>
                    <td class='num'>{totals['transport']:,.2f}</td>
                    <td class='num'>{totals['other']:,.2f}</td>
                    <td class='num'><b>{totals['total']:,.2f}</b></td>
                    <td class='num'>({totals['absence']:,.2f})</td>
                    <td class='num'>({totals['advance']:,.2f})</td>
                    <td class='num'>({totals['other_ded']:,.2f})</td>
                    <td class='num'><b>{totals['net']:,.2f}</b></td>
                </tr>"""
            elif "ot" in totals:
                total_row = f"""<tr class='total-row'>
                    <td colspan='2'><b>TOTAL</b></td>
                    <td class='num'>{totals['total']:,.2f}</td>
                    <td></td>
                    <td class='num'>{totals['ot']:,.2f}</td>
                    <td class='num'>{totals.get('bonus', Decimal('0')):,.2f}</td>
                    <td class='num'>({totals['absence']:,.2f})</td>
                    <td class='num'>({totals['advance']:,.2f})</td>
                    <td class='num'>({totals['other_ded']:,.2f})</td>
                    <td class='num'><b>{totals['net']:,.2f}</b></td>
                </tr>"""
            else:
                total_row = f"""<tr class='total-row'>
                    <td colspan='4'><b>TOTAL</b></td>
                    <td class='num'><b>{totals['net']:,.2f}</b></td>
                    <td></td>
                </tr>"""

            return f"""<!DOCTYPE html>
    <html><head><meta charset="UTF-8">
    <style>
        @page {{ size: A4 portrait; margin: 10mm; }}
        * {{ box-sizing: border-box; margin:0; padding:0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
        body {{ font-family: "Segoe UI", Arial, sans-serif; font-size: 10px; padding: 10px; }}
        .report-title {{ font-size: 18px; font-weight: bold; text-align: center; color: #000080; margin-bottom: 4px; }}
        .report-sub {{ font-size: 12px; text-align: center; color: #666; margin-bottom: 15px; }}
        .meta-box {{ background: #f5f5f5; padding: 10px 15px; border-radius: 6px; margin-bottom: 20px; line-height: 1.6; }}
        table {{ width: 100%; border-collapse: collapse; font-size: 9px; margin-top: 6px; }}
        th {{ background: #e8e8e8; border: 1px solid #999; padding: 6px; text-align: left; font-weight: bold; }}
        td {{ border: 1px solid #ccc; padding: 5px; }}
        .num {{ text-align: right; white-space: nowrap; }}
        tr:nth-child(even) {{ background: #fafafa; }}
        .total-row td {{ background: #e3f2fd; font-weight: bold; border-top: 2px solid #333; }}
    </style></head><body>
        {logo_bar_html}
        <div class="report-title">{title}</div>
        <div class="report-sub">Payment Method: {payment_method} | Generated: {date.today().strftime('%d-%b-%Y')}</div>
        <div class="meta-box">
            <b>Report Type:</b> {title}<br>
            <b>Date:</b> {date.today().strftime('%d-%b-%Y')}
        </div>
        <table>
            <thead><tr>{header_cells}</tr></thead>
            <tbody>{rows}</tbody>
            <tfoot>{total_row}</tfoot>
        </table>
        <script>window.onload = function() {{ window.print(); }}</script>
    </body></html>"""

    def staff_report(self, request):
        company = self.get_active_company(request)
        month = request.GET.get("month")
        qs = PayrollRecord.objects.filter(
            employee__employee_type="Staff",
            employee__payment_type="Bank",
            employee__company=company
        ).select_related("employee").order_by("-month")
        if month:
            try:
                # Convert YYYY-MM to first day of that month as a date object
                month_date = datetime.strptime(month, "%Y-%m").date().replace(day=1)
                qs = qs.filter(month=month_date)
            except ValueError:
                pass  # Invalid format, ignore filter

        rows = ""
        totals = {"basic": Decimal("0"), "housing": Decimal("0"), "transport": Decimal("0"),
                  "other": Decimal("0"), "total": Decimal("0"), "absence": Decimal("0"),
                  "advance": Decimal("0"), "other_ded": Decimal("0"), "net": Decimal("0")}

        for rec in qs:
            rows += f"""<<tr>
                <td>{rec.employee.employee_id}</td>
                <td>{rec.employee.name}</td>
                <td class='num'>{rec.basic_salary_snap:,.2f}</td>
                <td class='num'>{rec.housing_allowance_snap:,.2f}</td>
                <td class='num'>{rec.transport_allowance_snap:,.2f}</td>
                <td class='num'>{rec.other_allowances_snap:,.2f}</td>
                <td class='num'><b>{rec.total_salary_snap:,.2f}</b></td>
                <td class='num'>({rec.absence_deduction_snap:,.2f})</td>
                <td class='num'>({rec.salary_advance:,.2f})</td>
                <td class='num'>({rec.other_deduction:,.2f})</td>
                <td class='num'><b>{rec.net_salary_snap:,.2f}</b></td>
            </tr>"""
            totals["basic"] += rec.basic_salary_snap
            totals["housing"] += rec.housing_allowance_snap
            totals["transport"] += rec.transport_allowance_snap
            totals["other"] += rec.other_allowances_snap
            totals["total"] += rec.total_salary_snap
            totals["absence"] += rec.absence_deduction_snap
            totals["advance"] += rec.salary_advance
            totals["other_ded"] += rec.other_deduction
            totals["net"] += rec.net_salary_snap

        html = self._payroll_report_wrapper(
            request,
            "OFFICE STAFF PAYROLL REPORT",
            ["Emp ID", "Name", "Basic", "Housing", "Transport", "Other", "Total", "Absence", "Advance",
             "Other Ded.", "Net"],
            rows, totals, "Bank Transfer"
        )
        return HttpResponse(html)

    def wps_report(self, request):
        company = self.get_active_company(request)
        month = request.GET.get("month")
        qs = PayrollRecord.objects.filter(
            employee__employee_type="Site",
            employee__payment_type="WPS",
            employee__company=company
        ).select_related("employee").prefetch_related('cost_centers').order_by("-month")
        if month:
            try:
                month_date = datetime.strptime(month, "%Y-%m").date().replace(day=1)
                qs = qs.filter(month=month_date)
            except ValueError:
                pass

            rows = ""
            totals = {"total": Decimal("0"), "ot": Decimal("0"), "bonus": Decimal("0"),
                      "absence": Decimal("0"), "advance": Decimal("0"),
                      "other_ded": Decimal("0"), "net": Decimal("0")}

            for rec in qs:
                bonus = rec.cost_centers.aggregate(total=Sum('bonus'))['total'] or Decimal("0")
                rows += f"""<tr>
                    <td>{rec.employee.employee_id}</td>
                    <td>{rec.employee.name}</td>
                    <td class='num'>{rec.total_salary_snap:,.2f}</td>
                    <td class='num'>{rec.overtime_hours}h</td>
                    <td class='num'>{rec.overtime_amount_snap:,.2f}</td>
                    <td class='num'>{bonus:,.2f}</td>
                    <td class='num'>({rec.absence_deduction_snap:,.2f})</td>
                    <td class='num'>({rec.salary_advance:,.2f})</td>
                    <td class='num'>({rec.other_deduction:,.2f})</td>
                    <td class='num'><b>{rec.net_salary_snap:,.2f}</b></td>
                </tr>"""
                totals["total"] += rec.total_salary_snap
                totals["ot"] += rec.overtime_amount_snap
                totals["bonus"] += bonus
                totals["absence"] += rec.absence_deduction_snap
                totals["advance"] += rec.salary_advance
                totals["other_ded"] += rec.other_deduction
                totals["net"] += rec.net_salary_snap

            html = self._payroll_report_wrapper(
                request,
                "SITE WORKERS PAYROLL REPORT (WPS)",
                ["Emp ID", "Name", "Total Salary", "OT Hrs", "OT Amt", "Bonus", "Absence", "Advance", "Other Ded.",
                 "Net"],
                rows, totals, "WPS Agency"
            )
            return HttpResponse(html)

    def cash_report(self, request):
        company = self.get_active_company(request)
        month = request.GET.get("month")
        qs = PayrollRecord.objects.filter(
            employee__payment_type="Cash",
            employee__company=company
        ).select_related("employee").prefetch_related('cost_centers').order_by("-month")
        if month:
            try:
                month_date = datetime.strptime(month, "%Y-%m").date().replace(day=1)
                qs = qs.filter(month=month_date)
            except ValueError:
                pass

            rows = ""
            totals = {"bonus": Decimal("0"), "net": Decimal("0")}
            for rec in qs:
                bonus = rec.cost_centers.aggregate(total=Sum('bonus'))['total'] or Decimal("0")
                totals["bonus"] += bonus
                totals["net"] += rec.net_salary_snap
                rows += f"""<tr>
                    <td>{rec.employee.employee_id}</td>
                    <td>{rec.employee.name}</td>
                    <td>{rec.employee.get_employee_type_display()}</td>
                    <td class='num'>{bonus:,.2f}</td>
                    <td class='num'>{rec.net_salary_snap:,.2f}</td>
                    <td style='width:120px; border-bottom:1px solid #333;'></td>
                </tr>"""

            html = self._payroll_report_wrapper(
                request,
                "CASH PAYROLL REPORT",
                ["Emp ID", "Name", "Type", "Bonus", "Net Amount", "Signature"],
                rows, totals, "Cash"
            )
            return HttpResponse(html)

    def allocate_view(self, request):
            today = date.today()
            first_day = today.replace(day=1)
            last_month = (first_day - timedelta(days=1)).replace(day=1)
            company = self.get_active_company(request)
            unallocated = PayrollRecord.objects.filter(
                month=last_month, is_allocated=False, employee__company=company
            ).select_related("employee")

            if request.method == "POST":
                action = request.POST.get("action")
                if action == "yes":
                    from .payroll import allocate_payroll
                    done = 0
                    for rec in unallocated:
                        if allocate_payroll(rec):
                            done += 1
                    messages.success(request, f"{done} payroll record(s) allocated successfully.")
                    return redirect("..")
                else:
                    messages.info(request, "Allocation postponed. You will be reminded again.")
                    return redirect("..")

            rows = ""
            total_net = Decimal("0")
            for rec in unallocated:
                total_net += rec.net_salary_snap
                rows += f"""<tr>
                    <td>{rec.employee.employee_id}</td>
                    <td>{rec.employee.name}</td>
                    <td>{rec.employee.get_employee_type_display()}</td>
                    <td>{rec.employee.get_payment_type_display()}</td>
                    <td class='num'>{rec.net_salary_snap:,.2f}</td>
                </tr>"""

            html = f"""<!DOCTYPE html>
    <html><head><meta charset="UTF-8">
    <style>
        body {{ font-family: "Segoe UI", Arial, sans-serif; font-size: 12px; padding: 40px; background: #f5f5f5; }}
        .box {{ max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
        h2 {{ color: #000080; margin-bottom: 10px; }}
        .subtitle {{ color: #666; margin-bottom: 25px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; font-size: 11px; }}
        th {{ background: #e8e8e8; border: 1px solid #999; padding: 8px; text-align: left; }}
        td {{ border: 1px solid #ccc; padding: 8px; }}
        .num {{ text-align: right; }}
        .actions {{ margin-top: 25px; display: flex; gap: 15px; }}
        .btn {{ padding: 12px 30px; border: none; border-radius: 6px; font-size: 14px; cursor: pointer; font-weight: bold; }}
        .btn-yes {{ background: #2e7d32; color: white; }}
        .btn-no {{ background: #ed6c02; color: white; }}
        .total-row td {{ font-weight: bold; background: #e3f2fd; border-top: 2px solid #333; }}
    </style></head><body>
        <div class="box">
            <h2>Monthly Payroll Allocation</h2>
            <div class="subtitle">Month: {last_month.strftime('%B %Y')} | Unallocated Records: {unallocated.count()}</div>
            <table>
                <thead>
                    <tr><th>Emp ID</th><th>Name</th><th>Type</th><th>Payment</th><th class='num'>Net Salary</th></tr>
                </thead>
                <tbody>{rows}</tbody>
                <tfoot>
                    <tr class="total-row">
                        <td colspan="4"><b>TOTAL NET TO ALLOCATE</b></td>
                        <td class='num'><b>{total_net:,.2f}</b></td>
                    </tr>
                </tfoot>
            </table>
            <div style="background:#fff3cd; padding:12px; border-radius:6px; margin-bottom:20px; border-left:4px solid #ed6c02;">
                <b>Head Office employees</b> will be distributed across projects based on monthly work-done percentages.<br>
                <b>Project employees</b> will be allocated directly to their project's BOQ items.
            </div>
            <form method="post">
                <div class="actions">
                    <button type="submit" name="action" value="yes" class="btn btn-yes">YES - Allocate Now</button>
                    <button type="submit" name="action" value="no" class="btn btn-no">NO - Remind Me Tomorrow</button>
                </div>
            </form>
        </div>
    </body></html>"""
            return HttpResponse(html)

# =============================================================================
# PRICING ADMIN
# =============================================================================

class PricingBOQItemInline(admin.TabularInline):
    model = PricingBOQItem
    extra = 1
    fields = [
        "item_number", "description", "unit", "estimated_quantity",
        "reference_boq_item", "historical_rate", "historical_cost", "proposed_rate", "proposed_total"
    ]
    readonly_fields = ["historical_rate", "historical_cost", "proposed_total"]

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        company = CompanyProfile.get_active(request)
        if company and db_field.name == 'reference_boq_item':
            kwargs['queryset'] = BOQItem.objects.filter(project__company=company)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)



# =============================================================================
# SUPPLIER ADMIN
# =============================================================================

@admin.register(Supplier)
class SupplierAdmin(ProfessionalReportMixin, CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'company'
    list_display = [
        "name", "category", "contact_person", "payment_terms",
        "fmt_total_payable", "fmt_credit_limit", "is_active", "aging_button"
    ]
    list_filter = ["company", "category", "is_active", "payment_terms"]
    search_fields = ["name", "contact_person", "trn_number", "email"]
    fieldsets = (
        ("Basic Information", {
            "fields": ("company", "name", "category", "contact_person", "is_active")
        }),
        ("Contact Details", {
            "fields": ("phone", "email", "address", "trn_number")
        }),
        ("Payment Terms", {
            "fields": ("payment_terms", "credit_limit")
        }),
        ("Banking Details", {
            "fields": ("bank_name", "account_name", "account_number", "iban", "swift_code"),
            "classes": ("collapse",)
        }),
        ("Notes", {
            "fields": ("notes",),
            "classes": ("collapse",)
        }),
        ("Linked Expense", {
            "fields": ("linked_expense_display",),
            "description": "Auto-generated expense record linked to this invoice",
            "classes": ("collapse",)
        }),
    )

    def fmt_total_payable(self, obj):
        total = obj.total_payable
        if total > 0:
            return mark_safe(f'<div style="text-align:right;font-weight:bold;color:#c62828;">{total:,.2f}</div>')
        return mark_safe('<span style="color:#999;">—</span>')
    fmt_total_payable.short_description = "Total Payable"

    def fmt_credit_limit(self, obj):
        if obj.credit_limit > 0:
            usage = (obj.total_payable / obj.credit_limit * 100) if obj.credit_limit > 0 else 0
            color = "#d32f2f" if usage > 90 else "#ed6c02" if usage > 70 else "#2e7d32"
            return mark_safe(
                f'<div style="text-align:right;font-weight:bold;color:{color};">'
                f'{obj.credit_limit:,.2f}</div>'
                f'<div style="font-size:8px;color:#666;">{usage:.0f}% used</div>'
            )
        return mark_safe('<span style="color:#999;">—</span>')
    fmt_credit_limit.short_description = "Credit Limit"

    def aging_button(self, obj):
        url = reverse('admin:supplier_aging', args=[obj.pk])
        return format_html(
            '<a class="button" href="{}" target="_blank" style="background:#1a237e; color:white; padding:2px 8px; '
            'border-radius:4px; font-size:10px; text-decoration:none; font-weight:600;">📊 Aging</a>',
            url
        )
    aging_button.short_description = "Report"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('<int:pk>/aging/', self.admin_site.admin_view(self.aging_view), name='supplier_aging'),
        ]
        return custom + urls

    def aging_view(self, request, pk):
        """Supplier Aging Report"""
        company = self.get_active_company(request)
        supplier = self.get_object_or_404_scoped(request, Supplier, pk=pk)
        logo_url = company.logo.url if company and company.logo else ''

        invoices = SupplierInvoice.objects.filter(
            supplier=supplier
        ).exclude(status='Paid').exclude(status='Cancelled')

        if company:
            invoices = invoices.filter(company=company)

        # Bucket the invoices
        buckets = {
            'Current': [],
            '1-30 Days': [],
            '31-60 Days': [],
            '61-90 Days': [],
            '90+ Days': [],
        }
        bucket_totals = {k: Decimal("0") for k in buckets.keys()}

        for inv in invoices:
            bucket = inv.aging_bucket
            if bucket in buckets:
                buckets[bucket].append(inv)
                bucket_totals[bucket] += inv.balance_due

        # Build bucket sections
        bucket_html = ""
        colors = {
            'Current': ('#2e7d32', '#e8f5e9'),
            '1-30 Days': ('#f57c00', '#fff3e0'),
            '31-60 Days': ('#ed6c02', '#ffe0b2'),
            '61-90 Days': ('#d32f2f', '#ffcdd2'),
            '90+ Days': ('#c62828', '#ffebee'),
        }

        for bucket_name, inv_list in buckets.items():
            if not inv_list:
                continue
            color, bg = colors[bucket_name]
            rows = ""
            for inv in inv_list:
                rows += f"""
                <tr>
                    <td class="center">{inv.supplier_inv_number}</td>
                    <td class="center">{inv.invoice_date.strftime('%d-%b-%Y')}</td>
                    <td class="center">{inv.due_date.strftime('%d-%b-%Y')}</td>
                    <td class="text">{inv.description[:60]}</td>
                    <td class="num">{inv.total_amount:,.2f}</td>
                    <td class="num text-success">{inv.paid_amount:,.2f}</td>
                    <td class="num font-bold text-danger">{inv.balance_due:,.2f}</td>
                    <td class="center"><span class="badge badge-warning">{inv.days_overdue} days</span></td>
                </tr>
                """

            bucket_html += f"""
            <div class="card" style="margin-bottom: 16px; border-left: 4px solid {color};">
                <div class="card-header" style="color: {color}; border-color: {color};">
                    <span class="icon">⏱️</span> {bucket_name} 
                    <span class="badge badge-danger" style="margin-left: 10px;">{len(inv_list)} invoices</span>
                    <span style="float: right; color: {color}; font-weight: bold;">AED {bucket_totals[bucket_name]:,.2f}</span>
                </div>
                <table class="data-table" style="margin: 0;">
                    <thead>
                        <tr>
                            <th>Inv #</th>
                            <th>Inv Date</th>
                            <th>Due Date</th>
                            <th>Description</th>
                            <th class="num">Total</th>
                            <th class="num">Paid</th>
                            <th class="num">Balance</th>
                            <th>Overdue</th>
                        </tr>
                    </thead>
                    <tbody>{rows}</tbody>
                    <tfoot>
                        <tr style="background: {bg}; font-weight: bold;">
                            <td colspan="4"><b>{bucket_name} TOTAL</b></td>
                            <td class="num">—</td>
                            <td class="num">—</td>
                            <td class="num text-danger"><b>{bucket_totals[bucket_name]:,.2f}</b></td>
                            <td></td>
                        </tr>
                    </tfoot>
                </table>
            </div>
            """

        grand_total = sum(bucket_totals.values())

        body = f"""
        {self._build_meta_grid({
            'Supplier': supplier.name,
            'Category': supplier.get_category_display(),
            'TRN': supplier.trn_number or 'N/A',
            'Payment Terms': f"{supplier.payment_terms} days",
            'Credit Limit': f"AED {supplier.credit_limit:,.2f}" if supplier.credit_limit > 0 else 'N/A',
            'Report Date': date.today().strftime('%d-%b-%Y'),
        })}

        <div class="card" style="background: linear-gradient(135deg, #1a237e 0%, #283593 100%); color: white; margin-bottom: 20px;">
            <div class="card-header" style="color: white; border-color: rgba(255,255,255,0.3);">
                <span class="icon">📊</span> SUPPLIER AGING SUMMARY
            </div>
            <div class="grid-5">
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 20px;">{bucket_totals['Current']:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">Current</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 20px;">{bucket_totals['1-30 Days']:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">1-30 Days</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 20px;">{bucket_totals['31-60 Days']:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">31-60 Days</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 20px;">{bucket_totals['61-90 Days']:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">61-90 Days</div>
                </div>
                <div class="metric-card" style="background: rgba(255,255,255,0.15); border: none; color: white;">
                    <div class="metric-value" style="color: white; font-size: 20px;">{bucket_totals['90+ Days']:,.2f}</div>
                    <div class="metric-label" style="color: rgba(255,255,255,0.9);">90+ Days</div>
                </div>
            </div>
            <div style="text-align: center; margin-top: 16px; padding-top: 16px; border-top: 1px solid rgba(255,255,255,0.2);">
                <div style="font-size: 11px; text-transform: uppercase; letter-spacing: 1px; opacity: 0.8;">TOTAL OUTSTANDING</div>
                <div style="font-size: 28px; font-weight: 700;">AED {grand_total:,.2f}</div>
            </div>
        </div>

        {bucket_html if bucket_html else '<div class="card" style="text-align:center; padding:40px; color:#666;"><p>No outstanding invoices for this supplier.</p></div>'}

        <div class="signature-grid" style="margin-top: 40px;">
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">Finance Director</div>
            </div>
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">Procurement Manager</div>
            </div>
            <div class="signature-block">
                <div style="font-size: 8px; color: #666; margin-top: 4px;">General Manager</div>
            </div>
        </div>
        """

        return HttpResponse(self._report_base_wrapper(
            "SUPPLIER AGING REPORT",
            f"Accounts Payable Aging — {supplier.name}",
            body,
            logo_url
        ))


# =============================================================================
# SUPPLIER INVOICE (AP) ADMIN
# =============================================================================

class SupplierPaymentInline(admin.TabularInline):
    model = SupplierPayment
    extra = 0
    fields = ["payment_date", "amount", "payment_method", "bank_reference", "reference_number"]


@admin.register(SupplierInvoice)
class SupplierInvoiceAdmin(CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'company'
    inlines = [SupplierPaymentInline]
    list_display = [
        "supplier_inv_number", "supplier", "project", "fmt_amount",
        "fmt_total", "fmt_balance", "status_badge", "due_date",
        "days_overdue_display", "expected_payment_date", "is_recurring",
        "linked_expense_count",
    ]

    list_filter = [
        "company", "status", "supplier__category", "is_recurring", "recurring_frequency",
        "invoice_date", "due_date", "supplier"
    ]
    search_fields = ["supplier_inv_number", "description", "supplier__name", "reference_number"]
    date_hierarchy = "due_date"
    autocomplete_fields = ["supplier", "project", "expense_category"]

    fieldsets = (
        ("Invoice Details", {
            "fields": (
                ("company", "supplier"),
                ("supplier_inv_number", "reference_number"),
                "description",
                ("project", "boq_item", "expense_category"),
            )
        }),
        ("Amounts", {
            "fields": (
                ("amount", "vat_percent"),
                ("retention_percent",),
            )
        }),
        ("Dates & Status", {
            "fields": (
                ("invoice_date", "due_date", "expected_payment_date"),
                "status",
            )
        }),
        ("Recurring", {
            "fields": ("is_recurring", "recurring_frequency", "parent_invoice"),
            "classes": ("collapse",)
        }),
        ("Payment Tracking", {
            "fields": (
                ("paid_amount", "actual_payment_date"),
            ),
            "classes": ("collapse",)
        }),
        ("Notes", {
            "fields": ("notes",),
            "classes": ("collapse",)
        }),
    )

    class Media:
        js = (
            'admin/js/vendor/jquery/jquery.min.js',
            'admin/js/jquery.init.js',
            'billing/js/supplier_invoice_boq_filter.js',
        )

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        company = self.get_active_company(request)

        if db_field.name == 'boq_item':
            # Get project from existing object
            project_id = None
            if hasattr(request, '_supplier_invoice_project_id'):
                project_id = request._supplier_invoice_project_id
            elif request.resolver_match and hasattr(request.resolver_match, 'kwargs'):
                object_id = request.resolver_match.kwargs.get('object_id')
                if object_id:
                    try:
                        obj = SupplierInvoice.objects.get(pk=object_id)
                        project_id = obj.project_id
                    except SupplierInvoice.DoesNotExist:
                        pass

            # Also check POST data
            if not project_id and request.method == 'POST':
                project_id = request.POST.get('project')

            if project_id:
                kwargs['queryset'] = BOQItem.objects.filter(project_id=project_id)
            else:
                kwargs['queryset'] = BOQItem.objects.none()

        elif db_field.name == 'project':
            if company:
                kwargs['queryset'] = Project.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
        elif db_field.name == 'expense_category':
            if company:
                kwargs['queryset'] = ExpenseCategory.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )
        elif db_field.name == 'supplier':
            if company:
                kwargs['queryset'] = Supplier.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if obj and obj.project_id:
            request._supplier_invoice_project_id = obj.project_id
        return form

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == 'boq_item':
            # Try to get project from request (either POST data or existing object)
            project_id = None

            # Check if we're editing an existing object
            if request.resolver_match and hasattr(request.resolver_match, 'kwargs'):
                object_id = request.resolver_match.kwargs.get('object_id')
                if object_id:
                    try:
                        obj = SupplierInvoice.objects.get(pk=object_id)
                        if obj.project:
                            kwargs['queryset'] = BOQItem.objects.filter(project=obj.project)
                        else:
                            kwargs['queryset'] = BOQItem.objects.none()
                    except SupplierInvoice.DoesNotExist:
                        kwargs['queryset'] = BOQItem.objects.none()
                else:
                    # Adding new — can't know project yet, return empty or all
                    kwargs['queryset'] = BOQItem.objects.none()
            else:
                kwargs['queryset'] = BOQItem.objects.none()

        elif db_field.name == 'project':
            # Keep your existing company scoping for project
            company = self.get_active_company(request)
            if company:
                kwargs['queryset'] = Project.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )

        elif db_field.name == 'expense_category':
            company = self.get_active_company(request)
            if company:
                kwargs['queryset'] = ExpenseCategory.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )

        elif db_field.name == 'supplier':
            company = self.get_active_company(request)
            if company:
                kwargs['queryset'] = Supplier.objects.filter(
                    Q(company=company) | Q(company__isnull=True)
                )

        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def linked_expense_display(self, obj):
        """Display linked expenses with clickable links."""
        expenses = obj.linked_expenses.all()
        if not expenses.exists():
            return mark_safe(
                '<span style="color:#999;">No linked expenses yet. Save the invoice to auto-generate.</span>')

        html = '<div style="display:flex; flex-direction:column; gap:4px;">'
        for exp in expenses:
            url = reverse('admin:billing_expense_change', args=[exp.pk])
            badge = '<span style="background:#e8f5e9; color:#2e7d32; padding:1px 4px; border-radius:3px; font-size:8px;">Auto</span>' if exp.is_auto_generated else ''
            html += (
                f'<div style="display:flex; justify-content:space-between; align-items:center; padding:4px 8px; '
                f'background:#f5f5f5; border-radius:4px;">'
                f'<a href="{url}" target="_blank" style="color:#1a237e; font-weight:bold; text-decoration:none;">'
                f'{exp.category.name} — AED {exp.amount:,.2f}</a>'
                f'<span>{badge}</span>'
                f'</div>'
            )
        html += '</div>'
        return mark_safe(html)

    linked_expense_display.short_description = "Linked Expenses"

    def linked_expense_count(self, obj):
        count = obj.linked_expenses.count()
        if count > 0:
            return mark_safe(f'<span style="color:#2e7d32; font-weight:bold;">{count}</span>')
        return mark_safe('<span style="color:#999;">—</span>')

    linked_expense_count.short_description = "Expenses"

    def fmt_amount(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{obj.amount:,.2f}</div>')
    fmt_amount.short_description = "Net Amount"

    def fmt_total(self, obj):
        return mark_safe(f'<div style="text-align:right;">{obj.total_amount:,.2f}</div>')
    fmt_total.short_description = "Total + VAT"

    def fmt_balance(self, obj):
        balance = obj.balance_due
        if balance > 0:
            return mark_safe(f'<div style="text-align:right;font-weight:bold;color:#c62828;">{balance:,.2f}</div>')
        return mark_safe('<span style="color:#2e7d32; font-weight:bold;">PAID</span>')
    fmt_balance.short_description = "Balance Due"

    def status_badge(self, obj):
        colors = {
            'Draft': 'badge-warning',
            'Approved': 'badge-info',
            'Scheduled': 'badge-primary',
            'Paid': 'badge-success',
            'Cancelled': 'badge-danger',
            'Disputed': 'badge-danger',
        }
        badge_class = colors.get(obj.status, 'badge-warning')
        return mark_safe(f'<span class="badge {badge_class}">{obj.status}</span>')
    status_badge.short_description = "Status"

    def days_overdue_display(self, obj):
        days = obj.days_overdue
        if obj.status == 'Paid':
            return mark_safe('<span style="color:#2e7d32; font-size:10px;">✓ Paid</span>')
        if days == 0:
            return mark_safe(f'<span style="color:#2e7d32; font-size:10px;">{days} days</span>')
        elif days <= 30:
            return mark_safe(f'<span style="color:#f57c00; font-size:10px; font-weight:bold;">{days} days</span>')
        else:
            return mark_safe(f'<span style="color:#c62828; font-size:10px; font-weight:bold;">{days} days ⚠️</span>')
    days_overdue_display.short_description = "Overdue"

    actions = ["mark_approved", "mark_scheduled", "mark_paid", "generate_next_recurring"]

    @admin.action(description="Mark selected as Approved")
    def mark_approved(self, request, queryset):
        queryset.update(status='Approved')
        self.message_user(request, f"{queryset.count()} invoice(s) marked as Approved.")

    @admin.action(description="Mark selected as Scheduled for Payment")
    def mark_scheduled(self, request, queryset):
        queryset.update(status='Scheduled')
        self.message_user(request, f"{queryset.count()} invoice(s) marked as Scheduled.")

    @admin.action(description="Mark selected as Paid (full amount)")
    def mark_paid(self, request, queryset):
        today = date.today()
        for inv in queryset:
            if inv.status != 'Paid':
                inv.paid_amount = inv.total_amount
                inv.status = 'Paid'
                inv.actual_payment_date = today
                inv.save()
        self.message_user(request, f"{queryset.count()} invoice(s) marked as Paid.")

    @admin.action(description="Generate next recurring invoice")
    def generate_next_recurring(self, request, queryset):
        created = 0
        for inv in queryset.filter(is_recurring=True):
            # Calculate next date based on frequency
            if inv.recurring_frequency == 'Monthly':
                from dateutil.relativedelta import relativedelta
                next_date = inv.invoice_date + relativedelta(months=1)
                next_due = inv.due_date + relativedelta(months=1)
            elif inv.recurring_frequency == 'Quarterly':
                from dateutil.relativedelta import relativedelta
                next_date = inv.invoice_date + relativedelta(months=3)
                next_due = inv.due_date + relativedelta(months=3)
            elif inv.recurring_frequency == 'Annual':
                from dateutil.relativedelta import relativedelta
                next_date = inv.invoice_date + relativedelta(years=1)
                next_due = inv.due_date + relativedelta(years=1)
            else:
                continue

            # Check if next instance already exists
            if not SupplierInvoice.objects.filter(
                parent_invoice=inv, invoice_date=next_date
            ).exists():
                SupplierInvoice.objects.create(
                    company=inv.company,
                    supplier=inv.supplier,
                    project=inv.project,
                    expense_category=inv.expense_category,
                    supplier_inv_number=f"{inv.supplier_inv_number}-R",
                    description=inv.description,
                    amount=inv.amount,
                    vat_percent=inv.vat_percent,
                    invoice_date=next_date,
                    due_date=next_due,
                    expected_payment_date=next_due,
                    status='Draft',
                    is_recurring=True,
                    recurring_frequency=inv.recurring_frequency,
                    parent_invoice=inv,
                    notes=f"Auto-generated recurring invoice from {inv.supplier_inv_number}"
                )
                created += 1

        self.message_user(request, f"{created} recurring invoice(s) created.")


@admin.register(SupplierPayment)
class SupplierPaymentAdmin(CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'supplier_invoice__company'
    list_display = ["supplier_invoice", "payment_date", "fmt_amount", "payment_method", "bank_reference"]
    list_filter = ["payment_method", "payment_date"]
    search_fields = ["supplier_invoice__supplier__name", "bank_reference", "reference_number"]
    date_hierarchy = "payment_date"
    autocomplete_fields = ["supplier_invoice"]

    def fmt_amount(self, obj):
        return mark_safe(f'<div style="text-align:right;font-weight:bold;color:#2e7d32;">{obj.amount:,.2f}</div>')
    fmt_amount.short_description = "Amount"

@admin.register(PricingProject)
class PricingProjectAdmin(CompanyScopedAdminMixin, admin.ModelAdmin):
    company_field_path = 'company'
    inlines = [PricingBOQItemInline]
    list_display = ["project_name", "client", "company", "created_date", "fmt_total"]
    list_filter = ["company"]
    search_fields = ["project_name", "client__name"]
    filter_horizontal = ["reference_projects"]

    def fmt_total(self, obj):
        total = sum(item.proposed_total for item in obj.boq_items.all())
        return mark_safe(f'<div style="text-align:right;font-weight:bold;">{total:,.2f}</div>')
    fmt_total.short_description = "Proposed Total"
